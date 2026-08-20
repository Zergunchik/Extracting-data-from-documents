import sys
import os
import re
import pdfplumber
from openpyxl import load_workbook, Workbook
from pathlib import Path
import builtins
import logging

# Настраиваем логирование
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

def should_stop():
    """Проверяет, запросил ли пользователь остановку через GUI."""
    check = getattr(builtins, '_gui_stop_check', None)
    return check() if check else False

OUTPUT_DIR = None  # Глобальная переменная для папки вывода

def find_ktc_pdf(search_folder=None):
    """
    Ищет PDF файл, содержащий в названии 'КТС' или 'КРУС-КТС'
    Если search_folder не указан, ищет в папке со скриптом
    """
    if search_folder is None:
        search_folder = os.path.dirname(os.path.abspath(__file__))
    
    if not os.path.exists(search_folder):
        return None
    
    for file in os.listdir(search_folder):
        if file.lower().endswith('.pdf') and ('ктс' in file.lower() or 'крус-ктс' in file.lower()):
            return os.path.join(search_folder, file)
    return None

def expand_position_range(pos_str):
    """
    Раскрывает диапазоны позиционных обозначений
    Примеры: KBT1...KBT3 -> ['KBT1', 'KBT2', 'KBT3']
             KL4.1, KL4.2 -> ['KL4.1', 'KL4.2']
    """
    # Обработка диапазона через многоточие
    if '...' in pos_str:
        match = re.match(r'([A-Z]+)(\d+)(?:\.\d+)?\.\.\.\1(\d+)', pos_str)
        if match:
            prefix = match.group(1)
            start = int(match.group(2))
            end = int(match.group(3))
            return [f"{prefix}{i}" for i in range(start, end+1)]
        
        # С точкой в номере (например, KL4.1...KL4.3)
        match = re.match(r'([A-Z]+)(\d+\.\d+)\.\.\.\1(\d+\.\d+)', pos_str)
        if match:
            prefix = match.group(1)
            start_parts = match.group(2).split('.')
            end_parts = match.group(3).split('.')
            if len(start_parts) == 2 and len(end_parts) == 2:
                start_main = int(start_parts[0])
                start_sub = int(start_parts[1])
                end_main = int(end_parts[0])
                end_sub = int(end_parts[1])
                results = []
                for main_num in range(start_main, end_main + 1):
                    sub_start = start_sub if main_num == start_main else 1
                    sub_end = end_sub if main_num == end_main else 99
                    for sub_num in range(sub_start, sub_end + 1):
                        results.append(f"{prefix}{main_num}.{sub_num}")
                return results
        return [pos_str]
    
    # Обработка перечисления через запятую
    if ',' in pos_str:
        return [p.strip() for p in pos_str.split(',')]
    
    return [pos_str]

def load_nominals_library(library_path=None):
    """
    Загружает библиотеку номинальных токов из Library_nominals.xlsx
    Возвращает словарь: {тип_автомата: номинальный_ток}
    """
    if library_path is None:
        # Ищем в папке со скриптом
        script_folder = os.path.dirname(os.path.abspath(__file__))
        library_path = os.path.join(script_folder, "Library_nominals.xlsx")
    
    if not os.path.exists(library_path):
        logger.info(f"⚠️ Библиотека номиналов не найдена: {library_path}")
        logger.info(f"   Номинальные токи будут определяться по шаблону 'XX A' в PDF")
        return {}
    
    try:
        wb = load_workbook(library_path, data_only=True)
        # Пробуем найти лист "Номинальные токи", если нет - берем активный
        if "Номинальные токи" in wb.sheetnames:
            sheet = wb["Номинальные токи"]
        else:
            sheet = wb.active
        
        nominals = {}
        for row in range(2, sheet.max_row + 1):  # со 2-й строки (после заголовков)
            device_type = sheet.cell(row=row, column=2).value  # столбец B
            nominal = sheet.cell(row=row, column=3).value     # столбец C
            
            if device_type and nominal:
                # Приводим тип к единому формату (удаляем лишние пробелы)
                device_type_clean = str(device_type).strip()
                # Преобразуем номинал в число (если это возможно)
                try:
                    nominal_value = int(float(nominal)) if isinstance(nominal, (int, float)) else int(float(str(nominal).strip()))
                    nominals[device_type_clean] = nominal_value
                except (ValueError, TypeError):
                    logger.info(f"   ⚠️ Некорректный номинал для {device_type_clean}: {nominal}")
                    nominals[device_type_clean] = nominal
        
        logger.info(f"✅ Загружено номиналов из библиотеки: {len(nominals)}")
        return nominals
    
    except Exception as e:
        logger.info(f"❌ Ошибка загрузки библиотеки номиналов: {e}")
        return {}

def match_device_type_with_library(device_type, nominals_library):
    """
    Сопоставляет тип автомата с библиотекой:
    1. Точное совпадение
    2. Совпадение по маске (если тип из библиотеки содержит подстроку)
    3. Частичное совпадение (игнорируя суффиксы)
    """
    if not nominals_library:
        return None
    
    device_clean = clean_device_type(device_type)
    
    # Точное совпадение
    if device_clean in nominals_library:
        return nominals_library[device_clean]
    
    # Поиск по частичному совпадению (например, OptiMat D250F-TM016 должен найти OptiMat D250F-TM*)
    for lib_type, nominal in nominals_library.items():
        # Если тип из библиотеки длиннее, проверяем вхождение
        if len(lib_type) >= len(device_clean):
            if lib_type.startswith(device_clean) or device_clean in lib_type:
                return nominal
        else:
            if device_clean.startswith(lib_type) or lib_type in device_clean:
                return nominal
    
    return None

def clean_device_type(device_type_raw):
    """
    Очищает тип устройства от лишних слов
    Удаляет: 'Автоматический выключатель', 'Выключатель автоматический', 'Автомат'
    """
    words_to_remove = [
        'Автоматический выключатель', 'Выключатель автоматический',
        'Автоматический', 'Выключатель', 'Автомат', 'автоматический',
        'выключатель', 'автомат', 'автоматического', 'выключателя',
        'автоматический выключатель', 'выключатель автоматический'
    ]
    
    cleaned = device_type_raw
    for word in words_to_remove:
        cleaned = cleaned.replace(word, '')
    
    # Удаляем лишние пробелы и символы
    cleaned = ' '.join(cleaned.split())
    cleaned = cleaned.strip()
    
    return cleaned if cleaned else device_type_raw

def extract_circuit_breakers_from_pdf(pdf_path, nominals_library=None):
    """
    Извлекает из PDF позиционные обозначения и типы автоматических выключателей
    Номинальный ток берётся из библиотеки nominals_library
    """
    breakers_data = {}
    errors_automats = []  # Список для хранения ошибок (автоматов не найденных в библиотеке)
    
    if not pdf_path or not os.path.exists(pdf_path):
        logger.info("⚠️ PDF с КТС не найден. Автоматы не будут определены.")
        return breakers_data, errors_automats
    
    logger.info(f"📄 Обработка PDF: {os.path.basename(pdf_path)}")
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                # Извлекаем текст с сохранением позиций (табличный режим)
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
                
                # Пробуем извлечь таблицы напрямую
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if row and len(row) >= 2:
                            # Объединяем ячейки строки в один текст для анализа
                            row_text = ' '.join([str(cell) if cell else '' for cell in row])
                            full_text += row_text + "\n"
    except Exception as e:
        logger.info(f"❌ Ошибка при открытии PDF: {e}")
        return breakers_data, errors_automats
    
    # Разбиваем на строки
    lines = full_text.split('\n')
    
    # Паттерн для поиска SF позиций
    sf_pattern = re.compile(r'\b(SF\d*)\b')
    
    # Перебираем строки, ищем SF
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        
        # Ищем SF в текущей строке
        sf_matches = sf_pattern.findall(line)
        
        if sf_matches:
            for sf_pos in sf_matches:
                if sf_pos == 'SF' or sf_pos == 'None':
                    continue
                
                # Ищем тип автомата - сначала в текущей строке после SF
                device_type = None
                nominal_current = "Не указан"
                error_info = None  # Для хранения информации об ошибке
                
                # Вариант 1: SF и тип в одной строке (через табуляцию или пробелы)
                # Ищем после SF слова, похожие на "Автоматический выключатель XXX"
                after_sf = line[line.find(sf_pos) + len(sf_pos):]
                
                # Паттерн для поиска "Автоматический выключатель" с последующим типом
                auto_match = re.search(r'Автоматический\s+выключатель\s+([А-Яа-я0-9\-_]+)', after_sf, re.IGNORECASE)
                if not auto_match:
                    auto_match = re.search(r'Выключатель\s+автоматический\s+([А-Яа-я0-9\-_]+)', after_sf, re.IGNORECASE)
                if not auto_match:
                    auto_match = re.search(r'Автомат\s+([А-Яа-я0-9\-_]+)', after_sf, re.IGNORECASE)
                
                if auto_match:
                    device_type = auto_match.group(1)
                
                # Если не нашли в текущей строке, смотрим следующие 3 строки
                if not device_type:
                    for offset in range(1, min(4, len(lines) - i)):
                        next_line = lines[i + offset].strip()
                        auto_match = re.search(r'Автоматический\s+выключатель\s+([А-Яа-я0-9\-_]+)', next_line, re.IGNORECASE)
                        if not auto_match:
                            auto_match = re.search(r'Выключатель\s+автоматический\s+([А-Яа-я0-9\-_]+)', next_line, re.IGNORECASE)
                        if not auto_match:
                            auto_match = re.search(r'Автомат\s+([А-Яа-я0-9\-_]+)', next_line, re.IGNORECASE)
                        if auto_match:
                            device_type = auto_match.group(1)
                            break
                
                # Если всё ещё не нашли, проверяем соседние строки на наличие табличных данных
                if not device_type:
                    # Проверяем строки вокруг на наличие "ПРКЗ" или подобных паттернов
                    for offset in range(-2, 3):
                        check_idx = i + offset
                        if 0 <= check_idx < len(lines):
                            check_line = lines[check_idx].strip()
                            # Ищем типовые обозначения автоматов (расширенный паттерн)
                            type_match = re.search(r'\b(ПРК[А-Яа-я0-9\-_]+|[A-Za-z]+[\-_\d]*[A-Za-z]*[\d]*)\b', check_line)
                            if type_match:
                                potential_type = type_match.group(1)
                                # Исключаем SF, TA и другие служебные
                                if not re.match(r'^(SF|TA|QF|K|HL|HLG|HLR|HLY|None)$', potential_type, re.IGNORECASE):
                                    # Дополнительная проверка: тип должен содержать буквы или цифры
                                    if len(potential_type) >= 3:
                                        device_type = potential_type
                                        break
                
                # Если нашли тип, определяем номинал
                if device_type:
                    # Очищаем тип от лишних слов
                    device_type_clean = clean_device_type(device_type)
                    
                    # Пытаемся найти номинал в библиотеке
                    if nominals_library:
                        nominal_from_lib = match_device_type_with_library(device_type_clean, nominals_library)
                        if nominal_from_lib is not None:
                            nominal_current = nominal_from_lib
                            logger.info(f"   Найден автомат: {sf_pos} → {device_type_clean} (номинал из библиотеки: {nominal_current}A)")
                        else:
                            # Fallback: ищем номинал по шаблону 'XX A' в PDF
                            for offset in range(-2, 3):
                                check_idx = i + offset
                                if 0 <= check_idx < len(lines):
                                    check_line = lines[check_idx].strip()
                                    nominal_match = re.search(r'(\d+(?:\.\d+)?)\s*A', check_line)
                                    if nominal_match:
                                        nominal_current = int(float(nominal_match.group(1)))
                                        logger.info(f"   Найден автомат: {sf_pos} → {device_type_clean} (номинал из PDF: {nominal_current}A) [тип {device_type_clean} не найден в библиотеке]")
                                        # Запоминаем ошибку
                                        error_info = {
                                            'position': sf_pos,
                                            'type': device_type_clean,
                                            'nominal_from_pdf': nominal_current,
                                            'context': 'nominal_found_in_pdf'
                                        }
                                        break
                            else:
                                nominal_current = f"Не указан (тип '{device_type_clean}' не найден в библиотеке)"
                                logger.info(f"   Найден автомат: {sf_pos} → {device_type_clean} (номинал не определён)")
                                # Запоминаем ошибку
                                error_info = {
                                    'position': sf_pos,
                                    'type': device_type_clean,
                                    'nominal_from_pdf': None,
                                    'context': 'not_found_in_library'
                                }
                    else:
                        # Библиотека не загружена, ищем номинал по шаблону 'XX A' в PDF
                        for offset in range(-2, 3):
                            check_idx = i + offset
                            if 0 <= check_idx < len(lines):
                                check_line = lines[check_idx].strip()
                                nominal_match = re.search(r'(\d+(?:\.\d+)?)\s*A', check_line)
                                if nominal_match:
                                    nominal_current = int(float(nominal_match.group(1)))
                                    logger.info(f"   Найден автомат: {sf_pos} → {device_type_clean} (номинал из PDF: {nominal_current}A)")
                                    break
                        else:
                            nominal_current = "Не указан"
                            logger.info(f"   Найден автомат: {sf_pos} → {device_type_clean} (номинал не указан в PDF)")
                else:
                    device_type = "Тип не определён"
                    nominal_current = "Не указан"
                    logger.info(f"   Найден SF: {sf_pos} (тип не определён)")
                
                # Определяем номер схемы (нужно найти заголовок "Перечень элементов схемы X.X" выше по тексту)
                scheme_num = "Не определена"
                for j in range(max(0, i-20), i):
                    scheme_match = re.search(r'Перечень элементов схемы\s+([\d\.]+)', lines[j], re.IGNORECASE)
                    if scheme_match:
                        scheme_num = scheme_match.group(1)
                        break
                
                # Сохраняем результат
                if scheme_num not in breakers_data:
                    breakers_data[scheme_num] = []
                
                # Проверяем дубликаты
                if not any(pos == sf_pos for pos, _, _ in breakers_data[scheme_num]):
                    breakers_data[scheme_num].append((sf_pos, device_type, nominal_current))
                    
                    # Если есть информация об ошибке, добавляем в список ошибок
                    if error_info:
                        error_info['scheme'] = scheme_num
                        errors_automats.append(error_info)
        
        i += 1
    
    return breakers_data, errors_automats

def print_errors_summary(errors_automats, breakers_list=None):
    """
    Выводит сводку по автоматам, не найденным в библиотеке
    """
    if not errors_automats:
        return
    
    logger.info(f"\n⚠️ Строки автоматов, требующие добавления в библиотеку:")
    
    # Собираем уникальные типы для вывода
    unique_error_types = {}
    for err in errors_automats:
        error_type = err['type']
        if error_type not in unique_error_types:
            unique_error_types[error_type] = err
    
    # Выводим все уникальные типы
    for error_type in sorted(unique_error_types.keys()):
        err = unique_error_types[error_type]
        if err.get('nominal_from_pdf'):
            logger.info(f"   - [тип '{err['type']}' - номинал {err['nominal_from_pdf']}A найден в PDF]")
        else:
            logger.info(f"   - [тип '{err['type']}' не найден в библиотеке]")
    
    # Показываем статистику и примеры
    logger.info(f"\n📊 Всего уникальных типов, отсутствующих в библиотеке: {len(unique_error_types)}")
    logger.info(f"   Общее количество позиций с ошибками: {len(errors_automats)}")
    
    # Примеры (первые 5)
    if len(errors_automats) > 0:
        logger.info(f"\n   Примеры (первые 5 из {len(errors_automats)}):")
        for err in errors_automats[:5]:
            nominal_info = f"({err.get('nominal_from_pdf', '?')}A)" if err.get('nominal_from_pdf') else "(номинал не определён)"
            logger.info(f"      {err['position']} (схема {err.get('scheme', '?')}) → {err['type'][:40]} {nominal_info}")

def extract_basket_name(sheet):
    """Извлекает наименование корзины из строки 'Таблица фидеров:'"""
    for row in range(1, min(20, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                if "Таблица фидеров:" in cell_value:
                    match = re.search(r'Таблица фидеров:\s*([^\s]+)', cell_value)
                    if match:
                        return match.group(1)
    return "Корзина не определена"

def find_headers_row(sheet):
    """Находит строку с заголовками"""
    for row in range(1, min(30, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                if "Позиционное обозначение выкатного модуля" in cell_value:
                    return row
    return None

def find_column_by_header(sheet, header_row, target_headers):
    """Ищет номер столбца по заголовку"""
    if isinstance(target_headers, str):
        target_headers = [target_headers]
    
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=header_row, column=col).value
        if cell_value and isinstance(cell_value, str):
            cell_clean = cell_value.strip().lower()
            for target in target_headers:
                target_clean = target.strip().lower()
                if cell_clean == target_clean or cell_clean.startswith(target_clean):
                    return col
    return None

def extract_circuit_breakers_from_baskets(excel_file_path, pdf_breakers, errors_automats):
    """
    Извлекает корзины из Excel и сопоставляет с автоматами из PDF
    """
    # Проверяем, что pdf_breakers - это словарь (даже пустой)
    if pdf_breakers is None:
        pdf_breakers = {}
    
    # Проверяем, что errors_automats - это список (даже пустой)
    if errors_automats is None:
        errors_automats = []
    
    wb = load_workbook(excel_file_path, data_only=True)
    sheet = wb.active
    
    logger.info(f"📊 Обработка Excel файла: {os.path.basename(excel_file_path)}")
    logger.info("=" * 50)
    
    basket_name = extract_basket_name(sheet)
    logger.info(f"Наименование корзины: {basket_name}")
    
    header_row = find_headers_row(sheet)
    if header_row is None:
        logger.info("❌ Не найдена строка с заголовками!")
        return [], basket_name
    
    col_module = find_column_by_header(sheet, header_row, ["Позиционное обозначение выкатного модуля"])
    col_scheme = find_column_by_header(sheet, header_row, ["№ схемы КТС/Э3"])
    
    if col_module is None:
        logger.info("❌ Столбец 'Позиционное обозначение выкатного модуля' не найден!")
        return [], basket_name
    
    logger.info(f"✅ Столбец корзины: {col_module}")
    if col_scheme:
        logger.info(f"✅ Столбец схемы КТС: {col_scheme}")
    logger.info("=" * 50)
    
    # Собираем все строки (корзины) из Excel
    modules_data = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        if should_stop():
            logger.info("⏹️ Остановка по запросу пользователя")
            return [], basket_name
        module_value = sheet.cell(row=row_idx, column=col_module).value
        if not module_value or str(module_value).strip() == "":
            continue
        
        module_str = str(module_value).strip()
        scheme_value = sheet.cell(row=row_idx, column=col_scheme).value if col_scheme else None
        scheme_num = str(scheme_value).strip() if scheme_value else ""
        
        modules_data.append({
            'module': module_str,
            'scheme': scheme_num,
            'row_idx': row_idx
        })
    
    # Список для всех автоматов
    all_breakers = []
    counter = 1
    
    # Если pdf_breakers пуст, выводим предупреждение
    if not pdf_breakers:
        logger.info("\n⚠️ ВНИМАНИЕ: PDF с КТС не содержит данных об автоматах или не был найден!")
        logger.info("   Автоматические выключатели не будут сопоставлены с корзинами.")
    
    # Обрабатываем каждую корзину
    for module_info in modules_data:
        if should_stop():
            logger.info("⏹️ Остановка по запросу пользователя")
            return all_breakers, basket_name
        module_str = module_info['module']
        scheme_num = module_info['scheme']
        
        # Если для схемы этой корзины есть автоматы в PDF
        if scheme_num and scheme_num in pdf_breakers:
            breakers_list = pdf_breakers[scheme_num]  # список кортежей (позиция, тип, номинал)
            
            for breaker_position, device_type, nominal_current in breakers_list:
                # Формат: {оригинальная_позиция}-{корзина}
                scheme_designation = f"{breaker_position}-{module_str}"
                
                all_breakers.append({
                    "number": counter,
                    "panel": basket_name,
                    "position": breaker_position,
                    "scheme_designation": scheme_designation,
                    "device_type": device_type,
                    "nominal_current": nominal_current,
                    "source_scheme": f"Схема {scheme_num} (корзина {module_str})"
                })
                counter += 1
        elif scheme_num:
            # Схема есть в Excel, но в PDF для неё нет автоматов
            logger.info(f"   ⚠️ Для схемы {scheme_num} (корзина {module_str}) автоматы не найдены в PDF")
    
    # Вывод статистики
    logger.info(f"\n📊 ИТОГО:")
    logger.info(f"   - Корзин обработано: {len(modules_data)}")
    logger.info(f"   - Автоматов найдено: {len(all_breakers)}")
    
    # Выводим ошибки
    print_errors_summary(errors_automats, all_breakers)
    
    return all_breakers, basket_name

def save_breakers_to_xlsx(breakers_list, output_file_path):
    """Сохраняет автоматические выключатели в Excel файл"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Автоматы в корзинах"
    
    # Заголовки
    ws.append([
        "№ п/п",
        "Щит",
        "Позиция автомата (SF)",
        "Схемное обозначение",
        "Тип автомата",
        "Номинальный ток, А",
        "Источник (схема КТС)"
    ])
    
    for breaker in breakers_list:
        ws.append([
            breaker["number"],
            breaker["panel"],
            breaker["position"],
            breaker["scheme_designation"],
            breaker["device_type"],
            breaker["nominal_current"],
            breaker["source_scheme"]
        ])
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 45
    
    wb.save(output_file_path)
    return output_file_path

def main():
    # Проверка на режим GUI (вызов из Streamlit)
    is_gui_mode = '--from-gui' in sys.argv
    
    # Проверяем флаг --output-dir
    global OUTPUT_DIR
    output_dir = None
    if '--output-dir' in sys.argv:
        idx = sys.argv.index('--output-dir')
        if idx + 1 < len(sys.argv):
            output_dir = sys.argv[idx + 1]
            sys.argv.pop(idx)
            sys.argv.pop(idx)
    
    if output_dir:
        OUTPUT_DIR = Path(output_dir)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    if is_gui_mode:
        # Убираем аргумент --from-gui из списка
        args = [arg for arg in sys.argv[1:] if arg != '--from-gui']
        if not args:
            logger.info("❌ Ошибка: не указан путь к файлу!")
            return
        excel_file_path = args[0]
    else:
        # Обычный режим (из проводника)
        if len(sys.argv) < 2:
            logger.info("Использование: перетащите XLSX файл на этот скрипт")
            input("\nНажмите Enter для выхода...")
            return
        excel_file_path = sys.argv[1]
    
    if not os.path.exists(excel_file_path):
        logger.info(f"❌ Ошибка: Файл '{excel_file_path}' не найден!")
        if not is_gui_mode:
            input("Нажмите Enter для выхода...")
        return
    
    logger.info(f"\n📁 Обработка файла: {os.path.basename(excel_file_path)}")
    logger.info("=" * 50)
    
    # Загружаем библиотеку номиналов
    logger.info("📚 Загрузка библиотеки номинальных токов...")
    nominals_library = load_nominals_library()
    logger.info("=" * 50)
    
    # Ищем PDF с КТС в ПАПКЕ С EXCEL ФАЙЛОМ
    excel_folder = os.path.dirname(os.path.abspath(excel_file_path))
    logger.info(f"📁 Поиск PDF с КТС в папке Excel файла: {excel_folder}")
    ktc_pdf = find_ktc_pdf(excel_folder)
    
    if ktc_pdf:
        logger.info(f"✅ Найден PDF: {os.path.basename(ktc_pdf)}")
    else:
        logger.info(f"⚠️ PDF с 'КТС' или 'КРУС-КТС' в названии не найден в папке с Excel файлом!")
        logger.info(f"   Папка: {excel_folder}")
        logger.info(f"   Автоматы не будут определены.")
    
    # Извлекаем автоматы из PDF с использованием библиотеки номиналов
    pdf_breakers, errors_automats = extract_circuit_breakers_from_pdf(ktc_pdf, nominals_library)
    
    if pdf_breakers:
        logger.info(f"\n📊 Найдено схем с автоматами: {len(pdf_breakers)}")
        total_breakers = sum(len(v) for v in pdf_breakers.values())
        logger.info(f"📊 Всего позиций автоматов: {total_breakers}")
    else:
        logger.info("\n⚠️ Автоматы в PDF не найдены или PDF отсутствует.")
    
    logger.info("=" * 50)
    
    try:
        breakers, basket_name = extract_circuit_breakers_from_baskets(excel_file_path, pdf_breakers, errors_automats)
        
        if breakers:
            # Определяем путь для сохранения результата
            if OUTPUT_DIR:
                output_dir_path = OUTPUT_DIR
            else:
                output_dir_path = Path(excel_file_path).parent
            
            base_name = Path(excel_file_path).stem
            output_path = output_dir_path / f"{base_name}_Автоматические_выключатели_в_корзинах.xlsx"
            
            save_breakers_to_xlsx(breakers, str(output_path))
            logger.info(f"\n✅ Обработано автоматов: {len(breakers)}")
            logger.info(f"📁 Результат сохранён в: {output_path}")
            
            # Вывод примера
            logger.info("\n📋 ПРИМЕР найденных автоматов:")
            for breaker in breakers[:5]:
                logger.info(f"   {breaker['scheme_designation']} → {breaker['device_type']} ({breaker['nominal_current']}A)")
        else:
            logger.info("\n⚠️ Автоматы не найдены в файле!")
    
    except Exception as e:
        logger.info(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
    
    logger.info("\n" + "=" * 50)
    
    # Только в обычном режиме ждем нажатия Enter
    if not is_gui_mode:
        input("Нажмите Enter для выхода...")

if __name__ == "__main__":
    main()