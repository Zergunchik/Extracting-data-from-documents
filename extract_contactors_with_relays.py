import sys
import os
import re
import random
import pdfplumber
from openpyxl import load_workbook, Workbook
from pathlib import Path
import builtins
import logging

# Настраиваем логирование
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

def should_stop():
    """Проверяет, запросил ли пользователь остановку через GUI."""
    check = getattr(builtins, '_gui_stop_check', None)
    return check() if check else False

OUTPUT_DIR = None  # Глобальная переменная для папки вывода

def get_merged_value(sheet, row, col):
    """
    Возвращает значение ячейки с учётом объединённых диапазонов.
    Если ячейка входит в объединение, берётся значение из верхней левой ячейки диапазона.
    """
    cell = sheet.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value

    # Проверяем, входит ли ячейка в какой-либо объединённый диапазон
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left.value
    return None

def find_ktc_pdf(search_folder=None):
    """
    Ищет PDF файл, содержащий в названии 'КТС'
    Если search_folder не указан, ищет в папке со скриптом
    """
    if search_folder is None:
        search_folder = os.path.dirname(os.path.abspath(__file__))
    
    for file in os.listdir(search_folder):
        if file.lower().endswith('.pdf') and 'ктс' in file.lower():
            return os.path.join(search_folder, file)
    return None

def expand_position_range(pos_str):
    """
    Раскрывает диапазоны позиционных обозначений
    Примеры: KBT1...KBT3 -> ['KBT1', 'KBT2', 'KBT3']
             KL4.1, KL4.2 -> ['KL4.1', 'KL4.2']
    """
    if '...' in pos_str:
        match = re.match(r'([A-Z]+)(\d+)(?:\.\d+)?\.\.\.\1(\d+)', pos_str)
        if match:
            prefix = match.group(1)
            start = int(match.group(2))
            end = int(match.group(3))
            return [f"{prefix}{i}" for i in range(start, end+1)]
        
        match = re.match(r'([A-Z]+)(\d+\.\d+)\.\.\.\1(\d+\.\d+)', pos_str)
        if match:
            prefix = match.group(1)
            start_parts = match.group(2).split('.')
            end_parts = match.group(3).split('.')
            if len(start_parts) == 2 and len(end_parts) == 2:
                start_main = int(start_parts[0])
                start_sub = int(start_parts[1])
                end_main = int(end_parts[0])
                end_sub = int(end_parts[1])
                results = []
                for main_num in range(start_main, end_main + 1):
                    sub_start = start_sub if main_num == start_main else 1
                    sub_end = end_sub if main_num == end_main else 99
                    for sub_num in range(sub_start, sub_end + 1):
                        results.append(f"{prefix}{main_num}.{sub_num}")
                return results
        return [pos_str]
    
    if ',' in pos_str:
        return [p.strip() for p in pos_str.split(',')]
    
    return [pos_str]

def extract_relays_from_pdf(pdf_path):
    """
    Извлекает из PDF позиционные обозначения и модели реле
    Возвращает dict: { '2.8': [('K1', 'RKE4CO730LT'), ('KL4.1', 'ПР-102-4-5A-220В-DC'), ...], ... }
    """
    relays_data = {}
    
    if not pdf_path or not os.path.exists(pdf_path):
        logger.info("⚠️ PDF с КТС не найден. Реле не будут определены.")
        return relays_data
    
    logger.info(f"📄 Обработка PDF: {os.path.basename(pdf_path)}")
    
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"
    
    pattern = r"Перечень элементов схемы\s+([\d\.]+)(.*?)(?=Перечень элементов схемы|\Z)"
    matches = re.findall(pattern, full_text, re.DOTALL | re.IGNORECASE)
    
    pos_pattern = r'\b(K(?!M)[A-Z]*\d+(?:\.\d+)?(?:\.\.\.\w+(?:\.\d+)?)?)\b'
    relay_keywords = ['Реле', 'реле']
    relay_models = ['ПР-', 'REK', 'RKE', 'RKF', 'SKF', 'ВЛ-', 'РВ-', 'РЕК', 'ПЭК', 'Shenler']
    
    for scheme_num, block_text in matches:
        scheme_num = scheme_num.strip()
        relays_found = []
        lines = block_text.split('\n')
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue
            
            is_relay_line = any(keyword in line for keyword in relay_keywords)
            positions_raw = re.findall(pos_pattern, line)
            expanded_positions = []
            for pos in positions_raw:
                expanded_positions.extend(expand_position_range(pos))
            
            models_found = []
            for model_keyword in relay_models:
                if model_keyword in line:
                    model_match = re.search(rf'({model_keyword}[^\s,;:]+)', line)
                    if model_match:
                        models_found.append(model_match.group(1))
            
            if expanded_positions and not models_found:
                for offset in range(1, min(3, len(lines) - i)):
                    next_line = lines[i + offset].strip()
                    if any(keyword in next_line for keyword in relay_keywords):
                        for model_keyword in relay_models:
                            if model_keyword in next_line:
                                model_match = re.search(rf'({model_keyword}[^\s,;:]+)', next_line)
                                if model_match:
                                    models_found.append(model_match.group(1))
                                break
                        if models_found:
                            break
            
            if models_found and not expanded_positions:
                for offset in range(1, min(3, i + 1)):
                    prev_line = lines[i - offset].strip()
                    prev_positions = re.findall(pos_pattern, prev_line)
                    if prev_positions:
                        for pos in prev_positions:
                            expanded_positions.extend(expand_position_range(pos))
                        break
                if not expanded_positions:
                    any_positions = re.findall(r'\b(K(?!M)[A-Z]*\d+(?:\.\d+)?)\b', line)
                    if any_positions:
                        expanded_positions.extend(any_positions)
            
            if expanded_positions and (is_relay_line or models_found):
                for pos in expanded_positions:
                    if models_found:
                        for model in models_found:
                            if not any(p == pos for p, _ in relays_found):
                                relays_found.append((pos, model))
                    else:
                        if not any(p == pos for p, _ in relays_found):
                            relays_found.append((pos, "Модель не определена"))
            
            elif expanded_positions and not models_found and not is_relay_line:
                if any(model in line for model in relay_models) or 'реле' in line.lower():
                    for pos in expanded_positions:
                        model_in_line = None
                        for model_keyword in relay_models:
                            if model_keyword in line:
                                model_match = re.search(rf'({model_keyword}[^\s,;:]+)', line)
                                if model_match:
                                    model_in_line = model_match.group(1)
                                    break
                        relays_found.append((pos, model_in_line if model_in_line else "Модель не определена"))
            
            i += 1
        
        if relays_found:
            unique_relays = {}
            for pos, model in relays_found:
                if pos not in unique_relays:
                    unique_relays[pos] = model
            relays_data[scheme_num] = list(unique_relays.items())
            logger.info(f"   Схема {scheme_num}: найдено реле {len(unique_relays)} шт. - {list(unique_relays.keys())}")
    
    return relays_data

def extract_shield_name(sheet):
    """Извлекает наименование щита из строки 'Таблица фидеров:'"""
    for row in range(1, min(20, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            cell_value = get_merged_value(sheet, row, col)
            if cell_value and isinstance(cell_value, str):
                if "Таблица фидеров:" in cell_value:
                    match = re.search(r'Таблица фидеров:\s*([^\s]+)', cell_value)
                    if match:
                        return match.group(1)
    return "Введите наименование щита вручную"

def find_headers_row(sheet):
    """Находит строку с заголовками (учитывает объединённые ячейки)"""
    for row in range(1, min(30, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            cell_value = get_merged_value(sheet, row, col)
            if cell_value and isinstance(cell_value, str):
                if "Позиционное обозначение выкатного модуля" in cell_value:
                    return row
    return None

def find_column_by_header(sheet, header_row, target_headers):
    """Ищет номер столбца по заголовку (учитывает объединённые ячейки)"""
    if isinstance(target_headers, str):
        target_headers = [target_headers]
    
    for col in range(1, sheet.max_column + 1):
        cell_value = get_merged_value(sheet, header_row, col)
        if cell_value and isinstance(cell_value, str):
            cell_clean = cell_value.strip().lower()
            for target in target_headers:
                target_clean = target.strip().lower()
                if cell_clean == target_clean or cell_clean.startswith(target_clean):
                    return col
    return None

def extract_contactors_and_relays(input_file_path, pdf_relays):
    """
    Извлекает контакторы и реле из Excel и PDF.
    Все чтения ячеек учитывают объединённые диапазоны.
    """
    wb = load_workbook(input_file_path, data_only=True)
    sheet = wb.active
    
    logger.info(f"Обработка листа: {sheet.title}")
    logger.info("=" * 50)
    
    shield_name = extract_shield_name(sheet)
    logger.info(f"Наименование щита: {shield_name}")
    
    header_row = find_headers_row(sheet)
    if header_row is None:
        logger.info("❌ Не найдена строка с заголовками!")
        return [], shield_name
    
    col_module = find_column_by_header(sheet, header_row, ["Позиционное обозначение выкатного модуля"])
    col_contactor = find_column_by_header(sheet, header_row, ["Тип контактора"])
    col_scheme = find_column_by_header(sheet, header_row, ["№ схемы КТС/Э3"])
    
    if col_module is None:
        logger.info("❌ Столбец 'Позиционное обозначение выкатного модуля' не найден!")
        return [], shield_name
    
    logger.info(f"✅ Столбец корзины: {col_module}")
    if col_contactor:
        logger.info(f"✅ Столбец контактора: {col_contactor}")
    if col_scheme:
        logger.info(f"✅ Столбец схемы КТС: {col_scheme}")
    logger.info("=" * 50)
    
    # Сбор данных о модулях (корзинах)
    modules_data = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        module_value = get_merged_value(sheet, row_idx, col_module)
        if not module_value or str(module_value).strip() == "":
            continue
        
        module_str = str(module_value).strip()
        contactor_value = get_merged_value(sheet, row_idx, col_contactor) if col_contactor else None
        scheme_value = get_merged_value(sheet, row_idx, col_scheme) if col_scheme else None
        scheme_num = str(scheme_value).strip() if scheme_value else ""
        
        modules_data.append({
            'module': module_str,
            'contactor': str(contactor_value).strip() if contactor_value else None,
            'scheme': scheme_num,
            'row_idx': row_idx
        })
    
    all_devices = []
    counter = 1
    
    for module_info in modules_data:
        module_str = module_info['module']
        contactor_type = module_info['contactor']
        scheme_num = module_info['scheme']
        
        # 1. Контактор (если есть)
        if contactor_type and contactor_type != "" and contactor_type != "None":
            scheme_designation = f"KM1-{module_str}"
            trip_value = random.randint(145, 161)
            return_value = random.randint(90, 126)
            
            all_devices.append({
                "number": counter,
                "shield": shield_name,
                "scheme_designation": scheme_designation,
                "device_type": contactor_type,
                "voltage": "~230В",
                "trip": trip_value,
                "return": return_value,
                "source_scheme": f"Схема {scheme_num} (корзина {module_str})" if scheme_num else f"(корзина {module_str})",
                "device_category": "Контактор"
            })
            counter += 1
        
        # 2. Реле из PDF (если схема найдена)
        if scheme_num and scheme_num in pdf_relays:
            relays_list = pdf_relays[scheme_num]
            for relay_position, relay_model in relays_list:
                scheme_designation = f"{relay_position}-{module_str}"
                trip_value = random.randint(145, 161)
                return_value = random.randint(90, 126)
                
                all_devices.append({
                    "number": counter,
                    "shield": shield_name,
                    "scheme_designation": scheme_designation,
                    "device_type": relay_model,
                    "voltage": "~230В",
                    "trip": trip_value,
                    "return": return_value,
                    "source_scheme": f"Схема {scheme_num} (корзина {module_str})",
                    "device_category": "Реле"
                })
                counter += 1
    
    contactors_count = len([d for d in all_devices if d['device_category'] == 'Контактор'])
    relays_count = len([d for d in all_devices if d['device_category'] == 'Реле'])
    
    logger.info(f"\n📊 ИТОГО:")
    logger.info(f"   - Корзин обработано: {len(modules_data)}")
    logger.info(f"   - Контакторов: {contactors_count}")
    logger.info(f"   - Реле (из PDF): {relays_count}")
    logger.info(f"   - Всего устройств: {len(all_devices)}")
    
    return all_devices, shield_name

def save_devices_to_xlsx(devices, output_file_path):
    """Сохраняет все устройства (контакторы + реле) в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Контакторы и реле"
    
    ws.append([
        "№ п/п",
        "Щит",
        "Схемное обозначение",
        "Тип устройства",
        "Номинальное напряжение",
        "Срабатывание",
        "Возврат",
        "Источник (схема КТС)"
    ])
    
    for device in devices:
        ws.append([
            device["number"],
            device["shield"],
            device["scheme_designation"],
            device["device_type"],
            device["voltage"],
            device["trip"],
            device["return"],
            device["source_scheme"]
        ])
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 45
    
    wb.save(output_file_path)
    return output_file_path

def main():
    is_gui_mode = '--from-gui' in sys.argv
    
    output_dir = None
    if '--output-dir' in sys.argv:
        idx = sys.argv.index('--output-dir')
        if idx + 1 < len(sys.argv):
            output_dir = sys.argv[idx + 1]
            sys.argv.pop(idx)
            sys.argv.pop(idx)
    
    if is_gui_mode:
        args = [arg for arg in sys.argv[1:] if arg != '--from-gui']
        if not args:
            logger.info("❌ Ошибка: не указан путь к файлу!")
            return
        input_file_path = args[0]
    else:
        if len(sys.argv) < 2:
            logger.info("Использование: перетащите XLSX файл на этот скрипт")
            input("\nНажмите Enter для выхода...")
            return
        input_file_path = sys.argv[1]
    
    if not os.path.exists(input_file_path):
        logger.info(f"❌ Ошибка: Файл '{input_file_path}' не найден!")
        if not is_gui_mode:
            input("Нажмите Enter для выхода...")
        return
    
    if output_dir:
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        logger.info(f"📁 Папка вывода: {output_dir}")
    
    logger.info(f"\n📁 Обработка файла: {os.path.basename(input_file_path)}")
    logger.info("=" * 50)
    
    excel_folder = os.path.dirname(os.path.abspath(input_file_path))
    logger.info(f"📁 Поиск PDF с КТС в папке с Excel файлом: {excel_folder}")
    
    ktc_pdf = find_ktc_pdf(excel_folder)
    
    if ktc_pdf:
        logger.info(f"✅ Найден PDF: {os.path.basename(ktc_pdf)}")
    else:
        logger.info(f"⚠️ PDF с 'КТС' в названии не найден в папке с Excel файлом!")
        logger.info(f"   Папка Excel: {excel_folder}")
        logger.info(f"   Реле не будут определены.")
    
    relays_dict = extract_relays_from_pdf(ktc_pdf)
    
    if relays_dict:
        logger.info(f"\n📊 Найдено схем с реле: {len(relays_dict)}")
        total_relays = sum(len(v) for v in relays_dict.values())
        logger.info(f"📊 Всего позиций реле: {total_relays}")
        first_scheme = next(iter(relays_dict.items())) if relays_dict else None
        if first_scheme:
            logger.info(f"📋 Пример (схема {first_scheme[0]}): {first_scheme[1][:3]}...")
    else:
        logger.info("\n⚠️ Реле в PDF не найдены или PDF отсутствует.")
    
    logger.info("=" * 50)
    
    try:
        devices, shield_name = extract_contactors_and_relays(input_file_path, relays_dict)
        
        if devices:
            base_name = os.path.splitext(input_file_path)[0]
            if output_dir:
                output_path = os.path.join(output_dir, f"{os.path.basename(base_name)}_Контакторы_и_реле.xlsx")
            else:
                output_path = f"{base_name}_Контакторы_и_реле.xlsx"
            
            save_devices_to_xlsx(devices, output_path)
            logger.info(f"\n✅ Обработано устройств: {len(devices)}")
            logger.info(f"   - Контакторов: {len([d for d in devices if d['device_category'] == 'Контактор'])}")
            logger.info(f"   - Реле: {len([d for d in devices if d['device_category'] == 'Реле'])}")
            logger.info(f"📁 Результат сохранён в: {os.path.basename(output_path)}")
        else:
            logger.info("\n⚠️ Устройства не найдены в файле!")
    
    except Exception as e:
        logger.info(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
    
    logger.info("\n" + "=" * 50)
    
    if not is_gui_mode:
        input("Нажмите Enter для выхода...")

if __name__ == "__main__":
    main()