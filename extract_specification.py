# extract_specification.py - Объединенное извлечение реле и вторичных автоматов из PDF
import pdfplumber
import pandas as pd
import openpyxl
import re
import os
import sys
from pathlib import Path
import time
import builtins
import tkinter as tk
from tkinter import ttk

def should_stop():
    """Проверяет, запросил ли пользователь остановку через GUI."""
    check = getattr(builtins, '_gui_stop_check', None)
    return check() if check else False


def show_relay_voltage_dialog(unique_relays):
    """
    Показывает GUI диалог для выбора номинального напряжения для каждого типа реле.
    unique_relays: список словарей {'type': 'RCL424730', 'positions': [...]}
    Возвращает: словарь {relay_type: voltage} или None если отмена
    """
    result = {}
    cancelled = False
    
    root = tk.Tk()
    root.title("Выбор номинального напряжения реле")
    root.geometry("600x400")
    
    # Фрейм для таблицы
    table_frame = ttk.Frame(root)
    table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    # Создаем Treeview с колонками
    columns = ('num', 'type', 'v230', 'v220', 'v24')
    tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=min(len(unique_relays), 15))
    
    # Настраиваем заголовки
    tree.heading('num', text='№ п/п')
    tree.heading('type', text='Тип реле')
    tree.heading('v230', text='~230В')
    tree.heading('v220', text='=220В')
    tree.heading('v24', text='=24В')
    
    # Настраиваем ширину колонок
    tree.column('num', width=50, anchor=tk.CENTER)
    tree.column('type', width=200, anchor=tk.W)
    tree.column('v230', width=80, anchor=tk.CENTER)
    tree.column('v220', width=80, anchor=tk.CENTER)
    tree.column('v24', width=80, anchor=tk.CENTER)
    
    # Добавляем скроллбар
    scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    # Заполняем таблицу
    for idx, relay in enumerate(unique_relays, 1):
        relay_type = relay['type']
        tree.insert('', tk.END, values=(idx, relay_type, '', '', ''), tags=(relay_type,))
    
        # Обработчик клика по ячейке с напряжением
    def on_cell_click(event):
        region = tree.identify("region", event.x, event.y)
        if region == "cell":
            column = tree.identify_column(event.x)
            item = tree.identify_row(event.y)
            if column in ('#3', '#4', '#5'):  # Колонки с напряжением
                current_values = tree.item(item, 'values')
                
                # Сбрасываем все галочки в этой строке
                new_values = list(current_values)
                new_values[2] = ''  # ~230В
                new_values[3] = ''  # =220В
                new_values[4] = ''  # =24В
                
                # Ставим галочку в нажатой колонке (ИСПРАВЛЕНО)
                col_idx = int(column[1:]) - 1  
                new_values[col_idx] = '✓'
                
                tree.item(item, values=tuple(new_values))
    
    tree.bind('<Button-1>', on_cell_click)
    
    # Кнопки ОК и Отмена
    btn_frame = ttk.Frame(root)
    btn_frame.pack(fill=tk.X, padx=10, pady=10)
    
    def on_ok():
        for item in tree.get_children():
            values = tree.item(item, 'values')
            relay_type = values[1]
            if values[2]:  # ~230В
                result[relay_type] = '~230В'
            elif values[3]:  # =220В
                result[relay_type] = '=220В'
            elif values[4]:  # =24В
                result[relay_type] = '=24В'
            else:
                result[relay_type] = ''  # Не выбрано
        root.destroy()
    
    def on_cancel():
        nonlocal cancelled
        cancelled = True
        root.destroy()
    
    btn_ok = ttk.Button(btn_frame, text="OK", command=on_ok)
    btn_ok.pack(side=tk.LEFT, padx=5)
    
    btn_cancel = ttk.Button(btn_frame, text="Отмена", command=on_cancel)
    btn_cancel.pack(side=tk.LEFT, padx=5)
    
    # Инструкция
    lbl = ttk.Label(root, text="Выберите напряжение кликом по соответствующей колонке")
    lbl.pack(pady=(0, 10))
    
    root.mainloop()
    
    return None if cancelled else result


def get_unique_relay_types(relays_list):
    """Возвращает список уникальных типов реле с их позициями"""
    unique = {}
    for relay in relays_list:
        relay_type = relay.get('type', '')
        pos = relay.get('pos', '')
        if relay_type not in unique:
            unique[relay_type] = []
        unique[relay_type].append(pos)
    
    return [{'type': t, 'positions': p} for t, p in unique.items()]


# ============================================================
# ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ И КОНСТАНТЫ
# ============================================================

# --- СПИСОК ИСКЛЮЧАЕМЫХ КЛЮЧЕВЫХ СЛОВ (для реле) ---
NEGATIVE_KEYWORDS = [
    'компрессор', 'компресс',
    'насос', 'pump',
    'chiller',
    'кондиционер внешний',
    'двигатель',
    'мотор'
]

# --- СПИСОК КЛЮЧЕВЫХ СЛОВ ДЛЯ РЕЛЕ ---
RELAY_KEYWORDS = [
    'реле',
    'промежуточное реле',
    'реле промежуточное',
    'RKE4CO024LT',
    'RKE4CO730LT'
]

# --- ИСКЛЮЧАЕМЫЕ ТИПЫ РЕЛЕ ---
EXCLUDED_RELAY_TYPES = [
    'реле напряжения',
    'реле контроля напряжения',
    'KV',
    'KVZ',
    'контроль напряжения'
]


# ============================================================
# ФУНКЦИИ ИЗВЛЕЧЕНИЯ ДАННЫХ ИЗ PDF
# ============================================================

def extract_text_from_pdf(pdf_path):
    """Извлекает текст из PDF файла."""
    extracted_data = []
    print(f"   📄 Чтение текста из: {os.path.basename(pdf_path)}")
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            print(f"   📄 Всего страниц: {total_pages}")
            
            if total_pages == 0:
                print(f"   ⚠️ PDF файл не содержит страниц!")
                return None
                
            for page_num, page in enumerate(pdf.pages, 1):
                if page_num % 5 == 0 or page_num == total_pages:
                    print(f"   ⏳ Обработка текста: страница {page_num}/{total_pages}")
                
                text = page.extract_text()
                if text:
                    lines = text.split('\n')
                    for line_num, line in enumerate(lines, 1):
                        if line.strip():
                            extracted_data.append({
                                'Страница': page_num,
                                'Строка': line_num,
                                'Текст': line.strip()
                            })
            
            print(f"   ✅ Извлечено {len(extracted_data)} строк текста")
            return extracted_data
    except Exception as e:
        print(f"❌ Ошибка при чтении текста PDF: {e}")
        import traceback
        traceback.print_exc()
        return None


def extract_tables_from_pdf(pdf_path):
    """Извлекает таблицы из PDF файла."""
    all_tables = []
    print(f"   📊 Чтение таблиц из: {os.path.basename(pdf_path)}")
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            print(f"   📊 Всего страниц: {total_pages}")
            
            for page_num, page in enumerate(pdf.pages, 1):
                if page_num % 5 == 0 or page_num == total_pages:
                    print(f"   ⏳ Обработка таблиц: страница {page_num}/{total_pages}")
                
                try:
                    tables = page.extract_tables()
                    if tables:
                        print(f"   📊 Страница {page_num}: найдено {len(tables)} таблиц")
                    else:
                        tables = page.extract_tables({
                            "vertical_strategy": "lines",
                            "horizontal_strategy": "lines"
                        })
                        if tables:
                            print(f"   📊 Страница {page_num}: найдено {len(tables)} таблиц (альтернативный метод)")
                    
                    for table_num, table in enumerate(tables, 1):
                        if table:
                            clean_table = []
                            for row in table:
                                clean_row = [str(cell).strip() if cell is not None else '' for cell in row]
                                if any(clean_row):
                                    clean_table.append(clean_row)
                            if clean_table:
                                all_tables.append({
                                    'Страница': page_num,
                                    'Таблица': table_num,
                                    'Данные': clean_table
                                })
                                print(f"   📊 Таблица {table_num} на странице {page_num}: {len(clean_table)} строк, {len(clean_table[0]) if clean_table else 0} колонок")
                except Exception as page_e:
                    print(f"   ⚠️ Ошибка на странице {page_num}: {page_e}")
                    continue
    except Exception as e:
        print(f"❌ Ошибка при извлечении таблиц: {e}")
        import traceback
        traceback.print_exc()
        return all_tables
    
    print(f"   ✅ Всего извлечено {len(all_tables)} таблиц")
    return all_tables


# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ РЕЛЕ
# ============================================================

def normalize_text(text):
    """Нормализация текста (замена русских букв на латинские)"""
    if not text:
        return text
    text = str(text)
    text = text.replace('К', 'K').replace('к', 'k')
    text = text.replace('М', 'M').replace('м', 'm')
    return text


def has_negative_keywords(text):
    """Проверяет, содержит ли текст негативные ключевые слова"""
    if not text:
        return False
    text_lower = text.lower()
    for neg_word in NEGATIVE_KEYWORDS:
        if neg_word in text_lower:
            return True
    return False


def is_garbage_row(row):
    """Проверяет, является ли строка мусорной (служебной информацией)"""
    if not row:
        return True
    
    row_text = " ".join([str(cell).strip() for cell in row if cell])
    
    garbage_patterns = [
        r'^атад$',
        r'^и$',
        r'^ьсипдоП$',
        r'^\.лбуд$',
        r'^№\.внИ$',
        r'^№\.вни$',
        r'^\.мазВ$',
        r'^\.лдоп$',
        r'^Копировал:',
        r'^Изм\.',
        r'^Лист$',
        r'^№ докум\.',
        r'^Подпись$',
        r'^Дата$',
        r'^Формат',
        r'^K\-BLOCK\-',
    ]
    
    for pattern in garbage_patterns:
        if re.search(pattern, row_text, re.IGNORECASE):
            return True
    
    # Если строка содержит только короткие слова - возможно мусор
    words = row_text.split()
    if len(words) <= 4:
        short_words = [w for w in words if len(w) <= 3]
        if len(short_words) / max(len(words), 1) > 0.7:
            return True
    
    return False


def find_columns(data):
    """
    Находит столбцы с позициями и наименованиями
    Возвращает: (pos_col_idx, name_col_idx, start_row_idx)
    """
    if not data or not data[0]:
        return None, None, -1
    
    num_cols = len(data[0])
    pos_col_idx = None
    name_col_idx = None
    start_row = -1
    
    # Первый проход: ищем заголовки в первых 20 строках
    for row_idx, row in enumerate(data[:20]):
        if is_garbage_row(row):
            continue
            
        row_text_lower = " ".join([str(cell).lower() for cell in row if cell])
        
        if not pos_col_idx:
            for col_idx, cell in enumerate(row):
                if cell:
                    cell_str = str(cell).lower().strip()
                    if 'поз' in cell_str or 'обознач' in cell_str:
                        pos_col_idx = col_idx
                        break
        
        if not name_col_idx:
            for col_idx, cell in enumerate(row):
                if cell:
                    cell_str = str(cell).lower().strip()
                    if 'наименование' in cell_str or 'наим' in cell_str or 'описание' in cell_str:
                        name_col_idx = col_idx
                        break
        
        if pos_col_idx is not None and name_col_idx is not None:
            start_row = row_idx + 1
            break
    
    # Второй проход: если заголовки не найдены, ищем по данным
    if pos_col_idx is None or name_col_idx is None:
        print("   🔍 Ищу столбцы по содержимому данных...")
        
        for col_idx in range(num_cols):
            relay_positions = 0
            relay_names = 0
            
            for row in data[20:60]:
                if col_idx < len(row) and row[col_idx]:
                    cell_str = normalize_text(str(row[col_idx]))
                    
                    # Обновленное регулярное выражение с поддержкой KCC, KLP, KLB, KL, K
                    if re.search(r'\b\d*KCC\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLP\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLB\d+(?:-\d+)?(?:\.\d+)?|\b\d*KL?\d+(?:-\d+)?(?:\.\d+)?|\b\d*K\d+(?:-\d+)?(?:\.\d+)?(?!V)', cell_str):
                        relay_positions += 1
                    
                    cell_lower = str(row[col_idx]).lower()
                    for keyword in RELAY_KEYWORDS:
                        if keyword.lower() in cell_lower:
                            relay_names += 1
                            break
            
            if relay_positions > 3 and pos_col_idx is None:
                pos_col_idx = col_idx
                print(f"   ✅ Найдена колонка с позициями: {col_idx}")
            
            if relay_names > 2 and name_col_idx is None:
                name_col_idx = col_idx
                print(f"   ✅ Найдена колонка с наименованиями: {col_idx}")
    
    # Если не нашли стартовую строку, ищем первую строку с данными
    if start_row == -1 and pos_col_idx is not None:
        for row_idx, row in enumerate(data):
            if row_idx < 10:
                continue
            if pos_col_idx < len(row) and row[pos_col_idx]:
                cell_str = normalize_text(str(row[pos_col_idx]))
                # Обновленное регулярное выражение с поддержкой KCC, KLP, KLB, KL, K
                if re.search(r'\b\d*KCC\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLP\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLB\d+(?:-\d+)?(?:\.\d+)?|\b\d*KL?\d+(?:-\d+)?(?:\.\d+)?|\b\d*K\d+(?:-\d+)?(?:\.\d+)?(?!V)', cell_str):
                    start_row = row_idx
                    break
    
    print(f"   📊 Результат: Поз.={pos_col_idx}, Наим.={name_col_idx}, Начало={start_row}")
    return pos_col_idx, name_col_idx, start_row


def expand_positions_relay(text, full_row_text=None):
    """
    Извлекает позиции с реле (KL, K, KLB)
    Поддерживает: K1, K1.1, KL1, KLB1, K1...K10, K1.1...K1.9, 2K1, 4K1, 4K2, K1-4.1...K1-4.6
    """
    if not text or text == 'nan' or text == '':
        return []
    
    if full_row_text and has_negative_keywords(full_row_text):
        return []
    
    positions = []
    text = str(text).replace('\n', ' ').strip()
    normalized_text = normalize_text(text)

    # Исключаем реле напряжения
    if re.search(r'\bKV\d+(?:\.\d+)?|\bKVZ\d+(?:\.\d+)?', normalized_text, re.IGNORECASE):
        return []
    if re.search(r'реле\s+напряжени[яю]', normalized_text, re.IGNORECASE):
        return []

    # Проверяем наличие реле (обновленное регулярное выражение с поддержкой цифр перед буквами и дефисов)
    if not re.search(
        r'\b\d*KCC\d+(?:-\d+)?(?:\.\d+)?|'   # KCC1, KCC2-1, KCC3.1
        r'\b\d*KLP\d+(?:-\d+)?(?:\.\d+)?|'   # KLP1, KLP2-1, KLP3.1
        r'\b\d*KLB\d+(?:-\d+)?(?:\.\d+)?|'   # KLB1, KLB2-1, KLB3.1
        r'\b\d*KL?\d+(?:-\d+)?(?:\.\d+)?|'   # KL1, KL2-1 (и K1 если L нет)
        r'\b\d*K\d+(?:-\d+)?(?:\.\d+)?(?!V)', # K1, K2-1 (но не KV1)
        normalized_text
    ):
        return []

    parts = [p.strip() for p in text.split(',') if p.strip()]

    for part in parts:
        if not part:
            continue
             
        normalized_part = normalize_text(part)
        
        if re.search(r'\bKV\d+(?:\.\d+)?|\bKVZ\d+(?:\.\d+)?', normalized_part, re.IGNORECASE):
            continue
        if re.search(r'реле\s+напряжени[яю]', normalized_part, re.IGNORECASE):
            continue
        
        # Обновленное регулярное выражение с поддержкой KCC, KLP, KLB, KL, K
        if not re.search(r'\b\d*KCC\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLP\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLB\d+(?:-\d+)?(?:\.\d+)?|\b\d*KL?\d+(?:-\d+)?(?:\.\d+)?|\b\d*K\d+(?:-\d+)?(?:\.\d+)?(?!V)', normalized_part):
            continue
        
        # Обработка диапазонов с дефисом (например, K1-4.1...K1-4.6)
        if '...' in normalized_part:
            try:
                range_parts = normalized_part.split('...')
                if len(range_parts) == 2:
                    start_str = range_parts[0].strip()
                    end_str = range_parts[1].strip()
                    
                    # Проверяем, является ли это диапазоном с дефисом
                    match_start = re.search(r'^(\d*)(KCC|KLP|KLB|KL?B?|K?)(\d+)-(\d+)\.(\d+)$', start_str, re.IGNORECASE)
                    match_end = re.search(r'^(\d*)(KCC|KLP|KLB|KL?B?|K?)(\d+)-(\d+)\.(\d+)$', end_str, re.IGNORECASE)
                    
                    if match_start and match_end:
                        prefix_start = match_start.group(1) or ''
                        prefix_end = match_end.group(1) or ''
                        relay_type = match_start.group(2).upper()
                        base_num_start = int(match_start.group(3))
                        base_num_end = int(match_end.group(3))
                        sub_start = int(match_start.group(4))
                        sub_end = int(match_end.group(4))
                        dot_start = int(match_start.group(5))
                        dot_end = int(match_end.group(5))
                        
                        prefix = prefix_start if prefix_start else prefix_end
                        
                        # Генерируем все позиции в диапазоне
                        for base_num in range(base_num_start, base_num_end + 1):
                            if base_num == base_num_start:
                                start_dot = dot_start
                            else:
                                start_dot = 1
                            if base_num == base_num_end:
                                end_dot = dot_end
                            else:
                                end_dot = 99
                            
                            for dot_num in range(start_dot, end_dot + 1):
                                if dot_num > 99:
                                    break
                                positions.append(f"{prefix}{relay_type}{base_num}-{sub_start}.{dot_num}")
                    else:
                        # Пробуем стандартный диапазон без дефиса
                        match_start = re.search(r'^(\d*)(KCC|KLP|KLB|KL?B?|K?)(\d+)(?:\.(\d+))?$', start_str, re.IGNORECASE)
                        match_end = re.search(r'^(\d*)(KCC|KLP|KLB|KL?B?|K?)(\d+)(?:\.(\d+))?$', end_str, re.IGNORECASE)
                        
                        if match_start and match_end:
                            prefix_start = match_start.group(1) or ''
                            prefix_end = match_end.group(1) or ''
                            relay_type = match_start.group(2).upper()
                            base_num_start = int(match_start.group(3))
                            base_num_end = int(match_end.group(3))
                            dot_start = int(match_start.group(4)) if match_start.group(4) else 0
                            dot_end = int(match_end.group(4)) if match_end.group(4) else 0
                            
                            prefix = prefix_start if prefix_start else prefix_end
                            
                            # Проверяем, есть ли точки в диапазоне
                            has_dots = dot_start > 0 or dot_end > 0
                            
                            for base_num in range(base_num_start, base_num_end + 1):
                                if not has_dots:
                                    # Если точек нет, просто добавляем базовый номер
                                    positions.append(f"{prefix}{relay_type}{base_num}")
                                else:
                                    if base_num == base_num_start:
                                        start_dot = dot_start
                                    else:
                                        start_dot = 1
                                    if base_num == base_num_end:
                                        end_dot = dot_end
                                    else:
                                        end_dot = 99
                                    
                                    for dot_num in range(start_dot, min(end_dot + 1, 100)):
                                        if dot_num == 0:
                                            positions.append(f"{prefix}{relay_type}{base_num}")
                                        else:
                                            positions.append(f"{prefix}{relay_type}{base_num}.{dot_num}")
            except Exception:
                continue
        else:
            # Одиночная позиция
            # Обновленное регулярное выражение для поддержки форматов с дефисом (K2-1.1)
            # Группа 1: префикс (цифры)
            # Группа 2: тип реле (KCC, KLP, KLB, KL, K)
            # Группа 3: базовый номер (может включать дефис и цифру, например 2-1)
            # Группа 4: подномер после точки
            matches = re.findall(r'(\d*)(KCC|KLP|KLB|KL?B?|K?)(\d+(?:-\d+)?)(?:\.(\d+))?', normalized_part, re.IGNORECASE)
            for match in matches:
                prefix = match[0] or ''
                relay_type = match[1].upper()
                base_num = match[2]
                dot_num = match[3]
                
                if dot_num:
                    positions.append(f"{prefix}{relay_type}{base_num}.{dot_num}")
                else:
                    positions.append(f"{prefix}{relay_type}{base_num}")

    return positions


def extract_relay_type(text):
    """Определяет тип реле по тексту"""
    if not text:
        return None, False
    
    text_lower = text.lower()
    
    # Проверка на стоп-слова
    for kw in EXCLUDED_RELAY_TYPES:
        if kw.lower() in text_lower:
            return None, False
    
    # Описательные слова, которые НЕ являются частью артикула
    descriptors = r'(?:промежуточное|вспомогательное|модульное|электромагнитное|твердотельное)\s+'
    
    # Ищем "Реле" и захватываем артикул (буквы, цифры, дефисы, точки, пробелы между словами)
    # ОСТАНОВКА: запятая, скобка, плюс или любой другой спецсимвол прервет захват
    relay_article_match = re.search(
        rf'[Рр][Ее][Лл][Ее]\s+(?:{descriptors})?([A-Za-zА-Яа-я0-9\-\.]+(?:\s+[A-Za-zА-Яа-я0-9\-\.]+)*)', 
        text
    )
    if relay_article_match:
        article = relay_article_match.group(1).strip()
        
        # НОВАЯ БЕЗОПАСНАЯ ОЧИСТКА: удаляем только " - N шт." в самом конце
        article = re.sub(r'\s*[-,;]\s*\d+\s*шт\.?\s*$', '', article, flags=re.IGNORECASE)
        
        # Проверяем, что это похоже на артикул (содержит буквы и цифры)
        if re.search(r'[A-Za-zА-Яа-яЁё]', article) and re.search(r'[0-9]', article):
            return article, True
    
    # Поиск типа реле по ключевым словам (кроме общего слова "реле")
    for keyword in RELAY_KEYWORDS:
        if keyword.lower() in text_lower and keyword.lower() != 'реле':
            return keyword, True
    
    # Если найдено обозначение KCC/KLP/KLB/KL/K, но тип не определен
    if re.search(r'\b\d*KCC\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLP\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLB\d+(?:-\d+)?(?:\.\d+)?|\b\d*KL?\d+(?:-\d+)?(?:\.\d+)?|\b\d*K\d+(?:-\d+)?(?:\.\d+)?(?!V)', text):
        return None, False
    
    return None, False


# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ АВТОМАТОВ
# ============================================================

def is_valid_position(pos):
    """Проверяет корректность позиции автомата"""
    if not pos:
        return False
    pos = str(pos).strip()
    if pos.endswith('.'):
        return False
    # Поддерживаем SF, QFD, SFD
    if not re.search(r'(SF|QFD|SFD)', pos, re.IGNORECASE):
        return False
    return bool(re.search(r'(?:SF|QFD|SFD)-?\d+(\.\d+)?', pos, re.IGNORECASE))


def extract_short_number(cell_text):
    """Извлекает короткое число из ячейки"""
    cell = cell_text.strip()
    m = re.match(r'^(\d{1,5})(?:,.*)?$', cell)
    if m:
        return m.group(1)
    m2 = re.match(r'^(\d{1,5})$', cell)
    if m2:
        return m2.group(1)
    return None


def expand_positions_breaker(text):
    """Раскрывает позиции автоматов из диапазонов"""
    if not text or text == 'nan':
        return []
    positions = []
    text = str(text).replace('\n', ' ').strip()

    if not re.search(r'(SF|QFD|SFD)', text, re.IGNORECASE):
        return []

    parts = [p.strip() for p in text.split(',')]
    for part in parts:
        if not part:
            continue
        if part.endswith('.') and '...' not in part:
            continue

        if '...' in part:
            try:
                range_parts = part.split('...')
                if len(range_parts) != 2:
                    continue
                start_str = range_parts[0].strip()
                end_str = range_parts[1].strip()

                if end_str.endswith('.'):
                    continue

                start_match = re.search(r'(\d+)$', start_str)
                end_match = re.search(r'(\d+)$', end_str)
                if not start_match or not end_match:
                    continue

                start_prefix = start_str[:start_match.start()]
                end_prefix = end_str[:end_match.start()]
                start_num = int(start_match.group(1))
                end_num = int(end_match.group(1))

                if start_prefix != end_prefix:
                    continue

                if start_num > end_num:
                    start_num, end_num = end_num, start_num

                for num in range(start_num, end_num + 1):
                    full_pos = f"{start_prefix}{num}"
                    if is_valid_position(full_pos):
                        positions.append(full_pos)
            except Exception:
                continue
        else:
            if is_valid_position(part):
                positions.append(part)

    return positions


def get_current_from_type(type_str):
    """Извлекает номинальный ток из типа автомата"""
    if not type_str or type_str == "Не определён":
        return "не указан"
    prk_match = re.search(r'ПРК\d+-\d+', type_str)
    if prk_match:
        return prk_match.group(0).split('-')[1]
    bm_match = re.search(r'OptiDin\sBM\d+-\d+C(\d+)', type_str)
    if bm_match:
        return bm_match.group(1)
    d_match = re.search(r'OptiDin\sD\d+-\d+C(\d+)', type_str)
    if d_match:
        return d_match.group(1)
    optimat_match = re.search(r'OptiMat\s[A-Za-z]-(\d+)', type_str)
    if optimat_match:
        return optimat_match.group(1)
    va_match = re.search(r'ВА\d+-(\d+\.?\d*)A', type_str)
    if va_match:
        return va_match.group(1)
    va_new = re.search(r'ВА\d+[A-Za-z0-9\-]*?-(\d+)А', type_str, re.IGNORECASE)
    if va_new:
        return va_new.group(1)
    return "не указан"


def extract_device_type(text, debug=False):
    """Определяет тип устройства (автомата) по тексту"""
    if not text:
        return None, False

    text_lower = text.lower()

    # Стоп-слова
    stop_keywords = [
        "клеммная колодка", "источник питания", "реле", "трансформатор",
        "лампа", "контроллер", "переключатель", "резистор", "модуль буферный",
        "счетчик", "предохранитель", "розетка", "контактор", "дроссель",
        "вентилятор", "термостат", "датчик", "панель", "адаптер", "держатель предохр"
    ]
    for kw in stop_keywords:
        if kw in text_lower:
            if debug:
                print(f"      ⛔ Стоп-слово '{kw}' найдено, пропускаем")
            return None, True

    # --- ШАГ 1: Сначала ищем по шаблонам моделей (приоритет) ---
    patterns = [
    r'(OptiDin\sD\d+-\d+C\d+[A-Za-z0-9\-]*)',
    r'(OptiDin\sBM\d+-\d+C\d+)',
    r'(ПРК\s*\d+-\d+)',
    r'(OptiMat\s[A-Za-z0-9\-\.]+)',
    r'(АВДТ\s*\d+[A-Z]?)',
    r'(ВА\s*\d+[^\s,]+)',
    r'(ДИФ\d+[-\w]*)'
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            if debug:
                print(f"      🔍 Шаблон сработал: '{match.group(1)}'")
            return match.group(1), False

    # --- ШАГ 2: Если шаблоны не дали результат, пробуем маркер ---
    marker_match = re.search(
        r'(?:Автоматический выключатель|Выключатель автоматический|Дифференциальный автомат|Автомат дифференциальный|АВДТ с защитой от сверхтоков)\s+(.+)',
        text, re.IGNORECASE
    )
    if marker_match:
        after_marker = marker_match.group(1).strip()
        # Пытаемся извлечь модель из after_marker с помощью тех же паттернов
        for pattern in patterns:
            match = re.search(pattern, after_marker, re.IGNORECASE)
            if match:
                if debug:
                    print(f"      ✅ Маркер найден, модель из него: '{match.group(1)}'")
                return match.group(1), False
        # Если и так не вышло, ничего не возвращаем
        if debug:
            print(f"      ⚠️ Маркер найден, но не удалось извлечь модель из: '{after_marker}'")
        return None, False

    return None, False


# ============================================================
# ОБРАБОТКА ЛИСТА EXCEL ДЛЯ РЕЛЕ
# ============================================================

def process_worksheet_relays(ws, shield_name):
    """Обрабатывает лист Excel для поиска реле"""
    results_relays = []
    data = []
    
    print(f"   📋 Обработка листа: {ws.title}")
    
    for row in ws.iter_rows(values_only=True):
        data.append([str(c).strip() if c is not None else "" for c in row])

    if not data:
        print(f"   ⚠️ Лист {ws.title} пуст!")
        return results_relays
    
    print(f"   📊 Лист содержит {len(data)} строк, {len(data[0]) if data else 0} колонок")

    pos_col_idx, name_col_idx, start_row = find_columns(data)
    
    if pos_col_idx is None:
        print("   ❌ Не найден столбец с позициями!")
        return results_relays

    if name_col_idx is None:
        print("   ⚠️ Не найден столбец с наименованиями, использую все строки для поиска")

    print(f"   📊 Использую: Поз.={pos_col_idx}, Наим.={name_col_idx}, Начало={start_row}")

    if start_row == -1:
        for i, row in enumerate(data):
            if pos_col_idx < len(row) and row[pos_col_idx]:
                start_row = i
                break
    
    if start_row == -1:
        print("   ❌ Не найдены строки с данными!")
        return results_relays

    last_known_relay = None
    relay_count = 0

    def add_result_relay(positions, relay_type):
        if not relay_type:
            return
        for pos in positions:
            results_relays.append({
                'pos': pos,
                'type': relay_type,
                'shield': shield_name
            })

    for i in range(start_row, len(data)):
        if should_stop():
            print("⏹️ Остановка по запросу пользователя")
            return results_relays
        row = data[i]
        
        if not any(row):
            continue
        
        if is_garbage_row(row):
            continue
        
        pos_text = row[pos_col_idx] if pos_col_idx < len(row) else ""
        name_text = row[name_col_idx] if name_col_idx is not None and name_col_idx < len(row) else ""
        row_text = " ".join(row)
        
        positions_relay = expand_positions_relay(pos_text, row_text)
        
        if positions_relay:
            relay_type = None
            
            # Сначала пробуем найти тип реле в текущей строке
            if name_text:
                relay_type, is_relay = extract_relay_type(name_text)
                if not is_relay:
                    relay_type = None
            
            # Если не нашли, проверяем наличие "изделие в составе" и ищем в следующих строках
            if not relay_type:
                if "изделие в составе" in row_text.lower():
                    for j in range(1, 6):  # Увеличили глубину поиска до 5 строк
                        if i + j < len(data):
                            next_row = data[i + j]
                            if not is_garbage_row(next_row):
                                next_name = next_row[name_col_idx] if name_col_idx is not None and name_col_idx < len(next_row) else ""
                                next_row_text = " ".join(next_row)
                                relay_type, is_relay = extract_relay_type(next_name or next_row_text)
                                if is_relay and relay_type:
                                    break
            
            # Если всё ещё не нашли, ищем описание реле в следующей непустой строке
            if not relay_type:
                for j in range(1, 6):  # Увеличили глубину поиска до 5 строк
                    if i + j < len(data):
                        next_row = data[i + j]
                        if not is_garbage_row(next_row):
                            next_name = next_row[name_col_idx] if name_col_idx is not None and name_col_idx < len(next_row) else ""
                            next_row_text = " ".join(next_row)
                            # Проверяем, содержит ли следующая строка слово "Реле"
                            if 'реле' in next_row_text.lower():
                                relay_type, is_relay = extract_relay_type(next_name or next_row_text)
                                if is_relay and relay_type:
                                    break
                        else:
                            continue
                    break
            
            # Если не нашли впереди, пробуем найти в тексте текущей строки
            if not relay_type:
                relay_type_candidate, is_relay_candidate = extract_relay_type(row_text)
                # Используем найденный тип только если это конкретный артикул, а не просто "реле"
                if is_relay_candidate and relay_type_candidate and relay_type_candidate != "реле":
                    relay_type = relay_type_candidate
            
            # ВАЖНО: Сначала обновляем last_known_relay, если нашли новый тип реле
            # Это нужно сделать ДО использования last_known_relay для текущих позиций
            if relay_type and relay_type != "реле":
                last_known_relay = relay_type
                print(f"   🔄 Найдено новое реле: {relay_type} (строка {i})")
            
            # Теперь добавляем позиции с найденным или последним известным типом реле
            if relay_type and relay_type != "реле":
                add_result_relay(positions_relay, relay_type)
                relay_count += len(positions_relay)
            elif last_known_relay:
                # Используем последний известный тип реле, если новый не найден
                add_result_relay(positions_relay, last_known_relay)
                relay_count += len(positions_relay)
                print(f"   📌 Использован last_known_relay: {last_known_relay} для позиций {positions_relay}")
            else:
                add_result_relay(positions_relay, "реле")
                relay_count += len(positions_relay)
        
        # Дополнительно обновляем last_known_relay, если в текущей строке найдено описание реле
        # (даже если в этой строке нет позиций реле)
        if name_text and not positions_relay:
            relay_type_check, is_relay_check = extract_relay_type(name_text)
            if is_relay_check and relay_type_check and relay_type_check != "реле":
                last_known_relay = relay_type_check
                print(f"   🔄 Обновлено last_known_relay из описания: {last_known_relay} (строка {i})")

    print(f"   ✅ Найдено реле: {relay_count}")
    return results_relays


# ============================================================
# ОБРАБОТКА ЛИСТА EXCEL ДЛЯ АВТОМАТОВ
# ============================================================

def process_worksheet_breakers(ws, debug=False):
    """Обрабатывает лист Excel для поиска вторичных автоматов"""
    results = []
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(c).strip() if c is not None else '' for c in row])

    last_known_type = None
    invalid_count = 0

    def add_result(positions, device_type):
        nonlocal invalid_count
        for pos in positions:
            if is_valid_position(pos):
                results.append({
                    'pos': pos,
                    'type': device_type,
                    'current': get_current_from_type(device_type)
                })
            else:
                invalid_count += 1

    def look_ahead_for_type(current_row_idx, allow_positions=False):
        for offset in (1, 2):
            next_idx = current_row_idx + offset - 1
            if next_idx >= len(data):
                break
            future_row = data[next_idx]
            non_empty_future = [c for c in future_row if c]
            if not non_empty_future:
                continue
            future_text = " ".join(non_empty_future)
            if not allow_positions and expand_positions_breaker(future_text):
                if debug:
                    print(f"      🔎 Заглядываем вперёд на строку {next_idx+1}, но там есть позиции – пропускаем")
                continue
            ftype, _ = extract_device_type(future_text, debug=debug)
            if ftype:
                if debug:
                    print(f"      🔎 Заглянули вперёд на строку {next_idx+1} и нашли тип: '{ftype}'")
                return ftype
        return None

    for row_idx, row in enumerate(data, start=1):
        if should_stop():
            print("⏹️ Остановка по запросу пользователя")
            return results
        non_empty = [c for c in row if c]
        if not non_empty:
            continue

        row_text = " ".join(non_empty)

        # Извлечение позиций
        positions = []
        idx = 0
        while idx < len(non_empty):
            cell = non_empty[idx]

            if re.match(r'^\d+(,\d*)?$', cell):
                idx += 1
                continue

            if '...' in cell:
                parts = cell.split('...')
                if len(parts) == 2:
                    start_part = parts[0].strip()
                    end_part = parts[1].strip()
                    if end_part.endswith('.'):
                        if idx + 1 < len(non_empty):
                            next_cell = non_empty[idx + 1]
                            number = extract_short_number(next_cell)
                            if number:
                                end_base = end_part[:-1]
                                full_range = f"{start_part}...{end_base}.{number}"
                                positions.extend(expand_positions_breaker(full_range))
                                idx += 2
                                continue
                        idx += 1
                        continue
                    else:
                        positions.extend(expand_positions_breaker(cell))
                        idx += 1
                        continue
                else:
                    idx += 1
                    continue

            # Одиночная позиция с точкой
            m1 = re.search(r'(\b(?:SF|QFD|SFD)-?\d+)\.$', cell, re.IGNORECASE)
            if m1:
                if idx + 1 < len(non_empty):
                    next_cell = non_empty[idx + 1]
                    number = extract_short_number(next_cell)
                    if number:
                        base = m1.group(1)
                        full_range = f"{base}.1...{base}.{number}"
                        positions.extend(expand_positions_breaker(full_range))
                        idx += 2
                        continue
                idx += 1
                continue

            positions.extend(expand_positions_breaker(cell))
            idx += 1

        # Определяем тип
        current_type, _ = extract_device_type(row_text, debug=debug)

        if positions:
            if current_type:
                last_known_type = current_type
                add_result(positions, current_type)
                if debug:
                    print(f"    ➕ Добавлено {len(positions)} записей с типом '{current_type}'")
            else:
                if 'изделие в составе' in row_text.lower():
                    if last_known_type:
                        add_result(positions, last_known_type)
                        if debug:
                            print(f"    🔄 Использован last_known_type: '{last_known_type}' (изделие в составе)")
                    else:
                        found_type = look_ahead_for_type(row_idx, allow_positions=True)
                        if found_type:
                            last_known_type = found_type
                            add_result(positions, found_type)
                            if debug:
                                print(f"    🔍 Найден тип вперёд (изделие в составе): '{found_type}' → добавлено")
                        else:
                            add_result(positions, "Не определён")
                            if debug:
                                print(f"    ⚠️ Тип не определён, last_known_type отсутствует, вперёд не найден → 'Не определён'")
                else:
                    if last_known_type:
                        add_result(positions, last_known_type)
                        if debug:
                            print(f"    🔄 Использован last_known_type: '{last_known_type}' (нет явного типа)")
                    else:
                        found_type = look_ahead_for_type(row_idx, allow_positions=False)
                        if found_type:
                            last_known_type = found_type
                            add_result(positions, found_type)
                            if debug:
                                print(f"    🔍 Найден тип вперёд: '{found_type}' → добавлено")
                        else:
                            add_result(positions, "Не определён")
                            if debug:
                                print(f"    ⚠️ Тип не определён, last_known_type отсутствует, вперёд не найден → 'Не определён'")
        else:
            if current_type:
                last_known_type = current_type
                if debug:
                    print(f"    💾 Запомнили тип: '{current_type}' (позиций нет)")

    if invalid_count:
        print(f"   ⚠️ Пропущено некорректных позиций: {invalid_count}")
    return results


# ============================================================
# ОБРАБОТКА EXCEL ФАЙЛА
# ============================================================

def process_extracted_excel(input_file, shield_name):
    """Обрабатывает извлеченный Excel файл для поиска реле и автоматов"""
    print(f"   📂 Открываю файл: {os.path.basename(input_file)}...")
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
    except Exception as e:
        print(f"   ❌ Ошибка при открытии файла: {e}")
        return None, None
    
    all_relays = []
    all_breakers = []

    for sheet_name in wb.sheetnames:
        if should_stop():
            print("⏹️ Остановка по запросу пользователя")
            return all_relays, all_breakers
        if sheet_name.lower() in ['сводка', 'summary', 'текст_из_pdf']:
            continue
            
        print(f"\n   📄 Обработка листа: {sheet_name}")
        ws = wb[sheet_name]
        
        # Обрабатываем для реле
        relays = process_worksheet_relays(ws, shield_name)
        if relays:
            print(f"   ✅ Найдено {len(relays)} реле на листе {sheet_name}")
            all_relays.extend(relays)
        
        # Обрабатываем для автоматов
        breakers = process_worksheet_breakers(ws, debug=False)
        if breakers:
            print(f"   ✅ Найдено {len(breakers)} автоматов на листе {sheet_name}")
            all_breakers.extend(breakers)
        
    wb.close()
    return all_relays, all_breakers


# ============================================================
# БИБЛИОТЕКА НОМИНАЛОВ
# ============================================================

def load_nominals_library(library_path=None):
    """Загружает библиотеку номинальных токов"""
    if library_path is None:
        script_folder = os.path.dirname(os.path.abspath(__file__))
        library_path = os.path.join(script_folder, "Library_nominals.xlsx")
    if not os.path.exists(library_path):
        print(f"⚠️ Библиотека номиналов не найдена: {library_path}")
        return {}
    try:
        wb = openpyxl.load_workbook(library_path, data_only=True)
        sheet = wb["Номинальные токи"] if "Номинальные токи" in wb.sheetnames else wb.active
        nominals = {}
        for row in range(2, sheet.max_row + 1):
            device_type = sheet.cell(row=row, column=2).value
            nominal = sheet.cell(row=row, column=3).value
            if device_type and nominal:
                device_type_clean = str(device_type).strip()
                if isinstance(nominal, (int, float)):
                    nominal_value = float(nominal)
                else:
                    try:
                        nominal_value = float(str(nominal).strip().replace(',', '.'))
                    except ValueError:
                        nominal_value = nominal
                nominals[device_type_clean] = nominal_value
        print(f"✅ Загружено номиналов из библиотеки: {len(nominals)}")
        return nominals
    except Exception as e:
        print(f"❌ Ошибка загрузки библиотеки: {e}")
        return {}


def match_device_type_with_library(device_type, nominals_library):
    """Сопоставляет тип устройства с библиотекой номиналов"""
    if not nominals_library:
        return None
    device_clean = device_type.strip()
    if device_clean in nominals_library:
        return nominals_library[device_clean]
    for lib_type, nominal in nominals_library.items():
        if lib_type.startswith(device_clean) or device_clean in lib_type:
            return nominal
    return None


# ============================================================
# СОХРАНЕНИЕ РЕЗУЛЬТАТОВ
# ============================================================

def save_relays_to_xlsx(relays_list, output_file_path, voltage_map=None):
    """Сохраняет реле в Excel файл
    voltage_map: словарь {relay_type: voltage} для заполнения номинального напряжения
    """
    from openpyxl.styles import Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реле (KL, K, KLB)"
    ws.append([
        "№ п/п",
        "Позиционное обозначение",
        "Тип реле",
        "Номинальное напряжение, В",
        "Напряжение срабатывания",
        "Напряжение Возврата",
        "Контакты НО",
        "Контакты НЗ",
        "Заключение"
    ])
    for i, relay in enumerate(relays_list, 1):
        row_num = ws.max_row + 1
        relay_type = relay.get('type', '')
        pos = relay.get('pos', '')
        voltage = ''
        if voltage_map and relay_type in voltage_map:
            voltage = voltage_map.get(relay_type, '')
        
        # Записываем ячейки по одной, обрабатывая значения с '='
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=pos)
        
        # Тип реле - если начинается с '=', добавляем апостроф
        if isinstance(relay_type, str) and relay_type.startswith('='):
            cell_c = ws.cell(row=row_num, column=3, value="'" + relay_type)
            cell_c.number_format = '@'
            cell_c.alignment = Alignment(horizontal='left')
        else:
            ws.cell(row=row_num, column=3, value=relay_type)
        
        # Напряжение - если начинается с '=', добавляем апостроф
        if isinstance(voltage, str) and voltage.startswith('='):
            cell_d = ws.cell(row=row_num, column=4, value="'" + voltage)
            cell_d.number_format = '@'
            cell_d.alignment = Alignment(horizontal='left')
        else:
            ws.cell(row=row_num, column=4, value=voltage)
        
        # Остальные колонки
        ws.cell(row=row_num, column=5, value='')
        ws.cell(row=row_num, column=6, value='')
        ws.cell(row=row_num, column=7, value='')
        ws.cell(row=row_num, column=8, value='')
        ws.cell(row=row_num, column=9, value='Соотв.')
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    wb.save(output_file_path)
    return output_file_path

def save_breakers_to_xlsx(breakers_list, output_file_path):
    """Сохраняет автоматы в Excel файл"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Автоматы вторичных цепей"
    ws.append(["№ п/п", "Позиция (SF/QFD/SFD)", "Тип автомата", "Номинальный ток, А"])
    for i, breaker in enumerate(breakers_list, 1):
        ws.append([i, breaker.get('pos', ''), breaker.get('type', ''), breaker.get('current', 'не указан')])
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    wb.save(output_file_path)
    return output_file_path

# ============================================================
# РАБОТА С ФАЙЛОМ ПАМЯТИ НАПРЯЖЕНИЙ (Relay_voltage.xlsx)
# ============================================================

def load_relay_voltage_memory(file_path):
    """Загружает известные напряжения реле из файла памяти."""
    if not os.path.exists(file_path):
        return {}
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        voltages = {}
        for row in ws.iter_rows(min_row=2, values_only=True):  # Начинаем со 2 строки (без заголовка)
            if row[0] and row[1]:
                relay_type = str(row[0]).strip()
                voltage = str(row[1]).strip()
                # Удаляем апостроф в начале, если он есть (Excel добавляет его для текста с '=')
                if voltage.startswith("'"):
                    voltage = voltage[1:]
                if voltage:  # Записываем только если напряжение не пустое
                    voltages[relay_type] = voltage
        wb.close()
        print(f"   🧠 Загружено из памяти напряжений: {len(voltages)} записей")
        return voltages
    except Exception as e:
        print(f"   ⚠️ Ошибка чтения файла памяти напряжений: {e}")
        return {}


def save_relay_voltage_memory(file_path, new_voltages):
    """Обновляет или создает файл памяти напряжений реле."""
    # Сначала читаем то, что уже есть в файле
    existing_voltages = load_relay_voltage_memory(file_path)
    
    # Обновляем старые данные новыми (если тип реле уже был, напряжение перезапишется)
    existing_voltages.update(new_voltages)
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Память напряжений"
        
        # Заголовки
        ws.append(["Тип реле", "Номинальное напряжение"])
        
        # Записываем все данные, используя апостроф для значений с '='
        from openpyxl.styles import Alignment
        
        for r_type, voltage in existing_voltages.items():
            row_num = ws.max_row + 1
            
            # Тип реле
            ws.cell(row=row_num, column=1, value=r_type)
            
            # Напряжение - если начинается с '=', добавляем апостроф как префикс
            if isinstance(voltage, str) and voltage.startswith('='):
                # Добавляем апостроф перед значением - Excel распознает это как текст
                cell = ws.cell(row=row_num, column=2, value="'" + voltage)
                # Устанавливаем текстовый формат и выравнивание по левому краю
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left')
            else:
                ws.cell(row=row_num, column=2, value=voltage)
            
        # Красивые ширины колонок
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        
        wb.save(file_path)
        print(f"   💾 Память напряжений сохранена в: {os.path.basename(file_path)}")
    except PermissionError:
        print(f"   ❌ Ошибка: Файл {os.path.basename(file_path)} открыт в другой программе. Закройте его и попробуйте снова.")
    except Exception as e:
        print(f"   ❌ Ошибка сохранения файла памяти: {e}")

# ============================================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================================

def main():
    """Главная функция обработки PDF файлов"""
    args = sys.argv[1:]
    is_gui_mode = '--from-gui' in args

    for i, arg in enumerate(args):
        if arg == '--output-dir' and i + 1 < len(args):
            output_dir = args[i + 1]
            args = args[:i] + args[i+2:]
            break

    if '--from-gui' in args:
        args.remove('--from-gui')

    if not args:
        print("❌ Ошибка: не указан путь к файлу!")
        if not is_gui_mode:
            input("\nНажмите Enter для выхода...")
        return

    pdf_path = args[0]
    pdf_file = Path(pdf_path)
    if not pdf_file.exists():
        print(f"❌ Ошибка: Файл '{pdf_path}' не найден!")
        if not is_gui_mode:
            input("Нажмите Enter для выхода...")
        return

    print(f"\n📁 Обработка файла: {pdf_file.name}")
    print("=" * 50)

    # Проверяем кэш перед началом обработки
    from cache_manager import get_cache_manager, init_cache_manager
    cache_manager = init_cache_manager()
    cached_excel = cache_manager.get_cached_excel(pdf_file)
    
    if cached_excel and cached_excel.exists():
        print(f"✅ Найдены закэшированные данные: {cached_excel.name}")
        temp_excel_path = cached_excel
    else:
        print("Шаг 1: Извлечение текста и таблиц из PDF...")
        text_data = extract_text_from_pdf(pdf_path)
        tables_data = extract_tables_from_pdf(pdf_path)

        if not text_data and not tables_data:
            print("❌ Не удалось извлечь данные из PDF файла.")
            if not is_gui_mode:
                input("Нажмите Enter для выхода...")
            return

        # Создаем временный файл
        temp_excel_path = pdf_file.parent / f"{pdf_file.stem}_extracted_temp.xlsx"
        print(f"   Сохранение промежуточных данных: {temp_excel_path.name}")

        with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as writer:
            if text_data:
                df_text = pd.DataFrame(text_data)
                df_text.to_excel(writer, sheet_name='Текст_из_PDF', index=False)
            if tables_data:
                for i, table_info in enumerate(tables_data):
                    sheet_name = f"Таблица_{table_info['Страница']}_{table_info['Таблица']}"
                    sheet_name = sheet_name[:31]
                    df_table = pd.DataFrame(table_info['Данные'])
                    df_table.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        # Сохраняем в кэш
        temp_excel_path = cache_manager.save_to_cache(pdf_file, temp_excel_path)

    print("\nШаг 2: Поиск реле и автоматов в извлеченных данных...")
    all_relays, all_breakers = process_extracted_excel(str(temp_excel_path), pdf_file.stem)
    
    print(f"   Временный файл {temp_excel_path.name} сохранен в кэше.")
    
    # Определяем папку вывода
    if '--output-dir' in sys.argv:
        idx = sys.argv.index('--output-dir')
        if idx + 1 < len(sys.argv):
            output_folder = Path(sys.argv[idx + 1])
        else:
            output_folder = pdf_file.parent
    else:
        output_folder = pdf_file.parent
    
    output_folder.mkdir(parents=True, exist_ok=True)

    # ==========================================
    # ИЗМЕНЕННЫЙ БЛОК СОХРАНЕНИЯ РЕЛЕ
    # ==========================================
    if all_relays:
        relay_output_file = output_folder / f"{pdf_file.stem}_Реле.xlsx"
        if relay_output_file.exists():
            base_name = relay_output_file.stem
            ext = relay_output_file.suffix
            counter = 1
            while relay_output_file.exists():
                relay_output_file = output_folder / f"{base_name}_{counter}{ext}"
                counter += 1
        
        # 1. Получаем уникальные типы реле
        unique_relays = get_unique_relay_types(all_relays)
        
        # 2. Путь к файлу памяти в папке со скриптом
        script_dir = os.path.dirname(os.path.abspath(__file__))
        memory_file_path = os.path.join(script_dir, "Relay_voltage.xlsx")
        
        # 3. Загружаем уже известные напряжения из файла
        final_voltage_map = load_relay_voltage_memory(memory_file_path)
        
        # 4. Ищем реле, которых НЕТ в файле памяти
        unknown_relays = [r for r in unique_relays if r['type'] not in final_voltage_map]
        
        # 5. Показываем окно только для новых реле
        if unknown_relays:
            print(f"\n💡 Найдено {len(unknown_relays)} новых типов реле. Открытие окна выбора напряжения...")
            new_voltages = show_relay_voltage_dialog(unknown_relays)
            
            if new_voltages is not None:  # Нажали "ОК"
                final_voltage_map.update(new_voltages)
                save_relay_voltage_memory(memory_file_path, new_voltages) # Сохраняем в файл
                print(f"   ✅ Выбрано и сохранено в память новых напряжений: {len(new_voltages)}")
            else:  # Нажали "Отмена"
                print("   ⚠️ Выбор напряжения отменён для новых реле.")
        else:
            print("\n✅ Напряжения для всех найденных типов реле загружены из файла памяти (Relay_voltage.xlsx).")
        
        # 6. Сохраняем итоговый файл со спецификацией (передаем full map)
        save_relays_to_xlsx(all_relays, str(relay_output_file), voltage_map=final_voltage_map)
        
        print(f"\n✅ Реле сохранены: {relay_output_file.name}")
        print(f"   📊 Всего реле: {len(all_relays)}")
    else:
        print("\n⚠️ Реле не найдены в извлеченных данных.")
    # ==========================================

    # Сохраняем автоматы
    if all_breakers:
        # Загружаем библиотеку номиналов
        print("\n📚 Загрузка библиотеки номинальных токов...")
        nominals_library = load_nominals_library()
        
        # Применяем библиотеку номиналов
        for result in all_breakers:
            if nominals_library:
                nominal_from_lib = match_device_type_with_library(result['type'], nominals_library)
                if nominal_from_lib is not None:
                    result['current'] = nominal_from_lib
        
        breaker_output_file = output_folder / f"{pdf_file.stem}_Автоматические_выключатели_вторичных_цепей.xlsx"
        if breaker_output_file.exists():
            base_name = breaker_output_file.stem
            ext = breaker_output_file.suffix
            counter = 1
            while breaker_output_file.exists():
                breaker_output_file = output_folder / f"{base_name}_{counter}{ext}"
                counter += 1

        save_breakers_to_xlsx(all_breakers, str(breaker_output_file))
        print(f"\n✅ Автоматы сохранены: {breaker_output_file.name}")
        print(f"   📊 Всего автоматов: {len(all_breakers)}")
        
        # Пример найденных автоматов
        print("\n📋 ПРИМЕР найденных автоматов (первые 5):")
        for result in all_breakers[:5]:
            print(f"   {result['pos']} → {result['type']} ({result['current']}A)")
        
        # Отсутствующие типы
        if nominals_library:
            missing_types = set()
            for result in all_breakers:
                if result['type'] != "Не определён" and not match_device_type_with_library(result['type'], nominals_library):
                    missing_types.add(result['type'])
            if missing_types:
                print(f"\n⚠️ Типы, отсутствующие в библиотеке ({len(missing_types)}):")
                for t in sorted(missing_types):
                    print(f"   - {t}")
    else:
        print("\n⚠️ Автоматы не найдены в извлеченных данных.")

    print("\n" + "=" * 50)
    if not is_gui_mode:
        input("Нажмите Enter для выхода...")


if __name__ == "__main__":
    main()