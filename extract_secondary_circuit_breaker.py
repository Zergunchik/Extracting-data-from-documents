import pdfplumber
import pandas as pd
import openpyxl
import re
import os
import sys
from pathlib import Path
import builtins

def should_stop():
    """Проверяет, запросил ли пользователь остановку через GUI."""
    check = getattr(builtins, '_gui_stop_check', None)
    return check() if check else False


# ------------------------- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ -------------------------
def is_valid_position(pos):
    if not pos:
        return False
    pos = str(pos).strip()
    if pos.endswith('.'):
        return False
    # Поддерживаем SF, QFD, SFD
    if not re.search(r'(SF|QFD|SFD)', pos, re.IGNORECASE):
        return False
    return bool(re.search(r'(?:SF|QFD|SFD)-?\d+(\.\d+)?', pos, re.IGNORECASE))


def extract_short_number(cell_text):
    cell = cell_text.strip()
    m = re.match(r'^(\d{1,5})(?:,.*)?$', cell)
    if m:
        return m.group(1)
    m2 = re.match(r'^(\d{1,5})$', cell)
    if m2:
        return m2.group(1)
    return None


def expand_positions(text):
    if not text or text == 'nan':
        return []
    positions = []
    text = str(text).replace('\n', ' ').strip()

    if not re.search(r'(SF|QFD|SFD)', text, re.IGNORECASE):
        return []

    parts = [p.strip() for p in text.split(',')]
    for part in parts:
        if not part:
            continue
        if part.endswith('.') and '...' not in part:
            continue

        if '...' in part:
            try:
                range_parts = part.split('...')
                if len(range_parts) != 2:
                    continue
                start_str = range_parts[0].strip()
                end_str = range_parts[1].strip()

                if end_str.endswith('.'):
                    continue

                start_match = re.search(r'(\d+)$', start_str)
                end_match = re.search(r'(\d+)$', end_str)
                if not start_match or not end_match:
                    continue

                start_prefix = start_str[:start_match.start()]
                end_prefix = end_str[:end_match.start()]
                start_num = int(start_match.group(1))
                end_num = int(end_match.group(1))

                if start_prefix != end_prefix:
                    continue

                if start_num > end_num:
                    start_num, end_num = end_num, start_num

                for num in range(start_num, end_num + 1):
                    full_pos = f"{start_prefix}{num}"
                    if is_valid_position(full_pos):
                        positions.append(full_pos)
            except Exception:
                continue
        else:
            if is_valid_position(part):
                positions.append(part)

    return positions


def get_current_from_type(type_str):
    if not type_str or type_str == "Не определён":
        return "не указан"
    prk_match = re.search(r'ПРК\d+-(\d+)', type_str)
    if prk_match:
        return prk_match.group(1)
    bm_match = re.search(r'OptiDin\sBM\d+-\d+C(\d+)', type_str)
    if bm_match:
        return bm_match.group(1)
    d_match = re.search(r'OptiDin\sD\d+-\d+C(\d+)', type_str)
    if d_match:
        return d_match.group(1)
    optimat_match = re.search(r'OptiMat\s[A-Za-z]-(\d+)', type_str)
    if optimat_match:
        return optimat_match.group(1)
    va_match = re.search(r'ВА\d+-(\d+\.?\d*)A', type_str)
    if va_match:
        return va_match.group(1)
    va_new = re.search(r'ВА\d+[A-Za-z0-9\-]*?-(\d+)А', type_str, re.IGNORECASE)
    if va_new:
        return va_new.group(1)
    return "не указан"


# ------------------------- ОПРЕДЕЛЕНИЕ ТИПА УСТРОЙСТВА -------------------------
def extract_device_type(text, debug=False):
    if not text:
        return None, False

    text_lower = text.lower()

    # Стоп-слова
    stop_keywords = [
        "клеммная колодка", "источник питания", "реле", "трансформатор",
        "лампа", "контроллер", "переключатель", "резистор", "модуль буферный",
        "счетчик", "предохранитель", "розетка", "контактор", "дроссель",
        "вентилятор", "термостат", "датчик", "панель", "адаптер", "держатель предохр"
    ]
    for kw in stop_keywords:
        if kw in text_lower:
            if debug:
                print(f"      ⛔ Стоп-слово '{kw}' найдено, пропускаем")
            return None, True

    # --- ШАГ 1: Сначала ищем по шаблонам моделей (приоритет) ---
    patterns = [
    r'(OptiDin\sD\d+-\d+C\d+[A-Za-z0-9\-]*)',
    r'(OptiDin\sBM\d+-\d+C\d+)',
    r'(ПРК\s*\d+-\d+)',
    r'(OptiMat\s[A-Za-z0-9\-\.]+)',
    r'(АВДТ\s*\d+[A-Z]?)',
    r'(ВА\s*\d+[^\s,]+)',
    r'(ДИФ\d+[-\w]*)'
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            if debug:
                print(f"      🔍 Шаблон сработал: '{match.group(1)}'")
            return match.group(1), False

    # --- ШАГ 2: Если шаблоны не дали результат, пробуем маркер ---
    marker_match = re.search(
        r'(?:Автоматический выключатель|Выключатель автоматический|Дифференциальный автомат|Автомат дифференциальный|АВДТ с защитой от сверхтоков)\s+(.+)',
        text, re.IGNORECASE
    )
    if marker_match:
        after_marker = marker_match.group(1).strip()
        # Пытаемся извлечь модель из after_marker с помощью тех же паттернов
        for pattern in patterns:
            match = re.search(pattern, after_marker, re.IGNORECASE)
            if match:
                if debug:
                    print(f"      ✅ Маркер найден, модель из него: '{match.group(1)}'")
                return match.group(1), False
        # Если и так не вышло, ничего не возвращаем
        if debug:
            print(f"      ⚠️ Маркер найден, но не удалось извлечь модель из: '{after_marker}'")
        return None, False

    return None, False


# ------------------------- ОБРАБОТКА ОДНОГО ЛИСТА -------------------------
def process_worksheet(ws, debug=False):
    results = []
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(c).strip() if c is not None else '' for c in row])

    last_known_type = None
    invalid_count = 0

    def add_result(positions, device_type):
        nonlocal invalid_count
        for pos in positions:
            if is_valid_position(pos):
                results.append({
                    'pos': pos,
                    'type': device_type,
                    'current': get_current_from_type(device_type)
                })
            else:
                invalid_count += 1

    def look_ahead_for_type(current_row_idx, allow_positions=False):
        for offset in (1, 2):
            next_idx = current_row_idx + offset - 1
            if next_idx >= len(data):
                break
            future_row = data[next_idx]
            non_empty_future = [c for c in future_row if c]
            if not non_empty_future:
                continue
            future_text = " ".join(non_empty_future)
            if not allow_positions and expand_positions(future_text):
                if debug:
                    print(f"      🔎 Заглядываем вперёд на строку {next_idx+1}, но там есть позиции – пропускаем")
                continue
            ftype, _ = extract_device_type(future_text, debug=debug)
            if ftype:
                if debug:
                    print(f"      🔎 Заглянули вперёд на строку {next_idx+1} и нашли тип: '{ftype}'")
                return ftype
        return None

    for row_idx, row in enumerate(data, start=1):
        if should_stop():
            print("⏹️ Остановка по запросу пользователя")
            return results
        non_empty = [c for c in row if c]
        if not non_empty:
            continue

        row_text = " ".join(non_empty)

        # Извлечение позиций
        positions = []
        idx = 0
        while idx < len(non_empty):
            cell = non_empty[idx]

            if re.match(r'^\d+(,\d*)?$', cell):
                idx += 1
                continue

            if '...' in cell:
                parts = cell.split('...')
                if len(parts) == 2:
                    start_part = parts[0].strip()
                    end_part = parts[1].strip()
                    if end_part.endswith('.'):
                        if idx + 1 < len(non_empty):
                            next_cell = non_empty[idx + 1]
                            number = extract_short_number(next_cell)
                            if number:
                                end_base = end_part[:-1]
                                full_range = f"{start_part}...{end_base}.{number}"
                                positions.extend(expand_positions(full_range))
                                idx += 2
                                continue
                        idx += 1
                        continue
                    else:
                        positions.extend(expand_positions(cell))
                        idx += 1
                        continue
                else:
                    idx += 1
                    continue

            # Одиночная позиция с точкой
            m1 = re.search(r'(\b(?:SF|QFD|SFD)-?\d+)\.$', cell, re.IGNORECASE)
            if m1:
                if idx + 1 < len(non_empty):
                    next_cell = non_empty[idx + 1]
                    number = extract_short_number(next_cell)
                    if number:
                        base = m1.group(1)
                        full_range = f"{base}.1...{base}.{number}"
                        positions.extend(expand_positions(full_range))
                        idx += 2
                        continue
                idx += 1
                continue

            positions.extend(expand_positions(cell))
            idx += 1

        # Определяем тип
        current_type, _ = extract_device_type(row_text, debug=debug)

        if positions:
            if current_type:
                last_known_type = current_type
                add_result(positions, current_type)
                if debug:
                    print(f"    ➕ Добавлено {len(positions)} записей с типом '{current_type}'")
            else:
                if 'изделие в составе' in row_text.lower():
                    if last_known_type:
                        add_result(positions, last_known_type)
                        if debug:
                            print(f"    🔄 Использован last_known_type: '{last_known_type}' (изделие в составе)")
                    else:
                        found_type = look_ahead_for_type(row_idx, allow_positions=True)
                        if found_type:
                            last_known_type = found_type
                            add_result(positions, found_type)
                            if debug:
                                print(f"    🔍 Найден тип вперёд (изделие в составе): '{found_type}' → добавлено")
                        else:
                            add_result(positions, "Не определён")
                            if debug:
                                print(f"    ⚠️ Тип не определён, last_known_type отсутствует, вперёд не найден → 'Не определён'")
                else:
                    if last_known_type:
                        add_result(positions, last_known_type)
                        if debug:
                            print(f"    🔄 Использован last_known_type: '{last_known_type}' (нет явного типа)")
                    else:
                        found_type = look_ahead_for_type(row_idx, allow_positions=False)
                        if found_type:
                            last_known_type = found_type
                            add_result(positions, found_type)
                            if debug:
                                print(f"    🔍 Найден тип вперёд: '{found_type}' → добавлено")
                        else:
                            add_result(positions, "Не определён")
                            if debug:
                                print(f"    ⚠️ Тип не определён, last_known_type отсутствует, вперёд не найден → 'Не определён'")
        else:
            if current_type:
                last_known_type = current_type
                if debug:
                    print(f"    💾 Запомнили тип: '{current_type}' (позиций нет)")

    if invalid_count:
        print(f"   ⚠️ Пропущено некорректных позиций: {invalid_count}")
    return results


# ------------------------- ОБРАБОТКА ИЗВЛЕЧЁННОГО EXCEL -------------------------
def process_extracted_excel(input_file, debug=False):
    print(f"Открываю файл: {input_file}...")
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
    except Exception as e:
        print(f"Ошибка при открытии файла: {e}")
        return None

    all_results = []
    for sheet_name in wb.sheetnames:
        if should_stop():
            print("⏹️ Остановка по запросу пользователя")
            return all_results
        if sheet_name.lower() in ['сводка', 'summary', 'текст_из_pdf']:
            continue
        print(f"Обработка листа: {sheet_name}")
        ws = wb[sheet_name]
        sheet_data = process_worksheet(ws, debug=debug)
        all_results.extend(sheet_data)
        print(f"   Найдено устройств на листе: {len(sheet_data)}")
    wb.close()
    return all_results


# ------------------------- ИЗВЛЕЧЕНИЕ ИЗ PDF -------------------------
def extract_text_from_pdf(pdf_path):
    extracted_data = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if text:
                    lines = text.split('\n')
                    for line_num, line in enumerate(lines, 1):
                        if line.strip():
                            extracted_data.append({
                                'Страница': page_num,
                                'Строка': line_num,
                                'Текст': line.strip()
                            })
    except Exception as e:
        print(f"Ошибка при чтении PDF: {e}")
        return None
    return extracted_data


def extract_tables_from_pdf(pdf_path):
    all_tables = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()
                for table_num, table in enumerate(tables, 1):
                    if table:
                        clean_table = []
                        for row in table:
                            clean_row = [str(cell).strip() if cell is not None else '' for cell in row]
                            if any(clean_row):
                                clean_table.append(clean_row)
                        if clean_table:
                            all_tables.append({
                                'Страница': page_num,
                                'Таблица': table_num,
                                'Данные': clean_table
                            })
    except Exception as e:
        print(f"Ошибка при извлечении таблиц: {e}")
    return all_tables


# ------------------------- БИБЛИОТЕКА НОМИНАЛОВ -------------------------
def load_nominals_library(library_path=None):
    if library_path is None:
        script_folder = os.path.dirname(os.path.abspath(__file__))
        library_path = os.path.join(script_folder, "Library_nominals.xlsx")
    if not os.path.exists(library_path):
        print(f"⚠️ Библиотека номиналов не найдена: {library_path}")
        return {}
    try:
        wb = openpyxl.load_workbook(library_path, data_only=True)
        sheet = wb["Номинальные токи"] if "Номинальные токи" in wb.sheetnames else wb.active
        nominals = {}
        for row in range(2, sheet.max_row + 1):
            device_type = sheet.cell(row=row, column=2).value
            nominal = sheet.cell(row=row, column=3).value
            if device_type and nominal:
                device_type_clean = str(device_type).strip()
                # Сохраняем число как есть, без приведения к int
                if isinstance(nominal, (int, float)):
                    nominal_value = float(nominal)  # оставляем как float
                else:
                    # строка – пытаемся преобразовать, заменив запятую на точку
                    try:
                        nominal_value = float(str(nominal).strip().replace(',', '.'))
                    except ValueError:
                        nominal_value = nominal  # оставляем как есть (например, "не указан")
                nominals[device_type_clean] = nominal_value
        print(f"✅ Загружено номиналов из библиотеки: {len(nominals)}")
        return nominals
    except Exception as e:
        print(f"❌ Ошибка загрузки библиотеки: {e}")
        return {}


def match_device_type_with_library(device_type, nominals_library):
    if not nominals_library:
        return None
    device_clean = device_type.strip()
    if device_clean in nominals_library:
        return nominals_library[device_clean]
    for lib_type, nominal in nominals_library.items():
        if lib_type.startswith(device_clean) or device_clean in lib_type:
            return nominal
    return None


# ------------------------- СОХРАНЕНИЕ РЕЗУЛЬТАТОВ -------------------------
def save_breakers_to_xlsx(breakers_list, output_file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Автоматы вторичных цепей"
    ws.append(["№ п/п", "Позиция (SF/QFD/SFD)", "Тип автомата", "Номинальный ток, А"])
    for i, breaker in enumerate(breakers_list, 1):
        ws.append([i, breaker.get('pos', ''), breaker.get('type', ''), breaker.get('current', 'не указан')])
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    wb.save(output_file_path)
    return output_file_path


# ------------------------- ГЛАВНАЯ ФУНКЦИЯ -------------------------
def main():
    args = sys.argv[1:]
    is_gui_mode = '--from-gui' in args
    output_dir = None
    skip_cache_delete = '--skip-cache-delete' in args  # Новый флаг для сохранения кэша

    for i, arg in enumerate(args):
        if arg == '--output-dir' and i + 1 < len(args):
            output_dir = args[i + 1]
            args = args[:i] + args[i+2:]
            break

    if '--from-gui' in args:
        args.remove('--from-gui')
    
    if '--skip-cache-delete' in args:
        args.remove('--skip-cache-delete')

    if not args:
        print("❌ Ошибка: не указан путь к файлу!")
        if not is_gui_mode:
            input("\nНажмите Enter для выхода...")
        return

    pdf_path = args[0]
    pdf_file = Path(pdf_path)
    if not pdf_file.exists():
        print(f"❌ Ошибка: Файл '{pdf_path}' не найден!")
        if not is_gui_mode:
            input("Нажмите Enter для выхода...")
        return

    print(f"\n📁 Обработка файла: {pdf_file.name}")
    print("=" * 50)

    # Проверяем кэш перед началом обработки
    from cache_manager import get_cache_manager, init_cache_manager
    cache_manager = init_cache_manager()  # Инициализируем с логированием
    cached_excel = cache_manager.get_cached_excel(pdf_file)
    
    if cached_excel and cached_excel.exists():
        logger.info(f"✅ Найдены закэшированные данные: {cached_excel.name}")
        print(f"✅ Найдены закэшированные данные: {cached_excel.name}")
        temp_excel_path = cached_excel
    else:
        logger.info("Шаг 1: Извлечение текста и таблиц из PDF...")
        print("Шаг 1: Извлечение текста и таблиц из PDF...")
        text_data = extract_text_from_pdf(pdf_path)
        tables_data = extract_tables_from_pdf(pdf_path)

        if not text_data and not tables_data:
            logger.error("❌ Не удалось извлечь данные из PDF файла.")
            print("❌ Не удалось извлечь данные из PDF файла.")
            if not is_gui_mode:
                input("Нажмите Enter для выхода...")
            return

        # Создаем временный файл в той же папке где PDF (будет перемещен в кэш)
        temp_excel_path = pdf_file.parent / f"{pdf_file.stem}_extracted_temp.xlsx"
        logger.info(f"   Сохранение промежуточных данных: {temp_excel_path.name}")
        print(f"   Сохранение промежуточных данных: {temp_excel_path.name}")

        with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as writer:
            if text_data:
                df_text = pd.DataFrame(text_data)
                df_text.to_excel(writer, sheet_name='Текст_из_PDF', index=False)
            if tables_data:
                for i, table_info in enumerate(tables_data):
                    sheet_name = f"Таблица_{table_info['Страница']}_{table_info['Таблица']}"
                    sheet_name = sheet_name[:31]
                    df_table = pd.DataFrame(table_info['Данные'])
                    df_table.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        # Сохраняем в кэш через save_to_cache (перемещает файл в .pdf_cache)
        temp_excel_path = cache_manager.save_to_cache(pdf_file, temp_excel_path)

    print("\nШаг 2: Поиск автоматических выключателей в извлеченных данных...")
    # По умолчанию без отладочного вывода (debug=False)
    all_results = process_extracted_excel(str(temp_excel_path), debug=False)
    
    # Временный файл хранится в кэше для повторного использования - не удаляем
    print(f"   Временный файл {temp_excel_path.name} сохранен в кэше.")
    
    # Загружаем библиотеку номиналов только после успешного извлечения данных
    print("\n📚 Загрузка библиотеки номинальных токов...")
    nominals_library = load_nominals_library()
    print("=" * 50)

    if all_results:
        # Применяем библиотеку номиналов
        for result in all_results:
            if nominals_library:
                nominal_from_lib = match_device_type_with_library(result['type'], nominals_library)
                if nominal_from_lib is not None:
                    result['current'] = nominal_from_lib
                else:
                    # Оставляем значение из get_current_from_type, если оно не "не указан"
                    if result['current'] == "не указан":
                        pass  # так и останется "не указан"
            # Вывод в консоль для каждого устройства можно убрать или оставить краткий
            # Оставим только итоговую статистику

        if output_dir:
            output_folder = Path(output_dir)
        else:
            output_folder = pdf_file.parent
        output_folder.mkdir(parents=True, exist_ok=True)

        output_file = output_folder / f"{pdf_file.stem}_Автоматические_выключатели_вторичных_цепей.xlsx"
        if output_file.exists():
            base_name = output_file.stem
            ext = output_file.suffix
            counter = 1
            while output_file.exists():
                output_file = output_folder / f"{base_name}_{counter}{ext}"
                counter += 1

        save_breakers_to_xlsx(all_results, str(output_file))

        print(f"\n✅ Готово! Всего найдено устройств: {len(all_results)}")
        print(f"📁 Результат сохранен в: {output_file.name}")

        print("\n📋 ПРИМЕР найденных автоматов (первые 5):")
        for result in all_results[:5]:
            print(f"   {result['pos']} → {result['type']} ({result['current']}A)")

        if nominals_library:
            missing_types = set()
            for result in all_results:
                if result['type'] != "Не определён" and not match_device_type_with_library(result['type'], nominals_library):
                    missing_types.add(result['type'])
            if missing_types:
                print(f"\n⚠️ Типы, отсутствующие в библиотеке ({len(missing_types)}):")
                for t in sorted(missing_types):
                    print(f"   - {t}")
    else:
        print("⚠️ Устройства не найдены в извлеченных данных.")

    print("\n" + "=" * 50)
    if not is_gui_mode:
        input("Нажмите Enter для выхода...")


if __name__ == "__main__":
    main()