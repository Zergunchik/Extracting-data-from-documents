# stage2_extract_breakers.py - Этап 2: Извлечение автоматических выключателей
"""
Этап 2 конвейера обработки спецификаций.
Извлекает данные об автоматических выключателях из промежуточного Excel файла.

Входные данные: путь к Excel файлу с извлеченными данными из PDF
Выходные данные: список словарей с данными автоматов и путь к итоговому Excel файлу
"""

import openpyxl
import re
import os
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple


def is_valid_position(pos: Any) -> bool:
    """Проверяет корректность позиции автомата"""
    if not pos:
        return False
    pos = str(pos).strip()
    if pos.endswith('.'):
        return False
    # Поддерживаем SF, QFD, SFD
    if not re.search(r'(SF|QFD|SFD)', pos, re.IGNORECASE):
        return False
    return bool(re.search(r'(?:SF|QFD|SFD)-?\d+(\.\d+)?', pos, re.IGNORECASE))


def extract_short_number(cell_text: str) -> Optional[str]:
    """Извлекает короткое число из ячейки"""
    cell = cell_text.strip()
    m = re.match(r'^(\d{1,5})(?:,.*)?$', cell)
    if m:
        return m.group(1)
    m2 = re.match(r'^(\d{1,5})$', cell)
    if m2:
        return m2.group(1)
    return None


def expand_positions_breaker(text: Any) -> List[str]:
    """Раскрывает позиции автоматов из диапазонов"""
    if not text or text == 'nan':
        return []
    positions = []
    text = str(text).replace('\n', ' ').strip()

    if not re.search(r'(SF|QFD|SFD)', text, re.IGNORECASE):
        return []

    parts = [p.strip() for p in text.split(',')]
    for part in parts:
        if not part:
            continue
        if part.endswith('.') and '...' not in part:
            continue

        if '...' in part:
            try:
                range_parts = part.split('...')
                if len(range_parts) != 2:
                    continue
                start_str = range_parts[0].strip()
                end_str = range_parts[1].strip()

                if end_str.endswith('.'):
                    continue

                start_match = re.search(r'(\d+)$', start_str)
                end_match = re.search(r'(\d+)$', end_str)
                if not start_match or not end_match:
                    continue

                start_prefix = start_str[:start_match.start()]
                end_prefix = end_str[:end_match.start()]
                start_num = int(start_match.group(1))
                end_num = int(end_match.group(1))

                if start_prefix != end_prefix:
                    continue

                if start_num > end_num:
                    start_num, end_num = end_num, start_num

                for num in range(start_num, end_num + 1):
                    full_pos = f"{start_prefix}{num}"
                    if is_valid_position(full_pos):
                        positions.append(full_pos)
            except Exception:
                continue
        else:
            if is_valid_position(part):
                positions.append(part)

    return positions


def get_current_from_type(type_str: str) -> str:
    """Извлекает номинальный ток из типа автомата"""
    if not type_str or type_str == "Не определён":
        return "не указан"
    
    prk_match = re.search(r'ПРК\d+-\d+', type_str)
    if prk_match:
        return prk_match.group(0).split('-')[1]
    
    bm_match = re.search(r'OptiDin\sBM\d+-\d+C(\d+)', type_str)
    if bm_match:
        return bm_match.group(1)
    
    d_match = re.search(r'OptiDin\sD\d+-\d+C(\d+)', type_str)
    if d_match:
        return d_match.group(1)
    
    optimat_match = re.search(r'OptiMat\s[A-Za-z]-(\d+)', type_str)
    if optimat_match:
        return optimat_match.group(1)
    
    va_match = re.search(r'ВА\d+-(\d+\.?\d*)A', type_str)
    if va_match:
        return va_match.group(1)
    
    va_new = re.search(r'ВА\d+[A-Za-z0-9\-]*?-(\d+)А', type_str, re.IGNORECASE)
    if va_new:
        return va_new.group(1)
    
    return "не указан"


def extract_device_type(text: str, debug: bool = False) -> Tuple[Optional[str], bool]:
    """
    Определяет тип устройства (автомата) по тексту.
    
    Returns:
        (тип_устройства, is_stop_word) - тип или None, флаг стоп-слова
    """
    if not text:
        return None, False

    text_lower = text.lower()

    # Стоп-слова
    stop_keywords = [
        "клеммная колодка", "источник питания", "реле", "трансформатор",
        "лампа", "контроллер", "переключатель", "резистор", "модуль буферный",
        "счетчик", "предохранитель", "розетка", "контактор", "дроссель",
        "вентилятор", "термостат", "датчик", "панель", "адаптер", "держатель предохр"
    ]
    for kw in stop_keywords:
        if kw in text_lower:
            if debug:
                print(f"      ⛔ Стоп-слово '{kw}' найдено, пропускаем")
            return None, True

    # --- ШАГ 1: Сначала ищем по шаблонам моделей (приоритет) ---
    patterns = [
        r'(OptiDin\sD\d+-\d+C\d+[A-Za-z0-9\-]*)',
        r'(OptiDin\sBM\d+-\d+C\d+)',
        r'(ПРК\s*\d+-\d+)',
        r'(OptiMat\s[A-Za-z0-9\-\.]+)',
        r'(АВДТ\s*\d+[A-Z]?)',
        r'(ВА\s*\d+[^\s,]+)',
        r'(ДИФ\d+[-\w]*)'
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            if debug:
                print(f"      🔍 Шаблон сработал: '{match.group(1)}'")
            return match.group(1), False

    # --- ШАГ 2: Если шаблоны не дали результат, пробуем маркер ---
    marker_match = re.search(
        r'(?:Автоматический выключатель|Выключатель автоматический|Дифференциальный автомат|Автомат дифференциальный|АВДТ с защитой от сверхтоков)\s+(.+)',
        text, re.IGNORECASE
    )
    if marker_match:
        after_marker = marker_match.group(1).strip()
        # Пытаемся извлечь модель из after_marker с помощью тех же паттернов
        for pattern in patterns:
            match = re.search(pattern, after_marker, re.IGNORECASE)
            if match:
                if debug:
                    print(f"      ✅ Маркер найден, модель из него: '{match.group(1)}'")
                return match.group(1), False
        # Если и так не вышло, ничего не возвращаем
        if debug:
            print(f"      ⚠️ Маркер найден, но не удалось извлечь модель из: '{after_marker}'")
        return None, False

    return None, False


def process_worksheet_breakers(ws, debug: bool = False) -> List[Dict[str, Any]]:
    """
    Обрабатывает лист Excel для поиска вторичных автоматов.
    
    Args:
        ws: Лист Excel
        debug: Режим отладки
        
    Returns:
        Список словарей с данными автоматов
    """
    results = []
    data = []
    
    for row in ws.iter_rows(values_only=True):
        data.append([str(c).strip() if c is not None else '' for c in row])

    last_known_type = None
    invalid_count = 0

    def add_result(positions: List[str], device_type: str):
        nonlocal invalid_count
        for pos in positions:
            if is_valid_position(pos):
                results.append({
                    'pos': pos,
                    'type': device_type,
                    'current': get_current_from_type(device_type)
                })
            else:
                invalid_count += 1

    def look_ahead_for_type(current_row_idx: int, allow_positions: bool = False) -> Optional[str]:
        for offset in (1, 2):
            next_idx = current_row_idx + offset - 1
            if next_idx >= len(data):
                break
            future_row = data[next_idx]
            non_empty_future = [c for c in future_row if c]
            if not non_empty_future:
                continue
            future_text = " ".join(non_empty_future)
            if not allow_positions and expand_positions_breaker(future_text):
                if debug:
                    print(f"      🔎 Заглядываем вперёд на строку {next_idx+1}, но там есть позиции – пропускаем")
                continue
            ftype, _ = extract_device_type(future_text, debug=debug)
            if ftype:
                if debug:
                    print(f"      🔎 Заглянули вперёд на строку {next_idx+1} и нашли тип: '{ftype}'")
                return ftype
        return None

    for row_idx, row in enumerate(data, start=1):
        non_empty = [c for c in row if c]
        if not non_empty:
            continue

        row_text = " ".join(non_empty)

        # Извлечение позиций
        positions = []
        idx = 0
        while idx < len(non_empty):
            cell = non_empty[idx]

            if re.match(r'^\d+(,\d*)?$', cell):
                idx += 1
                continue

            if '...' in cell:
                parts = cell.split('...')
                if len(parts) == 2:
                    start_part = parts[0].strip()
                    end_part = parts[1].strip()
                    if end_part.endswith('.'):
                        if idx + 1 < len(non_empty):
                            next_cell = non_empty[idx + 1]
                            number = extract_short_number(next_cell)
                            if number:
                                end_base = end_part[:-1]
                                full_range = f"{start_part}...{end_base}.{number}"
                                positions.extend(expand_positions_breaker(full_range))
                                idx += 2
                                continue
                        idx += 1
                        continue
                    else:
                        positions.extend(expand_positions_breaker(cell))
                        idx += 1
                        continue
                else:
                    idx += 1
                    continue

            # Одиночная позиция с точкой
            m1 = re.search(r'(\b(?:SF|QFD|SFD)-?\d+)\.$', cell, re.IGNORECASE)
            if m1:
                if idx + 1 < len(non_empty):
                    next_cell = non_empty[idx + 1]
                    number = extract_short_number(next_cell)
                    if number:
                        base = m1.group(1)
                        full_range = f"{base}.1...{base}.{number}"
                        positions.extend(expand_positions_breaker(full_range))
                        idx += 2
                        continue
                idx += 1
                continue

            positions.extend(expand_positions_breaker(cell))
            idx += 1

        # Определяем тип
        current_type, _ = extract_device_type(row_text, debug=debug)

        if positions:
            if current_type:
                last_known_type = current_type
                add_result(positions, current_type)
                if debug:
                    print(f"    ➕ Добавлено {len(positions)} записей с типом '{current_type}'")
            else:
                if 'изделие в составе' in row_text.lower():
                    if last_known_type:
                        add_result(positions, last_known_type)
                        if debug:
                            print(f"    🔄 Использован last_known_type: '{last_known_type}' (изделие в составе)")
                    else:
                        found_type = look_ahead_for_type(row_idx, allow_positions=True)
                        if found_type:
                            last_known_type = found_type
                            add_result(positions, found_type)
                            if debug:
                                print(f"    🔍 Найден тип вперёд (изделие в составе): '{found_type}' → добавлено")
                        else:
                            add_result(positions, "Не определён")
                            if debug:
                                print(f"    ⚠️ Тип не определён, last_known_type отсутствует, вперёд не найден → 'Не определён'")
                else:
                    if last_known_type:
                        add_result(positions, last_known_type)
                        if debug:
                            print(f"    🔄 Использован last_known_type: '{last_known_type}' (нет явного типа)")
                    else:
                        found_type = look_ahead_for_type(row_idx, allow_positions=False)
                        if found_type:
                            last_known_type = found_type
                            add_result(positions, found_type)
                            if debug:
                                print(f"    🔍 Найден тип вперёд: '{found_type}' → добавлено")
                        else:
                            add_result(positions, "Не определён")
                            if debug:
                                print(f"    ⚠️ Тип не определён, last_known_type отсутствует, вперёд не найден → 'Не определён'")
        else:
            if current_type:
                last_known_type = current_type
                if debug:
                    print(f"    💾 Запомнили тип: '{current_type}' (позиций нет)")

    if invalid_count:
        print(f"   ⚠️ Пропущено некорректных позиций: {invalid_count}")
    
    return results


def load_nominals_library(library_path: Optional[str] = None) -> Dict[str, Any]:
    """Загружает библиотеку номинальных токов"""
    if library_path is None:
        script_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        library_path = os.path.join(script_folder, "Library_nominals.xlsx")
    
    if not os.path.exists(library_path):
        print(f"⚠️ Библиотека номиналов не найдена: {library_path}")
        return {}
    
    try:
        wb = openpyxl.load_workbook(library_path, data_only=True)
        sheet = wb["Номинальные токи"] if "Номинальные токи" in wb.sheetnames else wb.active
        nominals = {}
        
        for row in range(2, sheet.max_row + 1):
            device_type = sheet.cell(row=row, column=2).value
            nominal = sheet.cell(row=row, column=3).value
            
            if device_type and nominal:
                device_type_clean = str(device_type).strip()
                if isinstance(nominal, (int, float)):
                    nominal_value = float(nominal)
                else:
                    try:
                        nominal_value = float(str(nominal).strip().replace(',', '.'))
                    except ValueError:
                        nominal_value = nominal
                nominals[device_type_clean] = nominal_value
        
        print(f"✅ Загружено номиналов из библиотеки: {len(nominals)}")
        return nominals
    except Exception as e:
        print(f"❌ Ошибка загрузки библиотеки: {e}")
        return {}


def match_device_type_with_library(device_type: str, nominals_library: Dict[str, Any]) -> Optional[Any]:
    """Сопоставляет тип устройства с библиотекой номиналов"""
    if not nominals_library:
        return None
    
    device_clean = device_type.strip()
    if device_clean in nominals_library:
        return nominals_library[device_clean]
    
    for lib_type, nominal in nominals_library.items():
        if lib_type.startswith(device_clean) or device_clean in lib_type:
            return nominal
    
    return None


def save_breakers_to_xlsx(breakers_list: List[Dict[str, Any]], output_file_path: str) -> str:
    """Сохраняет автоматы в Excel файл"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Автоматы вторичных цепей"
    ws.append(["№ п/п", "Позиция (SF/QFD/SFD)", "Тип автомата", "Номинальный ток, А"])
    
    for i, breaker in enumerate(breakers_list, 1):
        ws.append([
            i, 
            breaker.get('pos', ''), 
            breaker.get('type', ''), 
            breaker.get('current', 'не указан')
        ])
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    wb.save(output_file_path)
    return output_file_path


def extract_breakers_from_excel(
    input_excel_path: str,
    output_dir: Optional[str] = None,
    apply_nominals: bool = True,
    pdf_name: Optional[str] = None
) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """
    Главная функция этапа 2: извлечение автоматов из Excel файла.
    
    Args:
        input_excel_path: Путь к Excel файлу с извлеченными данными
        output_dir: Директория для сохранения результата (опционально)
        apply_nominals: Применять библиотеку номиналов
        pdf_name: Имя исходного PDF файла для формирования имени выходного файла
        
    Returns:
        (список_автоматов, путь_к_файлу) или (список_автоматов, None) если автоматы не найдены
    """
    print(f"\nШаг 2: Поиск автоматов в извлеченных данных...")
    print(f"   📂 Открываю файл: {os.path.basename(input_excel_path)}...")
    
    try:
        wb = openpyxl.load_workbook(input_excel_path, data_only=True)
    except Exception as e:
        print(f"   ❌ Ошибка при открытии файла: {e}")
        return [], None

    all_breakers = []

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ['сводка', 'summary', 'текст_из_pdf']:
            continue

        print(f"\n   📄 Обработка листа: {sheet_name}")
        ws = wb[sheet_name]

        breakers = process_worksheet_breakers(ws, debug=False)
        if breakers:
            print(f"   ✅ Найдено {len(breakers)} автоматов на листе {sheet_name}")
            all_breakers.extend(breakers)

    wb.close()

    if not all_breakers:
        print("\n⚠️ Автоматы не найдены в извлеченных данных.")
        return all_breakers, None

    # Применяем библиотеку номиналов
    if apply_nominals:
        print("\n📚 Загрузка библиотеки номинальных токов...")
        nominals_library = load_nominals_library()

        for result in all_breakers:
            if nominals_library:
                nominal_from_lib = match_device_type_with_library(result['type'], nominals_library)
                if nominal_from_lib is not None:
                    result['current'] = nominal_from_lib

    # Определяем путь для выходного файла
    if output_dir:
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
    else:
        output_path = Path(input_excel_path).parent

    # Формируем имя файла
    if pdf_name:
        base_name = pdf_name
    else:
        base_name = Path(input_excel_path).stem.replace('_extracted_temp', '')

    breaker_output_file = output_path / f"{base_name}_Автоматические_выключатели.xlsx"
    
    if breaker_output_file.exists():
        base = breaker_output_file.stem
        ext = breaker_output_file.suffix
        counter = 1
        while breaker_output_file.exists():
            breaker_output_file = output_path / f"{base}_{counter}{ext}"
            counter += 1

    save_breakers_to_xlsx(all_breakers, str(breaker_output_file))
    print(f"\n✅ Автоматы сохранены: {breaker_output_file.name}")
    print(f"   📊 Всего автоматов: {len(all_breakers)}")

    # Пример найденных автоматов
    print("\n📋 ПРИМЕР найденных автоматов (первые 5):")
    for result in all_breakers[:5]:
        print(f"   {result['pos']} → {result['type']} ({result['current']}A)")

    # Отсутствующие типы
    if apply_nominals and nominals_library:
        missing_types = set()
        for result in all_breakers:
            if result['type'] != "Не определён" and not match_device_type_with_library(result['type'], nominals_library):
                missing_types.add(result['type'])
        if missing_types:
            print(f"\n⚠️ Типы, отсутствующие в библиотеке ({len(missing_types)}):")
            for t in sorted(missing_types):
                print(f"   - {t}")

    return all_breakers, str(breaker_output_file)


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python stage2_extract_breakers.py <путь_к_Excel> [выходная_директория]")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else None
    
    breakers, output_file = extract_breakers_from_excel(excel_path, output_dir)
    if breakers:
        print(f"\n✅ Этап 2 завершен. Найдено {len(breakers)} автоматов.")
        if output_file:
            print(f"   Результат сохранен: {output_file}")
    else:
        print("\n⚠️ Автоматы не найдены")
