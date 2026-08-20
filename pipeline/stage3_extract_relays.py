# stage3_extract_relays.py - Этап 3: Извлечение реле
"""
Этап 3 конвейера обработки спецификаций.
Извлекает данные о реле из промежуточного Excel файла.

Входные данные: путь к Excel файлу с извлеченными данными из PDF
Выходные данные: список словарей с данными реле и путь к итоговому Excel файлу
"""

import openpyxl
import re
import os
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

# Списки ключевых слов для реле
NEGATIVE_KEYWORDS = [
    'компрессор', 'компресс',
    'насос', 'pump',
    'chiller',
    'кондиционер внешний',
    'двигатель',
    'мотор'
]

RELAY_KEYWORDS = [
    'реле',
    'промежуточное реле',
    'реле промежуточное',
    'RKE4CO024LT',
    'RKE4CO730LT'
]

EXCLUDED_RELAY_TYPES = [
    'реле напряжения',
    'реле контроля напряжения',
    'KV',
    'KVZ',
    'контроль напряжения'
]


def normalize_text(text: Any) -> str:
    """Нормализация текста (замена русских букв на латинские)"""
    if not text:
        return text
    text = str(text)
    text = text.replace('К', 'K').replace('к', 'k')
    text = text.replace('М', 'M').replace('м', 'm')
    return text


def has_negative_keywords(text: Any) -> bool:
    """Проверяет, содержит ли текст негативные ключевые слова"""
    if not text:
        return False
    text_lower = text.lower()
    for neg_word in NEGATIVE_KEYWORDS:
        if neg_word in text_lower:
            return True
    return False


def is_garbage_row(row: List[Any]) -> bool:
    """Проверяет, является ли строка мусорной (служебной информацией)"""
    if not row:
        return True

    row_text = " ".join([str(cell).strip() for cell in row if cell])

    garbage_patterns = [
        r'^атад$',
        r'^и$',
        r'^ьсипдоП$',
        r'^\.лбуд$',
        r'^№\.внИ$',
        r'^№\.вни$',
        r'^\.мазВ$',
        r'^\.лдоп$',
        r'^Копировал:',
        r'^Изм\.',
        r'^Лист$',
        r'^№ докум\.',
        r'^Подпись$',
        r'^Дата$',
        r'^Формат',
        r'^K\-BLOCK\-',
    ]

    for pattern in garbage_patterns:
        if re.search(pattern, row_text, re.IGNORECASE):
            return True

    # Если строка содержит только короткие слова - возможно мусор
    words = row_text.split()
    if len(words) <= 4:
        short_words = [w for w in words if len(w) <= 3]
        if len(short_words) / max(len(words), 1) > 0.7:
            return True

    return False


def find_columns(data: List[List[str]]) -> Tuple[Optional[int], Optional[int], int]:
    """
    Находит столбцы с позициями и наименованиями
    Возвращает: (pos_col_idx, name_col_idx, start_row_idx)
    """
    if not data or not data[0]:
        return None, None, -1

    num_cols = len(data[0])
    pos_col_idx = None
    name_col_idx = None
    start_row = -1

    # Первый проход: ищем заголовки в первых 20 строках
    for row_idx, row in enumerate(data[:20]):
        if is_garbage_row(row):
            continue

        row_text_lower = " ".join([str(cell).lower() for cell in row if cell])

        if not pos_col_idx:
            for col_idx, cell in enumerate(row):
                if cell:
                    cell_str = str(cell).lower().strip()
                    if 'поз' in cell_str or 'обознач' in cell_str:
                        pos_col_idx = col_idx
                        break

        if not name_col_idx:
            for col_idx, cell in enumerate(row):
                if cell:
                    cell_str = str(cell).lower().strip()
                    if 'наименование' in cell_str or 'наим' in cell_str or 'описание' in cell_str:
                        name_col_idx = col_idx
                        break

        if pos_col_idx is not None and name_col_idx is not None:
            start_row = row_idx + 1
            break

    # Второй проход: если заголовки не найдены, ищем по данным
    if pos_col_idx is None or name_col_idx is None:
        print("   🔍 Ищу столбцы по содержимому данных...")

        for col_idx in range(num_cols):
            relay_positions = 0
            relay_names = 0

            for row in data[20:60]:
                if col_idx < len(row) and row[col_idx]:
                    cell_str = normalize_text(str(row[col_idx]))

                    # Обновленное регулярное выражение с поддержкой KCC, KLP, KLB, KL, K
                    if re.search(r'\b\d*KCC\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLP\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLB\d+(?:-\d+)?(?:\.\d+)?|\b\d*KL?\d+(?:-\d+)?(?:\.\d+)?|\b\d*K\d+(?:-\d+)?(?:\.\d+)?(?!V)', cell_str):
                        relay_positions += 1

                    cell_lower = str(row[col_idx]).lower()
                    for keyword in RELAY_KEYWORDS:
                        if keyword.lower() in cell_lower:
                            relay_names += 1
                            break

            if relay_positions > 3 and pos_col_idx is None:
                pos_col_idx = col_idx
                print(f"   ✅ Найдена колонка с позициями: {col_idx}")

            if relay_names > 2 and name_col_idx is None:
                name_col_idx = col_idx
                print(f"   ✅ Найдена колонка с наименованиями: {col_idx}")

    # Если не нашли стартовую строку, ищем первую строку с данными
    if start_row == -1 and pos_col_idx is not None:
        for row_idx, row in enumerate(data):
            if row_idx < 10:
                continue
            if pos_col_idx < len(row) and row[pos_col_idx]:
                cell_str = normalize_text(str(row[pos_col_idx]))
                # Обновленное регулярное выражение с поддержкой KCC, KLP, KLB, KL, K
                if re.search(r'\b\d*KCC\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLP\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLB\d+(?:-\d+)?(?:\.\d+)?|\b\d*KL?\d+(?:-\d+)?(?:\.\d+)?|\b\d*K\d+(?:-\d+)?(?:\.\d+)?(?!V)', cell_str):
                    start_row = row_idx
                    break

    print(f"   📊 Результат: Поз.={pos_col_idx}, Наим.={name_col_idx}, Начало={start_row}")
    return pos_col_idx, name_col_idx, start_row


def expand_positions_relay(text: Any, full_row_text: Optional[str] = None) -> List[str]:
    """
    Извлекает позиции с реле (KL, K, KLB)
    Поддерживает: K1, K1.1, KL1, KLB1, K1...K10, K1.1...K1.9, 2K1, 4K1, 4K2, K1-4.1...K1-4.6
    """
    if not text or text == 'nan' or text == '':
        return []

    if full_row_text and has_negative_keywords(full_row_text):
        return []

    positions = []
    text = str(text).replace('\n', ' ').strip()
    normalized_text = normalize_text(text)

    # Исключаем реле напряжения
    if re.search(r'\bKV\d+(?:\.\d+)?|\bKVZ\d+(?:\.\d+)?', normalized_text, re.IGNORECASE):
        return []
    if re.search(r'реле\s+напряжени[яю]', normalized_text, re.IGNORECASE):
        return []

    # Проверяем наличие реле (обновленное регулярное выражение с поддержкой цифр перед буквами и дефисов)
    if not re.search(
        r'\b\d*KCC\d+(?:-\d+)?(?:\.\d+)?|'   # KCC1, KCC2-1, KCC3.1
        r'\b\d*KLP\d+(?:-\d+)?(?:\.\d+)?|'   # KLP1, KLP2-1, KLP3.1
        r'\b\d*KLB\d+(?:-\d+)?(?:\.\d+)?|'   # KLB1, KLB2-1, KLB3.1
        r'\b\d*KL?\d+(?:-\d+)?(?:\.\d+)?|'   # KL1, KL2-1 (и K1 если L нет)
        r'\b\d*K\d+(?:-\d+)?(?:\.\d+)?(?!V)', # K1, K2-1 (но не KV1)
        normalized_text
    ):
        return []

    parts = [p.strip() for p in text.split(',') if p.strip()]

    for part in parts:
        if not part:
            continue

        normalized_part = normalize_text(part)

        if re.search(r'\bKV\d+(?:\.\d+)?|\bKVZ\d+(?:\.\d+)?', normalized_part, re.IGNORECASE):
            continue
        if re.search(r'реле\s+напряжени[яю]', normalized_part, re.IGNORECASE):
            continue

        # Обновленное регулярное выражение с поддержкой KCC, KLP, KLB, KL, K
        if not re.search(r'\b\d*KCC\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLP\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLB\d+(?:-\d+)?(?:\.\d+)?|\b\d*KL?\d+(?:-\d+)?(?:\.\d+)?|\b\d*K\d+(?:-\d+)?(?:\.\d+)?(?!V)', normalized_part):
            continue

        # Обработка диапазонов с дефисом (например, K1-4.1...K1-4.6)
        if '...' in normalized_part:
            try:
                range_parts = normalized_part.split('...')
                if len(range_parts) == 2:
                    start_str = range_parts[0].strip()
                    end_str = range_parts[1].strip()

                    # Проверяем, является ли это диапазоном с дефисом
                    match_start = re.search(r'^(\d*)(KCC|KLP|KLB|KL?B?|K?)(\d+)-(\d+)\.(\d+)$', start_str, re.IGNORECASE)
                    match_end = re.search(r'^(\d*)(KCC|KLP|KLB|KL?B?|K?)(\d+)-(\d+)\.(\d+)$', end_str, re.IGNORECASE)

                    if match_start and match_end:
                        prefix_start = match_start.group(1) or ''
                        prefix_end = match_end.group(1) or ''
                        relay_type = match_start.group(2).upper()
                        base_num_start = int(match_start.group(3))
                        base_num_end = int(match_end.group(3))
                        sub_start = int(match_start.group(4))
                        sub_end = int(match_end.group(4))
                        dot_start = int(match_start.group(5))
                        dot_end = int(match_end.group(5))

                        prefix = prefix_start if prefix_start else prefix_end

                        # Генерируем все позиции в диапазоне
                        for base_num in range(base_num_start, base_num_end + 1):
                            if base_num == base_num_start:
                                start_dot = dot_start
                            else:
                                start_dot = 1
                            if base_num == base_num_end:
                                end_dot = dot_end
                            else:
                                end_dot = 99

                            for dot_num in range(start_dot, end_dot + 1):
                                if dot_num > 99:
                                    break
                                positions.append(f"{prefix}{relay_type}{base_num}-{sub_start}.{dot_num}")
                    else:
                        # Пробуем стандартный диапазон без дефиса
                        match_start = re.search(r'^(\d*)(KCC|KLP|KLB|KL?B?|K?)(\d+)(?:\.(\d+))?$', start_str, re.IGNORECASE)
                        match_end = re.search(r'^(\d*)(KCC|KLP|KLB|KL?B?|K?)(\d+)(?:\.(\d+))?$', end_str, re.IGNORECASE)

                        if match_start and match_end:
                            prefix_start = match_start.group(1) or ''
                            prefix_end = match_end.group(1) or ''
                            relay_type = match_start.group(2).upper()
                            base_num_start = int(match_start.group(3))
                            base_num_end = int(match_end.group(3))
                            dot_start = int(match_start.group(4)) if match_start.group(4) else 0
                            dot_end = int(match_end.group(4)) if match_end.group(4) else 0

                            prefix = prefix_start if prefix_start else prefix_end

                            # Проверяем, есть ли точки в диапазоне
                            has_dots = dot_start > 0 or dot_end > 0

                            for base_num in range(base_num_start, base_num_end + 1):
                                if not has_dots:
                                    # Если точек нет, просто добавляем базовый номер
                                    positions.append(f"{prefix}{relay_type}{base_num}")
                                else:
                                    if base_num == base_num_start:
                                        start_dot = dot_start
                                    else:
                                        start_dot = 1
                                    if base_num == base_num_end:
                                        end_dot = dot_end
                                    else:
                                        end_dot = 99

                                    for dot_num in range(start_dot, min(end_dot + 1, 100)):
                                        if dot_num == 0:
                                            positions.append(f"{prefix}{relay_type}{base_num}")
                                        else:
                                            positions.append(f"{prefix}{relay_type}{base_num}.{dot_num}")
            except Exception:
                continue
        else:
            # Одиночная позиция
            # Обновленное регулярное выражение для поддержки форматов с дефисом (K2-1.1)
            matches = re.findall(r'(\d*)(KCC|KLP|KLB|KL?B?|K?)(\d+(?:-\d+)?)(?:\.(\d+))?', normalized_part, re.IGNORECASE)
            for match in matches:
                prefix = match[0] or ''
                relay_type = match[1].upper()
                base_num = match[2]
                dot_num = match[3]

                if dot_num:
                    positions.append(f"{prefix}{relay_type}{base_num}.{dot_num}")
                else:
                    positions.append(f"{prefix}{relay_type}{base_num}")

    return positions


def extract_relay_type(text: Any) -> Tuple[Optional[str], bool]:
    """
    Определяет тип реле по тексту.
    
    Returns:
        (тип_реле, is_relay) - тип или None, флаг что это реле
    """
    if not text:
        return None, False

    text_lower = text.lower()

    # Проверка на стоп-слова
    for kw in EXCLUDED_RELAY_TYPES:
        if kw.lower() in text_lower:
            return None, False

    # Описательные слова, которые НЕ являются частью артикула
    descriptors = r'(?:промежуточное|вспомогательное|модульное|электромагнитное|твердотельное)\s+'

    # Ищем "Реле" и захватываем артикул (буквы, цифры, дефисы, точки, пробелы между словами)
    relay_article_match = re.search(
        rf'[Рр][Ее][Лл][Ее]\s+(?:{descriptors})?([A-Za-zА-Яа-я0-9\-\.]+(?:\s+[A-Za-zА-Яа-я0-9\-\.]+)*)',
        text
    )
    if relay_article_match:
        article = relay_article_match.group(1).strip()

        # НОВАЯ БЕЗОПАСНАЯ ОЧИСТКА: удаляем только " - N шт." в самом конце
        article = re.sub(r'\s*[-,;]\s*\d+\s*шт\.?\s*$', '', article, flags=re.IGNORECASE)

        # Проверяем, что это похоже на артикул (содержит буквы и цифры)
        if re.search(r'[A-Za-zА-Яа-яЁё]', article) and re.search(r'[0-9]', article):
            return article, True

    # Поиск типа реле по ключевым словам (кроме общего слова "реле")
    for keyword in RELAY_KEYWORDS:
        if keyword.lower() in text_lower and keyword.lower() != 'реле':
            return keyword, True

    # Если найдено обозначение KCC/KLP/KLB/KL/K, но тип не определен
    if re.search(r'\b\d*KCC\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLP\d+(?:-\d+)?(?:\.\d+)?|\b\d*KLB\d+(?:-\d+)?(?:\.\d+)?|\b\d*KL?\d+(?:-\d+)?(?:\.\d+)?|\b\d*K\d+(?:-\d+)?(?:\.\d+)?(?!V)', text):
        return None, False

    return None, False


def process_worksheet_relays(ws, shield_name: str) -> List[Dict[str, Any]]:
    """
    Обрабатывает лист Excel для поиска реле.
    
    Args:
        ws: Лист Excel
        shield_name: Имя щита (из имени файла)
        
    Returns:
        Список словарей с данными реле
    """
    results_relays = []
    data = []

    print(f"   📋 Обработка листа: {ws.title}")

    for row in ws.iter_rows(values_only=True):
        data.append([str(c).strip() if c is not None else "" for c in row])

    if not data:
        print(f"   ⚠️ Лист {ws.title} пуст!")
        return results_relays

    print(f"   📊 Лист содержит {len(data)} строк, {len(data[0]) if data else 0} колонок")

    pos_col_idx, name_col_idx, start_row = find_columns(data)

    if pos_col_idx is None:
        print("   ❌ Не найден столбец с позициями!")
        return results_relays

    if name_col_idx is None:
        print("   ⚠️ Не найден столбец с наименованиями, использую все строки для поиска")

    print(f"   📊 Использую: Поз.={pos_col_idx}, Наим.={name_col_idx}, Начало={start_row}")

    if start_row == -1:
        for i, row in enumerate(data):
            if pos_col_idx < len(row) and row[pos_col_idx]:
                start_row = i
                break

    if start_row == -1:
        print("   ❌ Не найдены строки с данными!")
        return results_relays

    last_known_relay = None
    relay_count = 0

    def add_result_relay(positions: List[str], relay_type: str):
        if not relay_type:
            return
        for pos in positions:
            results_relays.append({
                'pos': pos,
                'type': relay_type,
                'shield': shield_name
            })

    for i in range(start_row, len(data)):
        row = data[i]

        if not any(row):
            continue

        if is_garbage_row(row):
            continue

        pos_text = row[pos_col_idx] if pos_col_idx < len(row) else ""
        name_text = row[name_col_idx] if name_col_idx is not None and name_col_idx < len(row) else ""
        row_text = " ".join(row)

        positions_relay = expand_positions_relay(pos_text, row_text)

        if positions_relay:
            relay_type = None

            # Сначала пробуем найти тип реле в текущей строке
            if name_text:
                relay_type, is_relay = extract_relay_type(name_text)
                if not is_relay:
                    relay_type = None

            # Если не нашли, проверяем наличие "изделие в составе" и ищем в следующих строках
            if not relay_type:
                if "изделие в составе" in row_text.lower():
                    for j in range(1, 6):  # Увеличили глубину поиска до 5 строк
                        if i + j < len(data):
                            next_row = data[i + j]
                            if not is_garbage_row(next_row):
                                next_name = next_row[name_col_idx] if name_col_idx is not None and name_col_idx < len(next_row) else ""
                                next_row_text = " ".join(next_row)
                                relay_type, is_relay = extract_relay_type(next_name or next_row_text)
                                if is_relay and relay_type:
                                    break

            # Если всё ещё не нашли, ищем описание реле в следующей непустой строке
            if not relay_type:
                for j in range(1, 6):  # Увеличили глубину поиска до 5 строк
                    if i + j < len(data):
                        next_row = data[i + j]
                        if not is_garbage_row(next_row):
                            next_name = next_row[name_col_idx] if name_col_idx is not None and name_col_idx < len(next_row) else ""
                            next_row_text = " ".join(next_row)
                            # Проверяем, содержит ли следующая строка слово "Реле"
                            if 'реле' in next_row_text.lower():
                                relay_type, is_relay = extract_relay_type(next_name or next_row_text)
                                if is_relay and relay_type:
                                    break
                        else:
                            continue
                    break

            # Если не нашли впереди, пробуем найти в тексте текущей строки
            if not relay_type:
                relay_type_candidate, is_relay_candidate = extract_relay_type(row_text)
                # Используем найденный тип только если это конкретный артикул, а не просто "реле"
                if is_relay_candidate and relay_type_candidate and relay_type_candidate != "реле":
                    relay_type = relay_type_candidate

            # ВАЖНО: Сначала обновляем last_known_relay, если нашли новый тип реле
            if relay_type and relay_type != "реле":
                last_known_relay = relay_type
                print(f"   🔄 Найдено новое реле: {relay_type} (строка {i})")

            # Теперь добавляем позиции с найденным или последним известным типом реле
            if relay_type and relay_type != "реле":
                add_result_relay(positions_relay, relay_type)
                relay_count += len(positions_relay)
            elif last_known_relay:
                # Используем последний известный тип реле, если новый не найден
                add_result_relay(positions_relay, last_known_relay)
                relay_count += len(positions_relay)
                print(f"   📌 Использован last_known_relay: {last_known_relay} для позиций {positions_relay}")
            else:
                add_result_relay(positions_relay, "реле")
                relay_count += len(positions_relay)

        # Дополнительно обновляем last_known_relay, если в текущей строке найдено описание реле
        if name_text and not positions_relay:
            relay_type_check, is_relay_check = extract_relay_type(name_text)
            if is_relay_check and relay_type_check and relay_type_check != "реле":
                last_known_relay = relay_type_check
                print(f"   🔄 Обновлено last_known_relay из описания: {last_known_relay} (строка {i})")

    print(f"   ✅ Найдено реле: {relay_count}")
    return results_relays


def get_unique_relay_types(relays_list: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Возвращает список уникальных типов реле с их позициями"""
    unique = {}
    for relay in relays_list:
        relay_type = relay.get('type', '')
        pos = relay.get('pos', '')
        if relay_type not in unique:
            unique[relay_type] = []
        unique[relay_type].append(pos)

    return [{'type': t, 'positions': p} for t, p in unique.items()]


def save_relays_to_xlsx(
    relays_list: List[Dict[str, Any]], 
    output_file_path: str, 
    voltage_map: Optional[Dict[str, str]] = None
) -> str:
    """
    Сохраняет реле в Excel файл.
    
    Args:
        relays_list: Список словарей с данными реле
        output_file_path: Путь к выходному файлу
        voltage_map: Словарь {relay_type: voltage} для заполнения номинального напряжения
        
    Returns:
        Путь к сохраненному файлу
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реле (KL, K, KLB)"
    ws.append([
        "№ п/п",
        "Позиционное обозначение",
        "Тип реле",
        "Номинальное напряжение, В",
        "Напряжение срабатывания",
        "Напряжение Возврата",
        "Контакты НО",
        "Контакты НЗ",
        "Заключение"
    ])
    
    for i, relay in enumerate(relays_list, 1):
        relay_type = relay.get('type', '')
        voltage = ''
        if voltage_map and relay_type in voltage_map:
            voltage = voltage_map.get(relay_type, '')
        ws.append([
            i,
            relay.get('pos', ''),
            relay_type,
            voltage,
            '',
            '',
            '',
            '',
            'Соотв.'
        ])

    # Принудительно указываем Excel, что ячейки с '=' это текст, а не формулы
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.data_type = 's'

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    wb.save(output_file_path)
    return output_file_path


def load_relay_voltage_memory(file_path: str) -> Dict[str, str]:
    """Загружает известные напряжения реле из файла памяти."""
    if not os.path.exists(file_path):
        return {}
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        voltages = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                relay_type = str(row[0]).strip()
                voltage = str(row[1]).strip()
                if voltage:
                    voltages[relay_type] = voltage
        wb.close()
        print(f"   🧠 Загружено из памяти напряжений: {len(voltages)} записей")
        return voltages
    except Exception as e:
        print(f"   ⚠️ Ошибка чтения файла памяти напряжений: {e}")
        return {}


def save_relay_voltage_memory(file_path: str, new_voltages: Dict[str, str]) -> None:
    """Обновляет или создает файл памяти напряжений реле."""
    existing_voltages = load_relay_voltage_memory(file_path)
    existing_voltages.update(new_voltages)

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Память напряжений"

        ws.append(["Тип реле", "Номинальное напряжение"])

        for r_type, voltage in existing_voltages.items():
            ws.append([r_type, voltage])

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    cell.data_type = 's'

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        wb.save(file_path)
        print(f"   💾 Память напряжений сохранена в: {os.path.basename(file_path)}")
    except PermissionError:
        print(f"   ❌ Ошибка: Файл {os.path.basename(file_path)} открыт в другой программе.")
    except Exception as e:
        print(f"   ❌ Ошибка сохранения файла памяти: {e}")


def extract_relays_from_excel(
    input_excel_path: str,
    output_dir: Optional[str] = None,
    pdf_name: Optional[str] = None,
    skip_voltage_dialog: bool = False
) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """
    Главная функция этапа 3: извлечение реле из Excel файла.
    
    Args:
        input_excel_path: Путь к Excel файлу с извлеченными данными
        output_dir: Директория для сохранения результата (опционально)
        pdf_name: Имя исходного PDF файла для формирования имени выходного файла
        skip_voltage_dialog: Пропустить диалог выбора напряжения (использовать только память)
        
    Returns:
        (список_реле, путь_к_файлу) или (список_реле, None) если реле не найдены
    """
    print(f"\nШаг 3: Поиск реле в извлеченных данных...")
    print(f"   📂 Открываю файл: {os.path.basename(input_excel_path)}...")
    
    try:
        wb = openpyxl.load_workbook(input_excel_path, data_only=True)
    except Exception as e:
        print(f"   ❌ Ошибка при открытии файла: {e}")
        return [], None

    all_relays = []
    
    # Получаем имя щита из имени файла
    if pdf_name:
        shield_name = pdf_name
    else:
        shield_name = Path(input_excel_path).stem.replace('_extracted_temp', '')

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ['сводка', 'summary', 'текст_из_pdf']:
            continue

        print(f"\n   📄 Обработка листа: {sheet_name}")
        ws = wb[sheet_name]

        relays = process_worksheet_relays(ws, shield_name)
        if relays:
            print(f"   ✅ Найдено {len(relays)} реле на листе {sheet_name}")
            all_relays.extend(relays)

    wb.close()

    if not all_relays:
        print("\n⚠️ Реле не найдены в извлеченных данных.")
        return all_relays, None

    # Определяем путь для выходного файла
    if output_dir:
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
    else:
        output_path = Path(input_excel_path).parent

    # Формируем имя файла
    if pdf_name:
        base_name = pdf_name
    else:
        base_name = Path(input_excel_path).stem.replace('_extracted_temp', '')

    relay_output_file = output_path / f"{base_name}_Реле.xlsx"
    
    if relay_output_file.exists():
        base = relay_output_file.stem
        ext = relay_output_file.suffix
        counter = 1
        while relay_output_file.exists():
            relay_output_file = output_path / f"{base}_{counter}{ext}"
            counter += 1

    # Получаем уникальные типы реле
    unique_relays = get_unique_relay_types(all_relays)

    # Путь к файлу памяти в папке со скриптом
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    memory_file_path = os.path.join(script_dir, "Relay_voltage.xlsx")

    # Загружаем уже известные напряжения из файла
    final_voltage_map = load_relay_voltage_memory(memory_file_path)

    # Ищем реле, которых НЕТ в файле памяти
    unknown_relays = [r for r in unique_relays if r['type'] not in final_voltage_map]

    # Показываем окно только для новых реле (если не пропущен диалог)
    if unknown_relays and not skip_voltage_dialog:
        print(f"\n💡 Найдено {len(unknown_relays)} новых типов реле.")
        print("   Для выбора напряжения используйте GUI (эта функция требует доработки для pipeline)")
        # В pipeline режиме просто используем пустые значения для неизвестных реле
        for relay in unknown_relays:
            final_voltage_map[relay['type']] = ''
    elif unknown_relays:
        print(f"\n⚠️ Найдено {len(unknown_relays)} новых типов реле без напряжения (режим без GUI)")
        for relay in unknown_relays:
            final_voltage_map[relay['type']] = ''
    else:
        print("\n✅ Напряжения для всех найденных типов реле загружены из файла памяти (Relay_voltage.xlsx).")

    # Сохраняем итоговый файл со спецификацией
    save_relays_to_xlsx(all_relays, str(relay_output_file), voltage_map=final_voltage_map)

    print(f"\n✅ Реле сохранены: {relay_output_file.name}")
    print(f"   📊 Всего реле: {len(all_relays)}")

    return all_relays, str(relay_output_file)


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python stage3_extract_relays.py <путь_к_Excel> [выходная_директория]")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else None
    
    relays, output_file = extract_relays_from_excel(excel_path, output_dir, skip_voltage_dialog=True)
    if relays:
        print(f"\n✅ Этап 3 завершен. Найдено {len(relays)} реле.")
        if output_file:
            print(f"   Результат сохранен: {output_file}")
    else:
        print("\n⚠️ Реле не найдены")
