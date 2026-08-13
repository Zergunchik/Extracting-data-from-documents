import sys
import os
import re
from openpyxl import load_workbook, Workbook
from pathlib import Path
import builtins

def should_stop():
    """Проверяет, запросил ли пользователь остановку через GUI."""
    check = getattr(builtins, '_gui_stop_check', None)
    return check() if check else False

OUTPUT_DIR = None  # Глобальная переменная для папки вывода

def find_column_by_header(sheet, target_headers):
    """Ищет номер столбца по тексту заголовка"""
    if isinstance(target_headers, str):
        target_headers = [target_headers]
    for row in range(1, min(20, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                cell_clean = cell_value.strip().lower()
                
                for target in target_headers:
                    target_clean = target.strip().lower()
                    if cell_clean == target_clean or cell_clean.startswith(target_clean):
                        if len(cell_clean) > len(target_clean):
                            next_char = cell_clean[len(target_clean)]
                            if next_char.isalnum() or next_char == '_':
                                continue
                        return col, row
    return None, None

def find_transformer_columns(sheet, col_fider):
    """Ищет столбцы, связанные с измерительным трансформатором"""
    col_type = None
    col_quantity = None
    header_row = None
    for row in range(1, min(20, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                cell_clean = cell_value.strip().lower()
                
                if "измерительный трансформатор" in cell_clean:
                    header_row = row
                    
                    next_row = row + 1
                    if next_row <= sheet.max_row:
                        for check_col in range(col, min(col + 5, sheet.max_column + 1)):
                            sub_cell = sheet.cell(row=next_row, column=check_col).value
                            if sub_cell and isinstance(sub_cell, str):
                                sub_clean = sub_cell.strip().lower()
                                if "тип" in sub_clean:
                                    col_type = check_col
                                elif "кол" in sub_clean or "количество" in sub_clean:
                                    col_quantity = check_col
                    
                    if col_type is None:
                        col_type = col
                    
                    return col_type, col_quantity, header_row

    return None, None, None

def find_data_start_row(sheet, header_row, col_fider, col_type):
    """Находит строку, с которой начинаются реальные данные"""
    start_row = header_row + 1
    for row in range(header_row + 1, min(header_row + 50, sheet.max_row + 1)):
        fider_val = sheet.cell(row=row, column=col_fider).value
        type_val = sheet.cell(row=row, column=col_type).value if col_type else None
        if (fider_val and str(fider_val).strip()) or (type_val and str(type_val).strip()):
            return row
    return start_row

def extract_full_fider_number(fider_str):
    """
    Извлекает полный номер фидера из позиционного обозначения
    Возвращает строку с номером (может содержать точку)
    Примеры: "QF1-2.1" -> "2.1", "Фидер 6.4" -> "6.4", "Ввод 1.1" -> "1.1"
    """
    if not fider_str:
        return None
    match = re.search(r'(\d+(?:\.\d+)?)$', fider_str)
    if match:
        return match.group(1)
    match = re.search(r'\d+', fider_str)
    if match:
        return match.group(0)
    return None

def extract_q_fider_number(fider_str):
    """Извлекает номер из фидера типа Q*"""
    if not fider_str:
        return None
    fider_clean = fider_str.strip()
    match = re.match(r'^[Qq](\d+)(?:[\.\-\d]*)?$', fider_clean)
    if match:
        return int(match.group(1))
    return None

def is_q_fider(fider_str):
    """Проверяет, является ли фидер типа Q*"""
    if not fider_str:
        return False
    fider_clean = fider_str.strip()
    match = re.match(r'^[Qq]\d+[\.\d]*$', fider_clean)
    return match is not None

def get_input_fider_number(fider_str):
    """
    Определяет номер вводного фидера (1 или 2)
    Возвращает 1 для "Ввод 1", "Ввод1", "Input 1" и т.д.
    Возвращает 2 для "Ввод 2", "Ввод2", "Input 2" и т.д.
    """
    if not fider_str:
        return None
    fider_lower = fider_str.strip().lower()
    patterns = [r'ввод\s*(\d+)', r'input\s*(\d+)', r'вв\.?\s*(\d+)']
    for pattern in patterns:
        match = re.search(pattern, fider_lower)
        if match:
            return int(match.group(1))
    return None

def is_input_fider(fider_str):
    """Проверяет, является ли фидер вводным"""
    if not fider_str:
        return False
    fider_lower = fider_str.strip().lower()
    patterns = [r'^ввод\d*$', r'^ввод\s+\d+$', r'^input\d*$', r'^input\s+\d+$', r'^вв\.?\s*\d*$']
    for pattern in patterns:
        if re.search(pattern, fider_lower):
            return True
    return False

def parse_transformer_cell(cell_value):
    """Парсит ячейку с измерительным трансформатором"""
    if not cell_value:
        return None, None
    cell_str = str(cell_value).strip()
    match = re.search(r'\)?\s*\(?(\d+)\)?\s*$', cell_str)
    if match:
        quantity = int(match.group(1))
        type_str = cell_str[:match.start()].strip()
        return type_str, quantity
    return cell_str, None

def extract_transformation_ratio(tt_type):
    """Извлекает коэффициент трансформации из типа трансформатора"""
    if not tt_type or not isinstance(tt_type, str):
        return None
    patterns = [
        r'(\d+)/([15])(?:\D|$)',
        r'(\d+)-([15])А?(?:\D|$)',
        r'(\d+)-([15])(?:\D|$)',
        r'(\d+)/([15])А?(?:\D|$)',
    ]
    matches = []
    for pattern in patterns:
        found = re.findall(pattern, tt_type)
        if found:
            matches.extend(found)

    if matches:
        numerator, denominator = matches[-1]
        return f"{numerator}/{denominator}"

    fallback_match = re.search(r'(\d+)[/\-]([15])А?(?:\D|$)', tt_type[::-1])
    if fallback_match:
        numerator = fallback_match.group(1)[::-1]
        denominator = fallback_match.group(2)
        return f"{numerator}/{denominator}"
    return None

def extract_scheme_designation(tt_number_in_fider, fider_number_full):
    """Формирует схемное обозначение в формате ТА{номер_ТТ_в_фидере}-{полный_номер_фидера}"""
    if fider_number_full is None:
        fider_number_full = "?"
    return f"ТА{tt_number_in_fider}-{fider_number_full}"

def extract_scheme_designation_for_q_fider(q_number, tt_index, total_tt_count):
    """Формирует схемное обозначение для фидеров типа Q*"""
    return f"3ТА1-{q_number}{tt_index}"

def extract_installation_location(fider_str, fider_number_full, load_name=None):
    """Формирует место установки"""
    if load_name and isinstance(load_name, str):
        load_name_clean = load_name.strip()
        if load_name_clean:
            load_lower = load_name_clean.lower()
            if any(keyword in load_lower for keyword in ["ввод", "input", "вв"]):
                return load_name_clean
            return load_name_clean
            
    if is_input_fider(fider_str):
        return fider_str.strip()
    elif is_q_fider(fider_str):
        return f"Нагрузка {fider_str}"
    elif fider_number_full:
        return f"Фидер {fider_number_full}"
    return "требуется уточнение"

def process_xlsx(input_file_path):
    """Обработка XLSX файла с таблицей фидеров (только трансформаторы)"""
    wb = load_workbook(input_file_path, data_only=True)
    sheet = wb.active
    print(f"Обработка листа: {sheet.title}")
    print("=" * 50)

    col_fider, header_row_fider = find_column_by_header(sheet, [
        "Позиционное обозначение фидера",
        "Позиционное обозначение фидеров по ИД",
        "Позиционное обозначение",
        "Обозначение фидера",
        "Фидер",
        "Поз. обозначение фидера"
    ])

    col_load_name, header_row_load = find_column_by_header(sheet, [
        "Наименование нагрузки",
        "Название нагрузки",
        "Наименование",
        "Нагрузка",
        "Потребитель"
    ])

    col_tt_type, col_tt_quantity, header_row_tt = find_transformer_columns(sheet, col_fider)

    if col_fider is None:
        print("❌ Столбец 'Позиционное обозначение фидера' не найден!")
        return [], None, None, None, None

    header_row = header_row_fider
    if header_row_tt:
        header_row = max(header_row, header_row_tt)
    if header_row_load:
        header_row = max(header_row, header_row_load)

    data_start_row = find_data_start_row(sheet, header_row, col_fider, col_tt_type)

    print(f"✅ Найден столбец фидера: {col_fider} (строка заголовка: {header_row_fider})")

    if col_load_name:
        print(f"✅ Найден столбец наименования нагрузки: {col_load_name} (строка заголовка: {header_row_load})")
    else:
        print("⚠️ Столбец 'Наименование нагрузки' не найден")

    if col_tt_type:
        print(f"✅ Найден столбец измерительного трансформатора (тип): {col_tt_type}")
    else:
        print("⚠️ Столбец 'Измерительный трансформатор' не найден!")
        return [], None, None, None, None
        
    if col_tt_quantity:
        print(f"✅ Найден столбец измерительного трансформатора (количество): {col_tt_quantity}")

    print(f"📌 Данные начинаются с строки: {data_start_row}")
    print("=" * 50)

    transformers = []
    counter = 1  # Общий порядковый номер для № п/п

    for row_idx in range(data_start_row, sheet.max_row + 1):
        if should_stop():
            print("⏹️ Остановка по запросу пользователя")
            return transformers, col_fider, col_tt_type, col_tt_quantity, header_row
        fider_value = sheet.cell(row=row_idx, column=col_fider).value 
        
        if fider_value is None or str(fider_value).strip() == "":
            continue
        
        fider_str = str(fider_value).strip()
        fider_number_full = extract_full_fider_number(fider_str)
        is_q = is_q_fider(fider_str)
        q_number = extract_q_fider_number(fider_str) if is_q else None
        is_input = is_input_fider(fider_str)
        input_number = get_input_fider_number(fider_str) if is_input else None
        
        load_name = None
        if col_load_name:
            load_cell = sheet.cell(row=row_idx, column=col_load_name).value
            if load_cell and str(load_cell).strip():
                load_name = str(load_cell).strip()
        
        tt_type = None
        tt_quantity = None
        
        if col_tt_type:
            tt_cell = sheet.cell(row=row_idx, column=col_tt_type).value
            if tt_cell and str(tt_cell).strip():
                if col_tt_quantity:
                    tt_type = str(tt_cell).strip()
                    qty_cell = sheet.cell(row=row_idx, column=col_tt_quantity).value
                    if qty_cell:
                        try:
                            tt_quantity = int(float(qty_cell))
                        except (ValueError, TypeError):
                            tt_quantity = 1
                    else:
                        tt_quantity = 1
                else:
                    tt_type, tt_quantity = parse_transformer_cell(tt_cell)
                    if tt_quantity is None:
                        tt_quantity = 1
                
                if tt_type:
                    ktt = extract_transformation_ratio(tt_type)
                    location = extract_installation_location(fider_str, fider_number_full, load_name)
                    actual_quantity = max(1, tt_quantity if tt_quantity else 1)
                    
                    for tt_index in range(1, actual_quantity + 1):
                        # === ОСОБАЯ ЛОГИКА НУМЕРАЦИИ ===
                        if is_input and input_number == 2:
                            # Для ввода 2: 3ТА21, 3ТА22, 3ТА23, 5ТА2
                            names = ["3ТА21", "3ТА22", "3ТА23", "5ТА2"]
                            scheme_designation = names[tt_index - 1] if tt_index <= len(names) else f"5ТА{tt_index - 2}"  
                                                   
                        elif is_q and q_number is not None:
                            # Для Q-фидеров
                            scheme_designation = extract_scheme_designation_for_q_fider(q_number, tt_index, actual_quantity)
                            
                        else:
                            # Для обычных фидеров: ТА1-2.1, ТА2-2.1 и т.д.
                            scheme_designation = extract_scheme_designation(
                                tt_index,
                                fider_number_full if fider_number_full else "?"
                            )
                        
                        transformers.append({
                            "number": counter,
                            "fider": fider_str,
                            "tt_type": tt_type,
                            "location": location,
                            "scheme_designation": scheme_designation,
                            "ktt": ktt if ktt else "требуется уточнение",
                            "quantity": 1,
                            "load_name": load_name,
                            "fider_number": fider_number_full
                        })
                        counter += 1

    print(f"\n📊 Статистика трансформаторов тока:")
    print(f"   - Найдено трансформаторов: {len(transformers)}")

    return transformers, col_fider, col_tt_type, col_tt_quantity, header_row

def save_transformers_to_xlsx(transformers, output_file_path):
    """Сохраняет результаты по трансформаторам тока в отдельный XLSX файл"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Трансформаторы тока"
    headers = ["№ п/п", "Тип ТТ", "Место установки", "Схемное обозначение", "Ктт", "Количество", "Исходный фидер"]
    ws.append(headers)

    for tt in transformers:
        ws.append([
            tt["number"],
            tt["tt_type"],
            tt["location"],
            tt["scheme_designation"],
            tt["ktt"],
            tt["quantity"],
            tt["fider"]
        ])

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20

    wb.save(output_file_path)
    print(f"📁 Файл с трансформаторами сохранён: {os.path.basename(output_file_path)}")
    return output_file_path

def main():
    # Проверка на режим GUI
    is_gui_mode = '--from-gui' in sys.argv
    
    # Проверяем флаг --output-dir
    global OUTPUT_DIR
    output_dir = None
    if '--output-dir' in sys.argv:
        idx = sys.argv.index('--output-dir')
        if idx + 1 < len(sys.argv):
            output_dir = sys.argv[idx + 1]
            sys.argv.pop(idx)
            sys.argv.pop(idx)
    
    if output_dir:
        OUTPUT_DIR = Path(output_dir)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    if is_gui_mode:
        # Убираем аргумент --from-gui из списка
        args = [arg for arg in sys.argv[1:] if arg != '--from-gui']
        if not args:
            print("❌ Ошибка: не указан путь к файлу!")
            return
        input_file_path = args[0]
    else:
        if len(sys.argv) < 2:
            print("Использование: перетащите XLSX файл на этот скрипт")
            print("\nСкрипт извлекает информацию о трансформаторах тока из таблицы фидеров")
            return
        input_file_path = sys.argv[1]

    # ... остальной код ...

    try:
        transformers, col_fider, col_tt_type, col_tt_quantity, header_row = process_xlsx(input_file_path)
        
        input_path = Path(input_file_path)
        if OUTPUT_DIR:
            output_dir_path = OUTPUT_DIR
        else:
            output_dir_path = input_path.parent
        
        base_name = input_path.stem
        
        if transformers:
            output_transformers_path = output_dir_path / f"{base_name}_Трансформаторы тока.xlsx"
            save_transformers_to_xlsx(transformers, str(output_transformers_path))
            print(f"\n✅ Обработано трансформаторов тока: {len(transformers)}")
            
            print("\n📊 Пример результатов по трансформаторам (первые 10):")
            for i, tt in enumerate(transformers[:10]):
                print(f"   {i+1}. №{tt['number']}: {tt['location']} → {tt['scheme_designation']} → {tt['tt_type'][:30]} → {tt['ktt']}")
        else:
            print("\n⚠️ Трансформаторы тока не найдены в файле")

    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()

    print("\n" + "=" * 50)
    
    # Только в обычном режиме ждем нажатия Enter
    if not is_gui_mode:
        input("Нажмите Enter для выхода...")

if __name__ == "__main__":
    main()