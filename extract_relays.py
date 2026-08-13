# extract_relays.py - Полная исправленная версия с поддержкой цифр перед буквами
import pdfplumber
import pandas as pd
import openpyxl
import re
import os
import sys
from pathlib import Path
import time
import logging
import builtins

def should_stop():
    """Проверяет, запросил ли пользователь остановку через GUI."""
    check = getattr(builtins, '_gui_stop_check', None)
    return check() if check else False

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

# --- ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ ---
MODE_RELAYS = 4

# --- СПИСОК ИСКЛЮЧАЕМЫХ КЛЮЧЕВЫХ СЛОВ ---
NEGATIVE_KEYWORDS = [
    'компрессор', 'компресс',
    'насос', 'pump',
    'chiller',
    'кондиционер внешний',
    'двигатель',
    'мотор'
]

# --- СПИСОК КЛЮЧЕВЫХ СЛОВ ДЛЯ РЕЛЕ ---
RELAY_KEYWORDS = [
    'реле',
    'реле напряжения',
    'реле контроля',
    'контроллер',
    'реле времени',
    'тепловое реле',
    'промежуточное реле',
    'реле тока',
    'реле температуры',
    'RKE4CO',
    'RKE4CO024LT',
    'RKE4CO730LT'
]

# --- ИСКЛЮЧАЕМЫЕ ТИПЫ РЕЛЕ ---
EXCLUDED_RELAY_TYPES = [
    'реле напряжения',
    'реле контроля напряжения',
    'KV',
    'KVZ',
    'контроль напряжения'
]

# ============================================================
# ЧАСТЬ 1: ИЗВЛЕЧЕНИЕ ИЗ PDF (полная логика из оригинального скрипта)
# ============================================================

def diagnose_pdf(pdf_path):
    """Диагностика содержимого PDF файла"""
    print(f"\n🔍 ДИАГНОСТИКА PDF: {os.path.basename(pdf_path)}")
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            print(f"  - Количество страниц: {total_pages}")
            
            for page_num in range(min(3, total_pages)):
                page = pdf.pages[page_num]
                text = page.extract_text()
                print(f"  - Страница {page_num + 1}: текст {'есть' if text else 'ОТСУТСТВУЕТ'}")
                if text:
                    preview = text[:200].replace('\n', ' ')
                    print(f"    Начало текста: {preview}...")
                
                tables = page.extract_tables()
                if tables:
                    print(f"  - Страница {page_num + 1}: найдено {len(tables)} таблиц")
                    for i, table in enumerate(tables[:2], 1):
                        if table:
                            print(f"    Таблица {i}: {len(table)} строк")
                            if len(table) > 1 and table[0]:
                                print(f"      Заголовок: {table[0][:5]}")
                else:
                    print(f"  - Страница {page_num + 1}: таблиц не найдено")
            print()
    except Exception as e:
        print(f"❌ Ошибка при диагностике: {e}")

def extract_text_from_pdf(pdf_path):
    extracted_data = []
    print(f"   📄 Чтение текста из: {os.path.basename(pdf_path)}")
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            print(f"   📄 Всего страниц: {total_pages}")
            
            if total_pages == 0:
                print(f"   ⚠️ PDF файл не содержит страниц!")
                return None
                
            for page_num, page in enumerate(pdf.pages, 1):
                if page_num % 5 == 0 or page_num == total_pages:
                    print(f"   ⏳ Обработка текста: страница {page_num}/{total_pages}")
                
                text = page.extract_text()
                if text:
                    lines = text.split('\n')
                    for line_num, line in enumerate(lines, 1):
                        if line.strip():
                            extracted_data.append({
                                'Страница': page_num,
                                'Строка': line_num,
                                'Текст': line.strip()
                            })
            
            print(f"   ✅ Извлечено {len(extracted_data)} строк текста")
            return extracted_data
    except Exception as e:
        print(f"❌ Ошибка при чтении текста PDF: {e}")
        import traceback
        traceback.print_exc()
        return None

def extract_tables_from_pdf(pdf_path):
    all_tables = []
    print(f"   📊 Чтение таблиц из: {os.path.basename(pdf_path)}")
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            print(f"   📊 Всего страниц: {total_pages}")
            
            for page_num, page in enumerate(pdf.pages, 1):
                if page_num % 5 == 0 or page_num == total_pages:
                    print(f"   ⏳ Обработка таблиц: страница {page_num}/{total_pages}")
                
                try:
                    tables = page.extract_tables()
                    if tables:
                        print(f"   📊 Страница {page_num}: найдено {len(tables)} таблиц")
                    else:
                        tables = page.extract_tables({
                            "vertical_strategy": "lines",
                            "horizontal_strategy": "lines"
                        })
                        if tables:
                            print(f"   📊 Страница {page_num}: найдено {len(tables)} таблиц (альтернативный метод)")
                    
                    for table_num, table in enumerate(tables, 1):
                        if table:
                            clean_table = []
                            for row in table:
                                clean_row = [str(cell).strip() if cell is not None else '' for cell in row]
                                if any(clean_row):
                                    clean_table.append(clean_row)
                            if clean_table:
                                all_tables.append({
                                    'Страница': page_num,
                                    'Таблица': table_num,
                                    'Данные': clean_table
                                })
                                print(f"   📊 Таблица {table_num} на странице {page_num}: {len(clean_table)} строк, {len(clean_table[0]) if clean_table else 0} колонок")
                except Exception as page_e:
                    print(f"   ⚠️ Ошибка на странице {page_num}: {page_e}")
                    continue
    except Exception as e:
        print(f"❌ Ошибка при извлечении таблиц: {e}")
        import traceback
        traceback.print_exc()
        return all_tables
    
    print(f"   ✅ Всего извлечено {len(all_tables)} таблиц")
    return all_tables

# ============================================================
# ЧАСТЬ 2: ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================

def normalize_text(text):
    """Нормализация текста (замена русских букв на латинские)"""
    if not text:
        return text
    text = str(text)
    text = text.replace('К', 'K').replace('к', 'k')
    text = text.replace('М', 'M').replace('м', 'm')
    return text

def has_negative_keywords(text):
    """Проверяет, содержит ли текст негативные ключевые слова"""
    if not text:
        return False
    text_lower = text.lower()
    for neg_word in NEGATIVE_KEYWORDS:
        if neg_word in text_lower:
            return True
    return False

def is_garbage_row(row):
    """Проверяет, является ли строка мусорной (служебной информацией)"""
    if not row:
        return True
    
    row_text = " ".join([str(cell).strip() for cell in row if cell])
    
    garbage_patterns = [
        r'^атад$',
        r'^и$',
        r'^ьсипдоП$',
        r'^\.лбуд$',
        r'^№\.внИ$',
        r'^№\.вни$',
        r'^\.мазВ$',
        r'^\.лдоп$',
        r'^Копировал:',
        r'^Изм\.',
        r'^Лист$',
        r'^№ докум\.',
        r'^Подпись$',
        r'^Дата$',
        r'^Формат',
        r'^K\-BLOCK\-',
    ]
    
    for pattern in garbage_patterns:
        if re.search(pattern, row_text, re.IGNORECASE):
            return True
    
    # Если строка содержит только короткие слова - возможно мусор
    words = row_text.split()
    if len(words) <= 4:
        short_words = [w for w in words if len(w) <= 3]
        if len(short_words) / max(len(words), 1) > 0.7:
            return True
    
    return False

def find_columns(data):
    """
    Находит столбцы с позициями и наименованиями
    Возвращает: (pos_col_idx, name_col_idx, start_row_idx)
    """
    if not data or not data[0]:
        return None, None, -1
    
    num_cols = len(data[0])
    pos_col_idx = None
    name_col_idx = None
    start_row = -1
    
    # Первый проход: ищем заголовки в первых 20 строках
    for row_idx, row in enumerate(data[:20]):
        if is_garbage_row(row):
            continue
            
        row_text_lower = " ".join([str(cell).lower() for cell in row if cell])
        
        if not pos_col_idx:
            for col_idx, cell in enumerate(row):
                if cell:
                    cell_str = str(cell).lower().strip()
                    if 'поз' in cell_str or 'обознач' in cell_str:
                        pos_col_idx = col_idx
                        break
        
        if not name_col_idx:
            for col_idx, cell in enumerate(row):
                if cell:
                    cell_str = str(cell).lower().strip()
                    if 'наименование' in cell_str or 'наим' in cell_str or 'описание' in cell_str:
                        name_col_idx = col_idx
                        break
        
        if pos_col_idx is not None and name_col_idx is not None:
            start_row = row_idx + 1
            break
    
    # Второй проход: если заголовки не найдены, ищем по данным
    if pos_col_idx is None or name_col_idx is None:
        print("   🔍 Ищу столбцы по содержимому данных...")
        
        for col_idx in range(num_cols):
            relay_positions = 0
            relay_names = 0
            
            for row in data[20:60]:
                if col_idx < len(row) and row[col_idx]:
                    cell_str = normalize_text(str(row[col_idx]))
                    
                    # Обновленное регулярное выражение с поддержкой цифр перед буквами
                    if re.search(r'\b\d*KL?\d+(?:\.\d+)?|\b\d*K\d+(?:\.\d+)?[^V]|\b\d*KLB\d+(?:\.\d+)?', cell_str):
                        relay_positions += 1
                    
                    cell_lower = str(row[col_idx]).lower()
                    for keyword in RELAY_KEYWORDS:
                        if keyword.lower() in cell_lower:
                            relay_names += 1
                            break
            
            if relay_positions > 3 and pos_col_idx is None:
                pos_col_idx = col_idx
                print(f"   ✅ Найдена колонка с позициями: {col_idx}")
            
            if relay_names > 2 and name_col_idx is None:
                name_col_idx = col_idx
                print(f"   ✅ Найдена колонка с наименованиями: {col_idx}")
    
    # Если не нашли стартовую строку, ищем первую строку с данными
    if start_row == -1 and pos_col_idx is not None:
        for row_idx, row in enumerate(data):
            if row_idx < 10:
                continue
            if pos_col_idx < len(row) and row[pos_col_idx]:
                cell_str = normalize_text(str(row[pos_col_idx]))
                # Обновленное регулярное выражение с поддержкой цифр перед буквами
                if re.search(r'\b\d*KL?\d+(?:\.\d+)?|\b\d*K\d+(?:\.\d+)?[^V]|\b\d*KLB\d+(?:\.\d+)?', cell_str):
                    start_row = row_idx
                    break
    
    print(f"   📊 Результат: Поз.={pos_col_idx}, Наим.={name_col_idx}, Начало={start_row}")
    return pos_col_idx, name_col_idx, start_row

# ============================================================
# ЧАСТЬ 3: ПОИСК РЕЛЕ
# ============================================================

def expand_positions_relay(text, full_row_text=None):
    """
    Извлекает позиции с реле (KL, K, KLB)
    Поддерживает: K1, K1.1, KL1, KLB1, K1...K10, K1.1...K1.9, 2K1, 4K1, 4K2, K1-4.1...K1-4.6
    """
    if not text or text == 'nan' or text == '':
        return []
    
    if full_row_text and has_negative_keywords(full_row_text):
        return []
    
    positions = []
    text = str(text).replace('\n', ' ').strip()
    normalized_text = normalize_text(text)

    # Исключаем реле напряжения
    if re.search(r'\bKV\d+(?:\.\d+)?|\bKVZ\d+(?:\.\d+)?', normalized_text, re.IGNORECASE):
        return []
    if re.search(r'реле\s+напряжени[яю]', normalized_text, re.IGNORECASE):
        return []

    # Проверяем наличие реле (обновленное регулярное выражение с поддержкой цифр перед буквами)
    if not re.search(r'\b\d*KL?\d+(?:\.\d+)?|\b\d*K\d+(?:\.\d+)?[^V]|\b\d*KLB\d+(?:\.\d+)?', normalized_text):
        return []

    parts = [p.strip() for p in text.split(',') if p.strip()]

    for part in parts:
        if not part:
            continue
             
        normalized_part = normalize_text(part)
        
        if re.search(r'\bKV\d+(?:\.\d+)?|\bKVZ\d+(?:\.\d+)?', normalized_part, re.IGNORECASE):
            continue
        if re.search(r'реле\s+напряжени[яю]', normalized_part, re.IGNORECASE):
            continue
        
        # Обновленное регулярное выражение с поддержкой цифр перед буквами
        if not re.search(r'\b\d*KL?\d+(?:\.\d+)?|\b\d*K\d+(?:\.\d+)?[^V]|\b\d*KLB\d+(?:\.\d+)?', normalized_part):
            continue
        
        # Обработка диапазонов с дефисом (например, K1-4.1...K1-4.6)
        if '...' in normalized_part:
            try:
                range_parts = normalized_part.split('...')
                if len(range_parts) == 2:
                    start_str = range_parts[0].strip()
                    end_str = range_parts[1].strip()
                    
                    # Проверяем, является ли это диапазоном с дефисом
                    match_start = re.search(r'^(\d*)(KL?B?)(\d+)-(\d+)\.(\d+)$', start_str, re.IGNORECASE)
                    match_end = re.search(r'^(\d*)(KL?B?)(\d+)-(\d+)\.(\d+)$', end_str, re.IGNORECASE)
                    
                    if match_start and match_end:
                        prefix_start = match_start.group(1) or ''
                        prefix_end = match_end.group(1) or ''
                        relay_type = match_start.group(2).upper()
                        base_num_start = int(match_start.group(3))
                        base_num_end = int(match_end.group(3))
                        sub_start = int(match_start.group(4))
                        sub_end = int(match_end.group(4))
                        dot_start = int(match_start.group(5))
                        dot_end = int(match_end.group(5))
                        
                        prefix = prefix_start if prefix_start else prefix_end
                        
                        # Генерируем все позиции в диапазоне
                        for base_num in range(base_num_start, base_num_end + 1):
                            if base_num == base_num_start:
                                start_dot = dot_start
                            else:
                                start_dot = 1
                            if base_num == base_num_end:
                                end_dot = dot_end
                            else:
                                end_dot = 99
                            
                            for dot_num in range(start_dot, end_dot + 1):
                                if dot_num > 99:
                                    break
                                positions.append(f"{prefix}{relay_type}{base_num}-{sub_start}.{dot_num}")
                    else:
                        # Пробуем стандартный диапазон без дефиса
                        match_start = re.search(r'^(\d*)(KL?B?)(\d+)(?:\.(\d+))?$', start_str, re.IGNORECASE)
                        match_end = re.search(r'^(\d*)(KL?B?)(\d+)(?:\.(\d+))?$', end_str, re.IGNORECASE)
                        
                        if match_start and match_end:
                            prefix_start = match_start.group(1) or ''
                            prefix_end = match_end.group(1) or ''
                            relay_type = match_start.group(2).upper()
                            start_num = int(match_start.group(3))
                            end_num = int(match_end.group(3))
                            start_sub = match_start.group(4)
                            end_sub = match_end.group(4)
                            
                            prefix = prefix_start if prefix_start else prefix_end
                            
                            if start_sub and end_sub:
                                start_sub_num = int(start_sub)
                                end_sub_num = int(end_sub)
                                for i in range(start_num, end_num + 1):
                                    for j in range(start_sub_num if i == start_num else 1, 
                                                  end_sub_num + 1 if i == end_num else 99):
                                        if j > 99:
                                            break
                                        positions.append(f"{prefix}{relay_type}{i}.{j}")
                            else:
                                for i in range(start_num, end_num + 1):
                                    positions.append(f"{prefix}{relay_type}{i}")
                        else:
                            positions.append(part)
                else:
                    positions.append(part)
            except Exception:
                positions.append(part)
        else:
            # Обновленное регулярное выражение с поддержкой цифр перед буквами
            if re.search(r'\b\d*KL?\d+(?:\.\d+)?|\b\d*K\d+(?:\.\d+)?[^V]|\b\d*KLB\d+(?:\.\d+)?', normalized_part):
                if not re.search(r'KV', normalized_part, re.IGNORECASE):
                    sub_parts = [p.strip() for p in part.split(',') if p.strip()]
                    for sub_part in sub_parts:
                        if re.search(r'\b\d*KL?\d+(?:\.\d+)?|\b\d*K\d+(?:\.\d+)?[^V]|\b\d*KLB\d+(?:\.\d+)?', normalize_text(sub_part)):
                            if not re.search(r'KV', normalize_text(sub_part), re.IGNORECASE):
                                positions.append(sub_part)
        
    return positions

def extract_relay_type(text):
    """Извлекает тип реле из текста"""
    if not text:
        return None, False
    
    text_lower = text.lower().strip()

    for excluded in EXCLUDED_RELAY_TYPES:
        if excluded.lower() in text_lower:
            return None, False
    
    is_relay = False
    for keyword in RELAY_KEYWORDS:
        if keyword.lower() in text_lower:
            is_relay = True
            break
    
    if not is_relay:
        return None, False
    
    patterns = [
        r'реле\s+([А-Яа-яA-Za-z0-9\-]+)',
        r'реле\s+([A-Za-z0-9\-]+)',
        r'([А-Яа-яA-Za-z0-9\-]+)\s+реле',
        r'реле\s+([А-Яа-яA-Za-z0-9\s]+?)(?:\s+на|\s+с|\s+до|\s+для|$)',
        r'([A-Z]{2,}\s*[A-Za-z0-9\-]{2,})',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            relay_type = match.group(1).strip()
            if len(relay_type) > 20:
                short_match = re.search(r'([А-Яа-яA-Za-z0-9\-]{2,10})', relay_type)
                if short_match:
                    return short_match.group(1), True
            return relay_type, True
    
    return "реле", True

# ============================================================
# ЧАСТЬ 4: ОБРАБОТКА ЛИСТА EXCEL
# ============================================================

def process_worksheet(ws, mode, shield_name):
    results_relays = []
    data = []
    
    print(f"   📋 Обработка листа: {ws.title}")
    
    for row in ws.iter_rows(values_only=True):
        data.append([str(c).strip() if c is not None else "" for c in row])

    if not data:
        print(f"   ⚠️ Лист {ws.title} пуст!")
        return results_relays
    
    print(f"   📊 Лист содержит {len(data)} строк, {len(data[0]) if data else 0} колонок")

    pos_col_idx, name_col_idx, start_row = find_columns(data)
    
    if pos_col_idx is None:
        print("   ❌ Не найден столбец с позициями!")
        return results_relays

    if name_col_idx is None:
        print("   ⚠️ Не найден столбец с наименованиями, использую все строки для поиска")

    print(f"   📊 Использую: Поз.={pos_col_idx}, Наим.={name_col_idx}, Начало={start_row}")

    if start_row == -1:
        for i, row in enumerate(data):
            if pos_col_idx < len(row) and row[pos_col_idx]:
                start_row = i
                break
    
    if start_row == -1:
        print("   ❌ Не найдены строки с данными!")
        return results_relays

    last_known_relay = None
    relay_count = 0

    def add_result_relay(positions, relay_type):
        if not relay_type:
            return
        for pos in positions:
            results_relays.append({
                'pos': pos,
                'type': relay_type,
                'shield': shield_name
            })

    for i in range(start_row, len(data)):
        if should_stop():
            print("⏹️ Остановка по запросу пользователя")
            return results_relays
        row = data[i]
        
        if not any(row):
            continue
        
        if is_garbage_row(row):
            continue
        
        pos_text = row[pos_col_idx] if pos_col_idx < len(row) else ""
        name_text = row[name_col_idx] if name_col_idx is not None and name_col_idx < len(row) else ""
        row_text = " ".join(row)
        
        positions_relay = expand_positions_relay(pos_text, row_text)
        
        if positions_relay:
            relay_type = None
            
            if name_text:
                relay_type, is_relay = extract_relay_type(name_text)
                if not is_relay:
                    relay_type = None
            
            if not relay_type:
                if "изделие в составе" in row_text.lower():
                    for j in range(1, 4):
                        if i + j < len(data):
                            next_row = data[i + j]
                            if not is_garbage_row(next_row):
                                next_name = next_row[name_col_idx] if name_col_idx is not None and name_col_idx < len(next_row) else ""
                                next_row_text = " ".join(next_row)
                                relay_type, is_relay = extract_relay_type(next_name or next_row_text)
                                if is_relay and relay_type:
                                    break
                
                if not relay_type:
                    relay_type, is_relay = extract_relay_type(row_text)
                    if not is_relay:
                        relay_type = None
            
            if relay_type:
                last_known_relay = relay_type
                add_result_relay(positions_relay, relay_type)
                relay_count += len(positions_relay)
            elif last_known_relay:
                add_result_relay(positions_relay, last_known_relay)
                relay_count += len(positions_relay)
            else:
                add_result_relay(positions_relay, "реле")
                relay_count += len(positions_relay)
        
        if name_text:
            relay_type, is_relay = extract_relay_type(name_text)
            if is_relay and relay_type:
                last_known_relay = relay_type

    print(f"   ✅ Найдено реле: {relay_count}")
    return results_relays

# ============================================================
# ЧАСТЬ 5: ОБРАБОТКА EXCEL ФАЙЛА И СОХРАНЕНИЕ
# ============================================================

def process_extracted_excel(input_file, mode, shield_name):
    print(f"   📂 Открываю файл: {os.path.basename(input_file)}...")
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
    except Exception as e:
        print(f"   ❌ Ошибка при открытии файла: {e}")
        return None
    
    all_relays = []

    for sheet_name in wb.sheetnames:
        if should_stop():
            print("⏹️ Остановка по запросу пользователя")
            return all_relays
        if sheet_name.lower() in ['сводка', 'summary', 'текст_из_pdf']:
            continue
            
        print(f"\n   📄 Обработка листа: {sheet_name}")
        ws = wb[sheet_name]
        relays = process_worksheet(ws, mode, shield_name)
        
        if relays:
            print(f"   ✅ Найдено {len(relays)} реле на листе {sheet_name}")
            all_relays.extend(relays)
        
    wb.close()
    return all_relays

def save_aggregated_results(relays_list, output_dir, merge_mode=False):
    if not relays_list:
        print("⚠️ Нет данных для сохранения в общий файл.")
        return

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if not merge_mode:
        # Обычный режим - сохраняем с меткой времени
        output_file = Path(output_dir) / f"Общий_Отчет_Реле_{timestamp}.xlsx"
        
        counter = 1
        while output_file.exists():
            output_file = Path(output_dir) / f"Общий_Отчет_Реле_{timestamp}_{counter}.xlsx"
            counter += 1
        
        # Создаем новый файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Реле (KL, K, KLB)"
        
        ws.append([
            "№ п/п",
            "Позиционное обозначение (KL/K/KLB)",
            "Тип реле",
            "Источник (Файл/Щит)"
        ])
        
        for i, relay in enumerate(relays_list, 1):
            ws.append([
                i,
                relay.get('pos', ''),
                relay.get('type', ''),
                relay.get('shield', '')
            ])
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30

    else:
        # Режим объединения - добавляем данные в существующий файл или создаем новый
        output_file = Path(output_dir) / "Реле_объединенные.xlsx"
        
        # Проверяем, существует ли файл
        if output_file.exists():
            print(f"   📂 Найден существующий файл: {output_file.name}")
            try:
                # Загружаем существующий файл
                wb = openpyxl.load_workbook(output_file)
                
                # Проверяем, есть ли нужный лист
                if "Реле (KL, K, KLB)" in wb.sheetnames:
                    ws = wb["Реле (KL, K, KLB)"]
                else:
                    # Если листа нет, создаем его
                    ws = wb.create_sheet("Реле (KL, K, KLB)")
                    ws.append([
                        "№ п/п",
                        "Позиционное обозначение (KL/K/KLB)",
                        "Тип реле",
                        "Источник (Файл/Щит)"
                    ])
                
                # Получаем существующие данные для проверки дубликатов
                existing_data = set()
                max_row = ws.max_row
                for row in range(2, max_row + 1):  # Пропускаем заголовок
                    pos_cell = ws.cell(row=row, column=2).value
                    if pos_cell:
                        existing_data.add(str(pos_cell))
                
                # Находим следующий номер
                next_num = max_row  # Начинаем с последней строки
                if max_row >= 2:
                    # Проверяем, есть ли данные в последней строке
                    last_pos = ws.cell(row=max_row, column=2).value
                    if last_pos:
                        next_num = max_row + 1
                    else:
                        # Ищем реальный последний номер
                        for row in range(max_row, 1, -1):
                            if ws.cell(row=row, column=1).value:
                                next_num = row + 1
                                break
                else:
                    next_num = 2
                
                # Добавляем новые данные, пропуская дубликаты
                added_count = 0
                skipped_count = 0
                
                for relay in relays_list:
                    pos = relay.get('pos', '')
                    if pos and str(pos) not in existing_data:
                        ws.append([
                            next_num,
                            pos,
                            relay.get('type', ''),
                            relay.get('shield', '')
                        ])
                        existing_data.add(str(pos))
                        next_num += 1
                        added_count += 1
                    else:
                        skipped_count += 1
                
                print(f"   ✅ Добавлено {added_count} новых реле, пропущено {skipped_count} дубликатов")
                
            except Exception as e:
                print(f"   ⚠️ Ошибка при загрузке существующего файла: {e}")
                print(f"   🔄 Создаю новый файл...")
                # Если не удалось загрузить, создаем новый
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Реле (KL, K, KLB)"
                
                ws.append([
                    "№ п/п",
                    "Позиционное обозначение (KL/K/KLB)",
                    "Тип реле",
                    "Источник (Файл/Щит)"
                ])
                
                for i, relay in enumerate(relays_list, 1):
                    ws.append([
                        i,
                        relay.get('pos', ''),
                        relay.get('type', ''),
                        relay.get('shield', '')
                    ])
        else:
            # Файл не существует - создаем новый
            print(f"   📝 Создаю новый файл: {output_file.name}")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Реле (KL, K, KLB)"
            
            ws.append([
                "№ п/п",
                "Позиционное обозначение (KL/K/KLB)",
                "Тип реле",
                "Источник (Файл/Щит)"
            ])
            
            for i, relay in enumerate(relays_list, 1):
                ws.append([
                    i,
                    relay.get('pos', ''),
                    relay.get('type', ''),
                    relay.get('shield', '')
                ])
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30

    try:
        wb.save(str(output_file))
        print(f"\n✅ ОТЧЕТ СОХРАНЕН: {output_file.name}")
        print(f"   📊 Реле: {len(relays_list)}")
        print(f"   📁 Путь: {output_file.parent}")
        if merge_mode:
            print(f"   🔄 Режим: Объединение с существующими данными")
        else:
            print(f"   🔄 Режим: Новый файл с меткой времени")
    except Exception as e:
        print(f"❌ Ошибка при сохранении файла: {e}")
        import traceback
        traceback.print_exc()

# ============================================================
# ЧАСТЬ 6: MAIN
# ============================================================

def main():
    print("=" * 70)
    print("ОТЛАДОЧНАЯ ИНФОРМАЦИЯ:")
    print(f"Аргументы командной строки: {sys.argv}")
    print(f"Текущая рабочая директория: {os.getcwd()}")
    if '--mode' in sys.argv:
        mode_idx = sys.argv.index('--mode')
        if mode_idx + 1 < len(sys.argv):
            print(f"Режим: {sys.argv[mode_idx + 1]}")
    print("=" * 70)
    print()
    
    pdf_files_in_dir = [f for f in os.listdir('.') if f.lower().endswith('.pdf')]
    print(f"Найдено PDF файлов в папке: {len(pdf_files_in_dir)}")
    if pdf_files_in_dir:
        print("Первые 5 файлов:")
        for f in pdf_files_in_dir[:5]:
            if os.path.exists(f):
                size = os.path.getsize(f)
                print(f"  - {f} (размер: {size} байт)")
            else:
                print(f"  - {f} (файл не найден)")
    print()
    
    args = sys.argv[1:]
    is_gui_mode = '--from-gui' in args
    batch_mode = '--batch-mode' in args
    output_dir = None
    mode_arg = None
    merge_mode = False  # НОВЫЙ ФЛАГ
    
    for i, arg in enumerate(args):
        if arg == '--output-dir' and i + 1 < len(args):
            output_dir = args[i + 1]
            args = args[:i] + args[i+2:]
            break

    for i, arg in enumerate(args):
        if arg == '--mode' and i + 1 < len(args):
            try:
                mode_arg = int(args[i + 1])
                args = args[:i] + args[i+2:]
                break
            except:
                pass
    
    # НОВЫЙ АРГУМЕНТ --merge
    if '--merge' in args:
        merge_mode = True
        args.remove('--merge')

    if '--from-gui' in args: args.remove('--from-gui')
    if '--batch-mode' in args: args.remove('--batch-mode')

    if mode_arg is None:
        mode_arg = MODE_RELAYS

    if not args:
        print("❌ Ошибка: не указаны пути к файлам!")
        if not is_gui_mode and not batch_mode: 
            input("\nНажмите Enter для выхода...")
        return

    if output_dir:
        output_folder = Path(output_dir)
    else:
        first_file = Path(args[0])
        output_folder = first_file.parent if first_file.parent.exists() else Path(".")
    
    output_folder.mkdir(parents=True, exist_ok=True)

    global_relays = []

    mode_names = {
        MODE_RELAYS: "Реле (KL, K, KLB)"
    }
    
    print(f"🔍 Режим: {mode_names.get(mode_arg, 'Неизвестный')}")
    print(f"📂 Папка вывода: {output_folder}")
    print(f"📁 Количество файлов для обработки: {len(args)}")
    print(f"🔄 Режим объединения: {'Включен' if merge_mode else 'Выключен'}")  # НОВОЕ
    print("=" * 60)

    for idx, pdf_path in enumerate(args, 1):
        pdf_file = Path(pdf_path)
        if not pdf_file.exists():
            print(f"❌ Файл '{pdf_path}' не найден, пропускаем.")
            continue

        shield_name = pdf_file.stem
        print(f"\n📁 [{idx}/{len(args)}] Обработка файла: {pdf_file.name} (Щит: {shield_name})")
        
        diagnose_pdf(str(pdf_file))
        
        print(f"\n📖 ИЗВЛЕЧЕНИЕ ДАННЫХ ИЗ {pdf_file.name}")
        
        # Проверяем кэш перед началом обработки
        from cache_manager import get_cache_manager, init_cache_manager
        cache_manager = init_cache_manager()  # Инициализируем с логированием
        cached_excel = cache_manager.get_cached_excel(pdf_file)
        
        if cached_excel and cached_excel.exists():
            logger.info(f"✅ Найдены закэшированные данные: {cached_excel.name}")
            print(f"✅ Найдены закэшированные данные: {cached_excel.name}")
            temp_excel_path = cached_excel
        else:
            logger.info("Шаг 1: Извлечение текста и таблиц из PDF...")
            text_data = extract_text_from_pdf(str(pdf_file))
            tables_data = extract_tables_from_pdf(str(pdf_file))

            if not text_data and not tables_data:
                logger.warning(f"   ⚠️ Не удалось извлечь данные из {pdf_file.name}")
                print(f"   ⚠️ Не удалось извлечь данные из {pdf_file.name}")
                print("   Возможные причины:")
                print("   1. PDF файл защищен паролем")
                print("   2. PDF содержит только сканированные изображения (нужен OCR)")
                print("   3. PDF не содержит текстовых слоев")
                continue

            print(f"\n💾 СОЗДАНИЕ ВРЕМЕННОГО EXCEL ФАЙЛА")
            
            # Создаем временный файл в той же папке где PDF (будет перемещен в кэш)
            temp_excel_path = pdf_file.parent / f"{pdf_file.stem}_extracted_temp.xlsx"
            logger.info(f"   Сохранение промежуточных данных: {temp_excel_path.name}")
            print(f"   Сохранение промежуточных данных: {temp_excel_path.name}")
            
            try:
                with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as writer:
                    if text_data:
                        df_text = pd.DataFrame(text_data)
                        df_text.to_excel(writer, sheet_name='Текст_из_PDF', index=False)
                        print(f"   ✅ Сохранено {len(text_data)} строк текста")
                    if tables_data:
                        for i, table_info in enumerate(tables_data):
                            sheet_name = f"Таблица_{table_info['Страница']}_{table_info['Таблица']}"[:31]
                            df_table = pd.DataFrame(table_info['Данные'])
                            df_table.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                            print(f"   ✅ Сохранена таблица {i+1}: {len(table_info['Данные'])} строк")
            except Exception as e:
                logger.error(f"❌ Ошибка при создании временного файла: {e}")
                print(f"❌ Ошибка при создании временного файла: {e}")
                continue
            
            if not temp_excel_path.exists():
                logger.error(f"❌ Временный файл не создан!")
                print(f"❌ Временный файл не создан!")
                continue
            else:
                logger.info(f"   ✅ Временный файл создан: {temp_excel_path.name} (размер: {temp_excel_path.stat().st_size} байт)")
                print(f"   ✅ Временный файл создан: {temp_excel_path.name} (размер: {temp_excel_path.stat().st_size} байт)")
            
            # Сохраняем в кэш через save_to_cache (перемещает файл в .pdf_cache)
            temp_excel_path = cache_manager.save_to_cache(pdf_file, temp_excel_path)
        
        print(f"\n🔍 ПОИСК РЕЛЕ В {temp_excel_path.name}")
        file_relays = process_extracted_excel(str(temp_excel_path), mode_arg, shield_name)
        
        if file_relays:
            print(f"   ✅ Найдено {len(file_relays)} реле")
            global_relays.extend(file_relays)
        else:
            print(f"   ⚠️ Реле не найдены!")

    print("\n" + "=" * 60)
    print("🏁 ЗАВЕРШЕНИЕ ОБРАБОТКИ ВСЕХ ФАЙЛОВ")
    print(f"📊 Итого: {len(global_relays)} реле")
    
    if global_relays:
        # ИЗМЕНЕНО: передаем merge_mode в функцию сохранения
        save_aggregated_results(global_relays, str(output_folder), merge_mode)
    else:
        print("⚠️ Нет данных для сохранения!")

    # --- УДАЛЕНИЕ ПРОМЕЖУТОЧНЫХ ФАЙЛОВ ТЕПЕРЬ ОТКЛЮЧЕНО ---
    # Временные файлы теперь хранятся в кэше для повторного использования
    print("\nℹ️ Временные файлы сохранены в кэше для ускорения последующей обработки.")
    print("   Кэш будет очищен при нажатии кнопки 'Очистить' или закрытии программы.")
    # --------------------------------------    
    
    if not is_gui_mode and not batch_mode:
        input("\nНажмите Enter для выхода...")

if __name__ == "__main__":
    main()