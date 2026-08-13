import os
import shutil
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import sys
import traceback
import builtins

def should_stop():
    """Проверяет, запросил ли пользователь остановку через GUI."""
    check = getattr(builtins, '_gui_stop_check', None)
    return check() if check else False

def find_files_with_suffixes(directory, suffixes):
    """Поиск файлов с заданными окончаниями в указанной директории"""
    found_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            # Игнорируем временные файлы Excel (~$)
            if file.startswith('~$'):
                continue
            # Проверяем все суффиксы
            for suffix in suffixes:
                if file.endswith(suffix):
                    found_files.append(os.path.join(root, file))
                    break  # Чтобы не добавлять файл дважды
    return found_files

def extract_data_from_source(source_file):
    """Извлечение данных из исходного файла из диапазона B2 и ниже (только столбцы B, C, D)"""
    try:
        print(f"  Чтение файла: {os.path.basename(source_file)}")
        source_wb = openpyxl.load_workbook(source_file, data_only=True)
        source_ws = source_wb.active
        
        # Проверяем наличие данных в колонке B
        has_any_data = False
        for r in range(2, min(100, source_ws.max_row + 1)):
            cell_value = source_ws.cell(row=r, column=2).value
            if cell_value is not None and str(cell_value).strip():
                has_any_data = True
                break
        
        if not has_any_data:
            print(f"  В файле не найдено данных в колонке B")
            return None
        
        # Собираем данные только из столбцов B, C, D (колонки 2, 3, 4)
        source_data = []
        row = 2
        
        while row <= source_ws.max_row:
            row_data = []
            has_data = False
            
            # Проверяем только столбцы B (2), C (3), D (4)
            for col in range(2, 5):  # Только столбцы 2, 3, 4 (B, C, D)
                cell = source_ws.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    row_data.append(cell_value)
                else:
                    row_data.append('')
            
            # Если есть данные, добавляем строку
            if has_data:
                source_data.append(row_data)
            
            row += 1
            # Безопасный лимит для предотвращения бесконечного цикла
            if row > 10000:
                break
        
        print(f"  Извлечено {len(source_data)} строк с данными (столбцы B, C, D)")
        return source_data
        
    except Exception as e:
        print(f"  Ошибка при чтении файла {source_file}: {e}")
        return None

def merge_data_to_template(template_file, all_data, output_file):
    """Объединение всех данных в шаблон"""
    try:
        print(f"Открываем файл шаблона: {template_file}")
        template_wb = openpyxl.load_workbook(template_file)
        template_ws = template_wb.active
        
        # Подсчитываем общее количество строк
        total_rows = 0
        for data in all_data:
            if data:
                total_rows += len(data)
        
        print(f"Всего строк для вставки: {total_rows}")
        
        # Очищаем целевой диапазон от L4 и ниже
        max_row = template_ws.max_row
        print("Очистка целевого диапазона...")
        for row in range(4, max_row + 1):
            for col in [12, 13, 14]:   # только L, M, N
                cell = template_ws.cell(row=row, column=col)
                # Проверяем, не является ли ячейка частью объединенной области
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell.value = None
        
        # Вставляем данные в целевой файл начиная с L4
        current_row = 4  # Начинаем с строки 4

        for file_index, file_data in enumerate(all_data):
            if not file_data:
                continue
            
            # Вставляем данные из текущего файла
            for i, row_data in enumerate(file_data):
                target_row = current_row + i
                # Вставляем 3 столбца (B, C, D) в колонки L, M, N (12, 13, 14)
                for j, value in enumerate(row_data):
                    target_col = j + 12  # Начинаем с колонки L (12)
                    if value:  # Если значение не пустое
                        target_cell = template_ws.cell(row=target_row, column=target_col)
                        # Проверяем, не является ли ячейка объединенной
                        if not isinstance(target_cell, openpyxl.cell.cell.MergedCell):
                            target_cell.value = value
                        else:
                            # Если ячейка объединена, пробуем найти первую ячейку в объединенной области
                            for merged_range in template_ws.merged_cells.ranges:
                                if target_row in range(merged_range.min_row, merged_range.max_row + 1) and \
                                target_col in range(merged_range.min_col, merged_range.max_col + 1):
                                    # Записываем в левую верхнюю ячейку объединенной области
                                    top_left_cell = template_ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                                    top_left_cell.value = value
                                    break
            
            current_row += len(file_data)
            print(f"  Данные из файла {file_index + 1} добавлены (столбцы B,C,D -> L,M,N)")
        
        # Сохраняем объединенный файл
        template_wb.save(output_file)
        print(f"Объединенные данные сохранены в: {output_file}")
        return True
        
    except Exception as e:
        print(f"Ошибка при объединении данных: {e}")
        traceback.print_exc()
        return False

def main():
    # Получаем путь к текущей директории скрипта
    script_dir = Path(__file__).parent.absolute()
    
    # Определяем папку для поиска файлов и сохранения результата
    # По умолчанию - папка "Результаты" в директории скрипта
    default_output_dir = script_dir / "Результаты"
    
    # Проверяем аргументы командной строки
    if len(sys.argv) >= 2:
        # Если передан путь как аргумент, используем его
        source_directory = sys.argv[1]
        output_directory = Path(source_directory)
    else:
        # Если аргументов нет, используем папку "Результаты" по умолчанию
        source_directory = str(default_output_dir)
        output_directory = default_output_dir
        print(f"Папка не указана, используется папка по умолчанию: {source_directory}")
    
    # Проверяем существование директории
    if not os.path.exists(source_directory):
        print(f"Директория не найдена: {source_directory}")
        print("Создаем папку по умолчанию...")
        output_directory.mkdir(parents=True, exist_ok=True)
    
    print(f"=" * 60)
    print(f"Поиск файлов в директории: {source_directory}")
    print(f"Результат будет сохранен в: {output_directory}")
    print(f"=" * 60)
    
    # Ищем файлы с несколькими вариантами окончаний
    search_suffixes = [
        "Автоматические_выключатели.xlsx",
        "Автоматические выключатели (объединенные).xlsx"
    ]
    
    found_files = find_files_with_suffixes(source_directory, search_suffixes)
    
    if not found_files:
        print(f"Не найдено файлов с указанными окончаниями:")
        for suffix in search_suffixes:
            print(f"  - {suffix}")
        sys.exit(1)
    
    print(f"Найдено файлов: {len(found_files)}")
    for i, file_path in enumerate(found_files, 1):
        print(f"  {i}. {os.path.basename(file_path)}")
    
    # Путь к файлу шаблона
    template_path = script_dir / "Шаблон для автоматов.xlsx"
    
    if not template_path.exists():
        print(f"Файл шаблона не найден: {template_path}")
        sys.exit(1)
    
    print(f"\nФайл шаблона найден: {template_path}")
    
    # Извлекаем данные из всех файлов
    print(f"\n" + "=" * 60)
    print("Извлечение данных из файлов (только столбцы B, C, D):")
    print("=" * 60)
    
    all_data = []
    successful_files = 0
    
    for file_path in found_files:
        data = extract_data_from_source(file_path)
        if data is not None:
            all_data.append(data)
            successful_files += 1
        else:
            all_data.append(None)
    
    print(f"\nУспешно прочитано файлов: {successful_files} из {len(found_files)}")
    
    if successful_files == 0:
        print("Нет данных для объединения")
        sys.exit(1)
    
    # Создаем имя для выходного файла с временной меткой
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"Таблица_автоматов_общая_{timestamp}.xlsx"
    
    # Сохраняем результат в указанную папку
    output_path = output_directory / output_filename
    
    # Создаем папку для результатов, если её нет
    output_directory.mkdir(parents=True, exist_ok=True)
    
    # Создаем резервную копию шаблона (в папке скрипта)
    backup_path = script_dir / "Шаблон для автоматов_backup.xlsx"
    if not backup_path.exists():
        shutil.copy2(template_path, backup_path)
        print(f"Создана резервная копия шаблона: {backup_path}")
    
    # Объединяем данные в шаблон
    print(f"\n" + "=" * 60)
    print("Объединение данных в шаблон (B,C,D -> L,M,N):")
    print("=" * 60)
    
    if merge_data_to_template(template_path, all_data, output_path):
        print(f"\n" + "=" * 60)
        print("УСПЕШНО! Все данные объединены в один файл:")
        print(f"  {output_path}")
        print(f"  Столбцы B,C,D скопированы в L,M,N соответственно")
        print("=" * 60)
        
        # Также сохраняем копию в папку скрипта для обратной совместимости
        #script_output_path = script_dir / output_filename
        #if output_path != script_output_path:
            #shutil.copy2(output_path, script_output_path)
            #print(f"Копия сохранена также в: {script_output_path}")
    else:
        print("\n[ОШИБКА] Не удалось объединить данные")
        # Восстанавливаем шаблон из резервной копии при ошибке
        if backup_path.exists():
            shutil.copy2(backup_path, template_path)
            print("Шаблон восстановлен из резервной копии")
        sys.exit(1)

if __name__ == "__main__":
    main()