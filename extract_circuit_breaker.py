import sys
import os
import re
from openpyxl import load_workbook, Workbook
from pathlib import Path
import builtins

def should_stop():
    """Проверяет, запросил ли пользователь остановку через GUI."""
    check = getattr(builtins, '_gui_stop_check', None)
    return check() if check else False

# Флаг для определения запуска из GUI (без интерактивного ввода)
RUNNING_FROM_GUI = len(sys.argv) > 1 and sys.argv[1] == "--from-gui"
OUTPUT_DIR = None  # Глобальная переменная для папки вывода

def load_nominal_library(library_path):
    """Загружает библиотеку соответствий типов автоматов и номинальных токов"""
    if not os.path.exists(library_path):
        print(f"❌ Ошибка: Файл библиотеки '{library_path}' не найден!")
        print("   Файл Library_nominals.xlsx должен находиться в той же папке, что и скрипт.")
        return None
    
    try:
        wb = load_workbook(library_path, data_only=True)
        sheet = wb.active
        
        col_type = None
        col_nominal = None
        header_row = None
        
        for row in range(1, min(10, sheet.max_row + 1)):
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    cell_clean = cell_value.strip().lower()
                    if "тип автомата" in cell_clean:
                        col_type = col
                        header_row = row
                    elif "номинальный ток" in cell_clean:
                        col_nominal = col
        
        if col_type is None or col_nominal is None:
            print("❌ Ошибка: Не удалось найти заголовки в файле библиотеки!")
            return None
        
        library = {}
        for row in range(header_row + 1, sheet.max_row + 1):
            type_value = sheet.cell(row=row, column=col_type).value
            nominal_value = sheet.cell(row=row, column=col_nominal).value
            
            if type_value and nominal_value:
                type_str = str(type_value).strip()
                try:
                    # ИСПРАВЛЕНИЕ: сохраняем как float, а не int
                    nominal_float = float(nominal_value)
                    library[type_str] = nominal_float
                except (ValueError, TypeError):
                    print(f"⚠️ Предупреждение: Пропущена строка {row}: неверное значение номинала '{nominal_value}'")
        
        print(f"✅ Загружено соответствий из библиотеки: {len(library)}")
        return library
        
    except Exception as e:
        print(f"❌ Ошибка при загрузке библиотеки: {e}")
        return None

def extract_nominal_current_from_library(type_value, library):
    """Извлекает номинальный ток из строки типа автомата из библиотеки"""
    if not type_value or not isinstance(type_value, str):
        return None, "пустое значение"
    
    type_str = type_value.strip()
    
    if library and type_str in library:
        return library[type_str], None
    
    return None, f"тип '{type_str[:50]}' не найден в библиотеке"

def find_column_by_header(sheet, target_headers, check_content_for_q=False):
    """Ищет номер столбца по тексту заголовка и опционально проверяет содержимое на наличие Q"""
    if isinstance(target_headers, str):
        target_headers = [target_headers]
    
    candidate_columns = []  # (col, header_row, sample_values)
    
    for row in range(1, min(20, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                cell_clean = cell_value.strip().lower()
                
                for target in target_headers:
                    target_clean = target.strip().lower()
                    if cell_clean == target_clean or cell_clean.startswith(target_clean):
                        if len(cell_clean) > len(target_clean):
                            next_char = cell_clean[len(target_clean)]
                            if next_char.isalnum() or next_char == '_':
                                continue
                        
                        # Собираем образцы значений из первых 5 непустых строк для проверки
                        samples = []
                        for check_row in range(row + 1, min(row + 20, sheet.max_row + 1)):
                            val = sheet.cell(row=check_row, column=col).value
                            if val and str(val).strip():
                                samples.append(str(val).strip())
                                if len(samples) >= 5:
                                    break
                        
                        candidate_columns.append({
                            "col": col,
                            "header_row": row,
                            "header_text": cell_value,
                            "samples": samples
                        })
                        break
    
    if not candidate_columns:
        return None, None
    
    # Если нужно проверить содержимое на наличие Q
    if check_content_for_q:
        # Ищем столбец, где в образцах есть буква Q (независимо от регистра)
        for candidate in candidate_columns:
            for sample in candidate["samples"]:
                if 'q' in sample.lower():
                    print(f"   ✅ Выбран столбец '{candidate['header_text']}' (содержит Q в значениях: {sample})")
                    return candidate["col"], candidate["header_row"]
        
        # Если ни в одном образце нет Q, но есть кандидаты - берём первый
        print(f"   ⚠️ В значениях не найдена буква Q, берём первый кандидат: '{candidate_columns[0]['header_text']}'")
        return candidate_columns[0]["col"], candidate_columns[0]["header_row"]
    
    # Иначе берём первый найденный
    return candidate_columns[0]["col"], candidate_columns[0]["header_row"]

def find_data_start_row(sheet, header_row, col_fider, col_type):
    """Находит строку, с которой начинаются реальные данные"""
    start_row = header_row + 1
    for row in range(header_row + 1, min(header_row + 50, sheet.max_row + 1)):
        fider_val = sheet.cell(row=row, column=col_fider).value
        type_val = sheet.cell(row=row, column=col_type).value
        if (fider_val and str(fider_val).strip()) or (type_val and str(type_val).strip()):
            return row
    return start_row

def extract_fider_number(fider_str):
    """Извлекает номер фидера из позиционного обозначения"""
    if not fider_str:
        return None
    
    match = re.search(r'[\d\.\-]+$', fider_str)
    if match:
        number_part = match.group()
        number_part = number_part.lstrip('-')
        return number_part
    return None

def is_input_fider(fider_str):
    """Проверяет, является ли фидер вводным"""
    if not fider_str:
        return False
    
    fider_lower = fider_str.strip().lower()
    patterns = [r'^ввод\d*$', r'^ввод\s+\d+$', r'^input\d*$', r'^вв\.?\s*\d*$']
    
    for pattern in patterns:
        if re.search(pattern, fider_lower):
            return True
    return False

def process_xlsx(input_file_path, library):
    """Обработка XLSX файла с таблицей фидеров (только автоматы)"""
    wb = load_workbook(input_file_path, data_only=True)
    sheet = wb.active
    
    print(f"Обработка листа: {sheet.title}")
    print("=" * 50)
    
    col_fider, header_row_fider = find_column_by_header(sheet, [
        "Позиционное обозначение фидера",
        "Позиционное обозначение фидеров по ИД",
        "Позиционное обозначение",
        "Обозначение фидера",
        "Фидер",
        "Поз. обозначение фидера"
    ], check_content_for_q=True)
    
    col_type, header_row_type = find_column_by_header(sheet, [
        "Тип автомата",
        "Тип автомата/коммутационного блока",
        "Тип автомата/ коммутационного блока",
        "Тип выключателя"
    ])
    
    if col_fider is None:
        print("❌ Столбец 'Позиционное обозначение фидера' не найден!")
        return [], [], None, None
    
    if col_type is None:
        print("❌ Столбец 'Тип автомата/коммутационного блока' не найден!")
        return [], [], None, None
    
    header_row = max(header_row_fider, header_row_type)
    data_start_row = find_data_start_row(sheet, header_row, col_fider, col_type)
    
    print(f"✅ Найден столбец фидера: {col_fider} (строка заголовка: {header_row_fider})")
    print(f"✅ Найден столбец типа автомата: {col_type} (строка заголовка: {header_row_type})")
    print(f"📌 Данные начинаются с строки: {data_start_row}")
    print("=" * 50)
    
    results_automats = []
    errors_automats = []
    found_matches = 0
    counter = 1
    
    for row_idx in range(data_start_row, sheet.max_row + 1):
        if should_stop():
            print("⏹️ Остановка по запросу пользователя")
            return results_automats, errors_automats, col_fider, col_type
        fider_value = sheet.cell(row=row_idx, column=col_fider).value
        type_value = sheet.cell(row=row_idx, column=col_type).value
        
        if fider_value is None or str(fider_value).strip() == "":
            continue
        
        fider_str = str(fider_value).strip()
        
        if type_value is None or str(type_value).strip() == "":
            errors_automats.append({
                "row": row_idx,
                "number": counter,
                "fider": fider_str,
                "type": "",
                "error": "тип автомата не указан"
            })
        else:
            type_str = str(type_value).strip()
            nominal_current, error_msg = extract_nominal_current_from_library(type_str, library)
            
            if nominal_current is not None:
                found_matches += 1
            
            results_automats.append({
                "number": counter,
                "fider": fider_str,
                "type": type_str,
                "nominal": nominal_current if nominal_current is not None else error_msg
            })
            
            if error_msg and nominal_current is None:
                errors_automats.append({
                    "row": row_idx,
                    "number": counter,
                    "fider": fider_str,
                    "type": type_str,
                    "error": error_msg
                })
        
        counter += 1
    
    if library:
        print(f"\n📊 Статистика определения токов автоматов:")
        print(f"   - Найдено в библиотеке: {found_matches}")
        print(f"   - Не найдено в библиотеке: {len(errors_automats)}")
    
    return results_automats, errors_automats, col_fider, col_type

def save_automats_to_xlsx(results, errors, output_file_path):
    """Сохраняет результаты по автоматам в XLSX файл"""
    wb = Workbook()
    
    ws_main = wb.active
    ws_main.title = "Номинальные токи"
    ws_main.append(["№ п/п", "Позиционное обозначение фидера", "Тип автомата/коммутационного блока", "Номинальный ток, А"])
    
    for row in results:
        ws_main.append([row["number"], row["fider"], row["type"], row["nominal"]])
    
    ws_main.column_dimensions['A'].width = 8
    ws_main.column_dimensions['B'].width = 25
    ws_main.column_dimensions['C'].width = 50
    ws_main.column_dimensions['D'].width = 20
    
    if errors:
        unique_errors = {}
        for err in errors:
            error_type = err["type"]
            if error_type not in unique_errors:
                unique_errors[error_type] = err
        
        ws_errors = wb.create_sheet("Неопределённые токи")
        ws_errors.append(["Тип автомата/коммутационного блока", "Причина", "Пример фидера"])
        
        for error_type in sorted(unique_errors.keys()):
            err = unique_errors[error_type]
            ws_errors.append([
                err["type"],
                err["error"],
                err["fider"]
            ])
        
        ws_errors.column_dimensions['A'].width = 50
        ws_errors.column_dimensions['B'].width = 40
        ws_errors.column_dimensions['C'].width = 25
        
        print(f"\n⚠️ Найдено уникальных типов автоматов, отсутствующих в библиотеке: {len(unique_errors)}")
    
    wb.save(output_file_path)
    return output_file_path

def wait_for_exit():
    """Ожидает нажатия Enter, только если скрипт запущен не из GUI"""
    if not RUNNING_FROM_GUI:
        input("\nНажмите Enter для выхода...")

def main():
    # Проверяем флаг --from-gui в аргументах
    global RUNNING_FROM_GUI, OUTPUT_DIR
    if "--from-gui" in sys.argv:
        RUNNING_FROM_GUI = True
        # Удаляем флаг из аргументов
        sys.argv = [arg for arg in sys.argv if arg != "--from-gui"]
    
    # Проверяем флаг --output-dir
    output_dir = None
    if "--output-dir" in sys.argv:
        idx = sys.argv.index("--output-dir")
        if idx + 1 < len(sys.argv):
            output_dir = sys.argv[idx + 1]
            # Удаляем оба аргумента
            sys.argv.pop(idx)
            sys.argv.pop(idx)
    
    if output_dir:
        OUTPUT_DIR = Path(output_dir)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    if len(sys.argv) < 2:
        print("Использование: перетащите XLSX файл на этот скрипт")
        print("\nПримечание: Файл библиотеки 'Library_nominals.xlsx' должен находиться в той же папке, что и скрипт")
        wait_for_exit()
        return
    
    input_file_path = sys.argv[1]
    
    if not os.path.exists(input_file_path):
        print(f"Ошибка: Файл '{input_file_path}' не найден!")
        wait_for_exit()
        return
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    library_path = os.path.join(script_dir, "Library_nominals.xlsx")
    
    print(f"\n📁 Обработка файла: {os.path.basename(input_file_path)}")
    print("=" * 50)
    
    library = load_nominal_library(library_path)
    
    if library is None:
        print("\n❌ Невозможно продолжить: библиотека номиналов не загружена!")
        wait_for_exit()
        return
    
    print("=" * 50)
    
    try:
        results_automats, errors_automats, col_fider, col_type = process_xlsx(input_file_path, library)
        
        # Определяем путь для сохранения
        input_path = Path(input_file_path)
        if OUTPUT_DIR:
            output_dir_path = OUTPUT_DIR
        else:
            output_dir_path = input_path.parent
        
        base_name = input_path.stem
        output_automats_path = output_dir_path / f"{base_name}_Автоматические_выключатели.xlsx"
        
        if results_automats:
            save_automats_to_xlsx(results_automats, errors_automats, str(output_automats_path))
            print(f"\n✅ Обработано фидеров (автоматы): {len(results_automats)}")
            print(f"❌ Не удалось определить ток (нет в библиотеке): {len(errors_automats)}")
            print(f"📁 Результат по автоматам сохранён в: {output_automats_path}")
        
            if errors_automats:
                print(f"\n⚠️ Строки автоматов, требующие добавления в библиотеку:")
                
                # Собираем уникальные типы для вывода в формате, понятном веб-интерфейсу
                unique_error_types = {}
                for err in errors_automats:
                    error_type = err['type']
                    if error_type not in unique_error_types:
                        unique_error_types[error_type] = err
                
                # Выводим все уникальные типы в формате, который распознаёт app.py
                for error_type in sorted(unique_error_types.keys()):
                    err = unique_error_types[error_type]
                    # Формат: [тип 'XXX' не найден в библиотеке]
                    print(f"   - [тип '{err['type']}' не найден в библиотеке]")
                
                # Для информации показываем статистику
                print(f"\n📊 Всего уникальных типов, отсутствующих в библиотеке: {len(unique_error_types)}")
                print(f"   Общее количество строк с ошибками: {len(errors_automats)}")
                
                # Показываем первые 5 для наглядности (опционально)
                if len(errors_automats) > 0:
                    print(f"\n   Примеры (первые 5 из {len(errors_automats)}):")
                    for err in errors_automats[:5]:
                        print(f"      №{err.get('number', '?')}: {err['fider']} → {err['type'][:40]}")
    
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "=" * 50)
    wait_for_exit()

if __name__ == "__main__":
    main()